"""
automation.adapters.statssa — Statistics South Africa adapter.

Responsible datasets
--------------------
  - unemployment       (QLFS P0211 — part of the QLFS family)
  - youth-unemployment (QLFS P0211 — same release)
  - labour-force       (QLFS P0211 — same release)
  - gdp                (GDP P0441)
  - inflation          (CPI component only, P0141)
  - population         (MYPE P0302)
  - housing            (GHS component only, P0318)
  - census             (Census 2022 — erratum watch only)
  - municipalities     (Census 2022 Municipal Fact Sheet — erratum watch only)

Design principles (from architecture doc)
------------------------------------------
- The QLFS family (unemployment, youth-unemployment, labour-force) is ONE
  release.  ``check_for_updates`` returns the same result for all three.
- Census and municipalities are static — their check is a lightweight
  page-hash watch, not a download attempt.
- Population MUST use Stats SA P0302, not World Bank.  The source guard
  will be enforced in the download step (Phase B).
- Retry policy: STATSSA_POLICY (exponential backoff, up to 2 h — the
  release-day site-load scenario).

Phase 1 scope (QLFS) — download + archive
------------------------------------------
This upgrade implemented:
  - Real ETag/content-hash change detection on the P0211 release hub
  - Discover → download → archive the raw QLFS publication file

Phase 2 scope (QLFS) — parse + transform + stage (this build)
----------------------------------------------------------------
Continuing directly from where Phase 1 stopped, ``fetch_and_apply()`` now
also:
  1. Parses the archived Excel workbook (``parse_qlfs_workbook()``) by
     header/label matching — not fixed cell coordinates — since Stats SA's
     table layout is a per-release idiom, not a stable contract (see
     ``IMPLEMENTATION-SPEC-STATSSA-PHASE2.md`` §5). If the expected tables
     cannot be located, this raises loudly; there is no PDF-scraping or
     stale-value fallback.
  2. Transforms the parsed values into the three existing JSON schemas
     (``unemployment.json``, ``youth-unemployment.json``,
     ``labour-force.json``) via ``_transform_unemployment()`` /
     ``_transform_youth_unemployment()`` / ``_transform_labour_force()``,
     each following the exact deep-copy / rate-bearing-fields-only /
     seed-or-append-series pattern already established by
     ``SARBAdapter._transform_interest_rates()``.
  3. Validates each candidate document: rate bounds, quarterly label
     format, ``check_protected_fields()`` (reused unchanged from
     ``core/metadata.py``), and a quarter-over-quarter plausibility check
     that flags (does not hard-fail) large jumps for the human reviewer.
  4. Stages each of the three documents that actually changed via
     ``automation.core.staging.write_staged_dataset()`` and records one
     ``pending`` version entry per dataset (three total) — the same
     staging → approval → promote gate already enforced for
     ``interest-rates.json``. **No direct write to any dataset JSON ever
     happens here.**

Design note (open question resolved, see spec §9): this adapter records
**one version entry per output dataset** (three total per QLFS release),
not one shared entry for all three. ``version.py``'s store and
``promote_version()`` are both keyed by a single ``dataset_id`` per call,
so one entry per dataset is the natural fit and keeps `--approve`/
`--promote` invocations for one QLFS output independent of the other two
(e.g. a reviewer can approve ``unemployment`` while still checking
``labour-force``). No change was made to ``version.py`` to support this —
it is simply called three times.

Phase 3a scope (GDP) — parse + transform + stage (this build)
-----------------------------------------------------------------
Continuing from where Phase 2 left QLFS, ``fetch_and_apply()`` now also
processes GDP (P0441) as a second, fully independent flow within the same
method call (see ``IMPLEMENTATION-SPEC-GDP.md``):
  1. Real ETag/content-hash detection against the P0441 release hub
     (``_check_gdp()``, mirroring ``_check_qlfs()``).
  2. Discover → download → archive the raw GDP Excel publication.
  3. Parse the workbook (``parse_gdp_workbook()``) by header/label
     matching, extracting **every** available quarter-column value in
     the GDP growth table — not just the latest — because Stats SA
     routinely revises previously published quarters (``gdp.yaml``'s
     ``overwrites_historical_points: true``).
  4. Transform (``_transform_gdp()`` / ``_apply_gdp_growth_points()``)
     into ``gdp.json``'s existing ``gdp-growth`` stat only, overwriting
     any revised historical series point in place and appending any
     genuinely new one; ``gdp-annual-growth``, ``gdp-nominal``, and
     ``gdp-per-capita`` are never read or modified by this flow.
  5. Validate (plausibility range for growth rates — which, unlike QLFS
     rates, can be negative — quarterly label format, protected-field
     diff, and a GDP-specific quarter-over-quarter anomaly threshold)
     and stage the candidate via the same
     ``core/staging.py``/``core/version.py`` pipeline already proven for
     SARB and QLFS. No direct write to ``gdp.json`` ever happens here.

Phase 3b scope (CPI) — parse + transform + stage (this build)
-----------------------------------------------------------------
Continuing from where Phase 3a left GDP, ``fetch_and_apply()`` now also
processes CPI (P0141) as a third, fully independent flow within the same
method call (see ``IMPLEMENTATION-SPEC-CPI.md``):
  1. Real ETag/content-hash detection against the P0141 release hub
     (``_check_cpi()``, mirroring ``_check_qlfs()``/``_check_gdp()``).
  2. Discover → download → archive the raw CPI Excel publication.
  3. Parse the workbook (``parse_cpi_workbook()``) by header/label
     matching, extracting only the **latest month's** value per metric —
     unlike GDP, CPI does not routinely revise previously published
     months, so a single-column read (mirroring
     ``parse_qlfs_workbook()``'s approach, not ``parse_gdp_workbook()``'s)
     is sufficient (assumption flagged for confirmation, see below).
  4. Transform (``_transform_inflation()``, reusing ``_apply_qlfs_rate_map()``
     unchanged — see §12.1 of the spec) into ``inflation.json``'s
     ``cpi-headline`` and ``food-inflation`` stats only; ``repo-rate``
     (SARB-owned) and ``annual-cpi-avg`` (deferred, see §0.2) are never
     read or modified by this flow.
  5. Validate (a genuinely new plausibility range tolerating negative/
     deflationary values, monthly label format, protected-field diff via
     the reused ``check_protected_fields()``, and a CPI-specific
     month-over-month anomaly threshold) **and** enforce, as a dedicated
     hard-fail check distinct from ``check_protected_fields()``, the
     ownership boundary against ``repo-rate``/``annual-cpi-avg``
     (``_assert_cpi_ownership_boundary()`` — deep-compares every
     non-owned stat between the previous and proposed document and
     hard-fails staging on any difference, including a stat silently
     added or removed). Then stage the candidate via the same
     ``core/staging.py``/``core/version.py`` pipeline already proven for
     SARB, QLFS, and GDP. No direct write to ``inflation.json`` ever
     happens here.

The ``repo-rate``/``repo-rate-sarb`` de-duplication and ``annual-cpi-avg``
automation are explicitly deferred to their own follow-on milestones (see
``IMPLEMENTATION-SPEC-CPI.md`` §0.1 and §0.2) — this build's blast radius
is limited to ``cpi-headline`` and ``food-inflation``.

Housing, census, and municipalities remain Phase A stubs — explicitly out
of scope for this build (see ``IMPLEMENTATION-SPEC-CPI.md`` §1). Population
was a Phase A stub as of CPI Phase 3b; it is now implemented — see "Phase 4
scope (Population)" below.

Phase 4 scope (Population) — parse + transform + stage (this build)
-----------------------------------------------------------------------
Continuing from where Phase 3b left CPI, ``fetch_and_apply()`` now also
processes Population (MYPE, P0302) as a fourth, fully independent flow
within the same method call (see ``IMPLEMENTATION-SPEC-POPULATION.md``).
Unlike QLFS/GDP/CPI, this is not a green-field automation — it is a
data-integrity fix shaped like one: the current ``population.json``
``population-total`` value is suspected to be sourced from World Bank,
not Stats SA (see ``SA-Data-Hub-Dataset-Sourcing-Plan.md`` §9), so this
milestone's ``fetch_and_apply()`` flow must make that class of error
structurally impossible going forward:
  1. Real ETag/content-hash detection against the P0302 release hub
     (``_check_population()``, mirroring ``_check_qlfs()``/``_check_gdp()``/
     ``_check_cpi()``).
  2. Discover → download → archive the raw MYPE Excel publication.
  3. Parse the workbook (``parse_population_workbook()``) by header/label
     matching, extracting only the **latest year's** total-population
     value — MYPE has one release per year, so a single-column read
     (mirroring ``parse_qlfs_workbook()``'s/``parse_cpi_workbook()``'s
     approach) is sufficient.
  4. **Enforce Stats SA provenance** (``_assert_population_source_guard()``)
     as a hard-fail gate, checked twice — immediately after discovery
     (URL-level, before download) and again immediately before staging
     (extract-level) — deliberately redundant rather than relying on a
     single check. This is the genuinely new complication this milestone
     introduces and the direct enforcement of ``population.yaml``'s
     ``source_guard_required: true`` / ``source_guard_domain:
     "statssa.gov.za"``.
  5. Transform (``_transform_population()`` / ``_apply_population_total_point()``,
     a new, small, single-purpose helper — **not** a reuse of
     ``_apply_qlfs_rate_map()``, since ``population-total``'s display shape
     is materially different: bare-year series labels and a millions
     magnitude with a dual millions/raw-headcount on-disk convention, see
     §8.2 of the spec) into ``population.json``'s ``population-total`` stat
     only; ``population-urban`` and ``population-median-age`` are never
     read or modified by this flow.
  6. Validate (a wide plausibility range for South Africa's slow-moving
     population, annual label format, protected-field diff via the reused
     ``check_protected_fields()``, and a Population-specific
     year-over-year anomaly threshold expected to fire on the one-time
     production correction run — see §9/§14 item 7 of the spec) **and**
     enforce, as a dedicated hard-fail check distinct from
     ``check_protected_fields()``, the ownership boundary against
     ``population-urban``/``population-median-age``
     (``_assert_population_ownership_boundary()`` — a direct structural
     copy of ``_assert_cpi_ownership_boundary()``). Then stage the
     candidate via the same ``core/staging.py``/``core/version.py``
     pipeline already proven for SARB, QLFS, GDP, and CPI. No direct write
     to ``population.json`` ever happens here.

Housing, census, and municipalities remain Phase A stubs — explicitly out
of scope for this build (see ``IMPLEMENTATION-SPEC-POPULATION.md`` §3).

Population Excel layout — verification status (read before touching the parser)
-----------------------------------------------------------------------------------
Exactly as with QLFS, GDP, and CPI, no archived MYPE ``.xlsx`` file was
available in this implementation session to inspect (no session to date
has had network access to ``statssa.gov.za``). ``parse_population_workbook()``
and ``_POPULATION_METRIC_SPECS``'s label-matching rules were built against
the *documented* Stats SA convention (a header row of bare 4-digit years,
e.g. ``2026``, with a "total population" indicator row identified by label
text), not empirically verified against a real P0302 release file — only
against synthetic fixtures (see ``automation/adapters/tests/test_statss.py``).
The P0302 URL-naming convention used by ``_build_population_candidate_urls()``
is likewise unconfirmed, carried forward the same way the QLFS/GDP/CPI URL
conventions were at the start of their own milestones. The
raw-headcount-vs-millions representation in the source workbook is also
unconfirmed — ``parse_population_workbook()`` applies a documented,
single heuristic (values over 1000 are treated as a raw headcount and
divided by 1,000,000) and logs which branch it took, rather than guessing
silently. The numeric judgement calls ``_POPULATION_PLAUSIBLE_RANGE =
(40.0, 90.0)`` and ``_POPULATION_JUMP_WARNING_THRESHOLD = 2.0`` are this
document's own assumptions, not sourced from ``dataset-analysis.md`` or
the sourcing plan — flagged for stakeholder confirmation. This is
mitigated by design (fail loudly, no guessing, no PDF fallback), not
resolved by observation — the first live run against a real downloaded
MYPE workbook is the actual empirical test of this parser, and is also
the vehicle for the one-time production correction described in
``IMPLEMENTATION-SPEC-POPULATION.md`` §9.

CPI Excel layout — verification status (read before touching the parser)
--------------------------------------------------------------------------
Exactly as with QLFS and GDP, no archived CPI ``.xlsx`` file was available
in this implementation session to inspect (no session to date has had
network access to ``statssa.gov.za``). ``parse_cpi_workbook()`` and
``_CPI_METRIC_SPECS``'s label-matching rules were built against the
*documented* Stats SA convention (a header row of month-year labels, e.g.
``May 2026``, with "All items" and "Food" indicator rows identified by
label text), not empirically verified against a real P0141 release file —
only against synthetic fixtures (see
``automation/adapters/tests/test_statss.py``). The P0141 URL-naming
convention used by ``_build_cpi_candidate_urls()`` is likewise
unconfirmed, carried forward the same way the QLFS and GDP URL
conventions were at the start of their own milestones. The numeric
judgement calls ``_CPI_PLAUSIBLE_RANGE = (-5.0, 30.0)`` and
``_CPI_JUMP_WARNING_THRESHOLD = 1.5`` are this document's own assumptions,
not sourced from ``dataset-analysis.md`` or the sourcing plan — flagged
for stakeholder confirmation. This is mitigated by design (fail loudly,
no guessing, no PDF fallback), not resolved by observation — the first
live run against a real downloaded CPI workbook is the actual empirical
test of this parser.

GDP Excel layout — verification status (read before touching the parser)
--------------------------------------------------------------------------
Exactly as with QLFS in Phase 2, no archived GDP ``.xlsx`` file was
available in this implementation session to inspect (no session to date
has had network access to ``statssa.gov.za``). ``parse_gdp_workbook()``
and ``_GDP_GROWTH_SPEC``'s label-matching rules were built against the
*documented* Stats SA convention (a header row of quarter labels with a
"GDP growth" indicator row beneath it), not empirically verified against
a real P0441 release file — only against synthetic fixtures (see
``automation/adapters/tests/test_statss.py``). The P0441 URL-naming
convention used by ``_build_gdp_candidate_urls()`` is likewise
unconfirmed, carried forward the same way the QLFS URL convention was at
the start of Phase 2. This is mitigated by design (fail loudly, no
guessing, no PDF fallback), not resolved by observation — the first live
run against a real downloaded GDP workbook is the actual empirical test
of this parser.

Excel layout — verification status (read before touching the parser)
----------------------------------------------------------------------
No archived QLFS ``.xlsx`` file was available in this implementation
session to inspect (no session to date has had network access to
``statssa.gov.za`` — see the WAF finding below — and the reports/archive
tree contains no prior Phase 1 download to fall back on). The parser
below was therefore built against the *documented* Stats SA convention
(a header row of quarter labels, e.g. ``Q1 2026``, with indicator rows
identified by label text below it) using header/label matching rather
than fixed coordinates specifically so it can tolerate the layout drift
the sourcing plan warns about — but this has **not** been empirically
validated against a real release file, only against synthetic fixtures
(see ``automation/adapters/tests/test_statss.py``). This is the same
class of open item as the WAF-hash-determinism question below: mitigated
by design (fail loudly, label-based lookup, no guessing), not resolved by
observation. The first live run against a real downloaded workbook should
be treated as the empirical test of this parser, and any mismatch should
update this note rather than silently patching around it.

QLFS release hub URL
--------------------
  https://www.statssa.gov.za/?page_id=1854&PPN=P0211

Excel link pattern (observed, may change per release)
-----------------------------------------------------
  The release hub links to one or more Excel data table files alongside
  the PDF statistical release.  URL patterns observed:
    - Direct xlsx: statssa.gov.za/publications/P0211/...
    - SuperWEB2 time-series: statsssa.gov.za/publications/... (varies)

  We scan the release hub HTML for the first .xlsx or .xls href that
  belongs to statssa.gov.za and contains "P0211" in the path, falling back
  to any .xlsx link on the page.
"""

from __future__ import annotations

import copy
import json
import re
import urllib.parse
from dataclasses import dataclass
from datetime import date, datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any, Callable

import openpyxl

from automation.adapters import register
from automation.adapters.base import BaseAdapter, DatasetCheckResult
from automation.core.config import AutomationConfig, DatasetConfig, SourceConfig
from automation.core.files import portable_archive_path, save_to_archive
from automation.core.http_client import AutomationHTTPError, HTTPClient
from automation.core.logging import get_logger
from automation.core.metadata import check_protected_fields
from automation.core.retry import STATSSA_POLICY, WATCH_POLICY, with_retry
from automation.core.staging import write_staged_dataset
from automation.core.version import new_version_entry, save_version_entry

log = get_logger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_STATSSA_BASE = "https://www.statssa.gov.za"
_RELEASE_HUB_BASE = "https://www.statssa.gov.za/?page_id=1854"

# QLFS (P0211) release hub — used for ETag/content-hash change detection.
# Note: this page is protected by Incapsula WAF and requires a browser
# session to render fully.  We use the raw response hash to detect page
# changes (the redirect/challenge page itself changes when content changes).
_QLFS_HUB_URL = f"{_RELEASE_HUB_BASE}&PPN=P0211"
_QLFS_PUBLICATION_CODE = "P0211"

# Direct publication base URL (confirmed accessible without WAF).
# Individual files at this path can be fetched directly.
_QLFS_PUBLICATION_BASE = "https://www.statssa.gov.za/publications/P0211/"

# Ordinal suffixes used in Stats SA URL naming convention (confirmed 2026-07-12)
# Pattern: "Presentation QLFS QN YYYY.pdf" → "Presentation%20QLFS%20QN%20YYYY.pdf"
_QUARTER_ORDINALS: dict[int, str] = {
    1: "1st",
    2: "2nd",
    3: "3rd",
    4: "4th",
}

# Quarter to month name (for release window context)
_QUARTER_MONTH_NAMES: dict[int, str] = {
    1: "March",
    2: "June",
    3: "September",
    4: "December",
}

# Excel file discovery patterns (for HTML scraping fallback, if page is accessible)
_EXCEL_HREF_PATTERNS: list[re.Pattern[str]] = [
    # Full URL with P0211 in path
    re.compile(r'href=["\']([^"\']*P0211[^"\']*\.xlsx?)["\']', re.IGNORECASE),
    # Any statssa.gov.za xlsx link
    re.compile(r'href=["\']([^"\']*statssa\.gov\.za[^"\']*\.xlsx?)["\']', re.IGNORECASE),
    # Any relative xlsx link
    re.compile(r'href=["\']([^"\']*\.xlsx?)["\']', re.IGNORECASE),
]

# Known quarter labels for QLFS — used to identify the release period from text
_QUARTER_PATTERN = re.compile(
    r"Q([1-4])\s+(\d{4})|Quarter\s+([1-4]).*?(\d{4})", re.IGNORECASE
)

# ---------------------------------------------------------------------------
# QLFS release calendar reference (from SA-Data-Hub-Automation-Architecture.md)
# Each quarter-end → approximate release window
# ---------------------------------------------------------------------------

_QLFS_RELEASE_WINDOWS: dict[str, str] = {
    "Q1": "May–June (6 weeks after 31 March)",
    "Q2": "August (6 weeks after 30 June)",
    "Q3": "November (6 weeks after 30 September)",
    "Q4": "February (6 weeks after 31 December)",
}

_GDP_RELEASE_WINDOWS: dict[str, str] = {
    "Q1": "June (~65–70 days after 31 March)",
    "Q2": "September",
    "Q3": "December",
    "Q4": "March",
}

# GDP (P0441) release hub — used for ETag/content-hash change detection.
# Same WAF caveats as the QLFS hub (see _QLFS_HUB_URL above).
_GDP_HUB_URL = f"{_RELEASE_HUB_BASE}&PPN=P0441"
_GDP_PUBLICATION_CODE = "P0441"

# Direct publication base URL — mirrors _QLFS_PUBLICATION_BASE's pattern.
# **Unconfirmed** against a real Stats SA release (see module docstring's
# "GDP Excel layout — verification status" section).
_GDP_PUBLICATION_BASE = "https://www.statssa.gov.za/publications/P0441/"

# ---------------------------------------------------------------------------
# Datasets managed by this adapter
# ---------------------------------------------------------------------------

_STATSSA_DATASETS: list[str] = [
    "unemployment",
    "youth-unemployment",
    "labour-force",
    "gdp",
    "inflation",     # CPI component only
    "population",
    "housing",       # GHS component only
    "census",
    "municipalities",
]

# Datasets that are effectively static — light erratum-watch only
_STATIC_DATASETS: frozenset[str] = frozenset({"census", "municipalities"})

# QLFS family — all three are one release
_QLFS_FAMILY: frozenset[str] = frozenset({
    "unemployment",
    "youth-unemployment",
    "labour-force",
})

# Path to the canonical dataset JSON files (relative to project root,
# resolved at run time) — mirrors SARBAdapter's _DATASETS_DIR/_INTEREST_RATES_JSON.
_DATASETS_DIR = Path(__file__).resolve().parent.parent.parent / "src" / "data" / "datasets"
_QLFS_DATASET_JSON: dict[str, Path] = {
    "unemployment": _DATASETS_DIR / "unemployment.json",
    "youth-unemployment": _DATASETS_DIR / "youth-unemployment.json",
    "labour-force": _DATASETS_DIR / "labour-force.json",
}

# gdp.json — the sole dataset JSON touched by the GDP flow.
_GDP_DATASET_JSON: Path = _DATASETS_DIR / "gdp.json"
_GDP_GROWTH_STAT_ID = "gdp-growth"

# Quarter-over-quarter jump beyond this many percentage points is flagged
# as an anomaly for the human reviewer (not a hard failure — see
# IMPLEMENTATION-SPEC-STATSSA-PHASE2.md §6, item 3).
_QOQ_JUMP_WARNING_THRESHOLD = 3.0

# GDP growth is inherently more volatile quarter-to-quarter than QLFS
# rates, so its anomaly threshold is wider (IMPLEMENTATION-SPEC-GDP.md §6
# item 4) — a tighter QLFS-style threshold would flag routine, non-
# anomalous revisions.
_GDP_GROWTH_JUMP_WARNING_THRESHOLD = 5.0

# GDP growth is not a [0, 100] percentage like QLFS rates — it can be
# negative (gdp.json already contains -6.2% for 2020) — so this is a
# genuinely new plausibility range, not a reuse of _validate_percentage()'s
# [0, 100] assumption (IMPLEMENTATION-SPEC-GDP.md §6 item 1).
_GDP_GROWTH_PLAUSIBLE_RANGE: tuple[float, float] = (-20.0, 20.0)

# CPI (P0141) release hub — used for ETag/content-hash change detection.
# Same WAF caveats as the QLFS/GDP hubs above.
_CPI_HUB_URL = f"{_RELEASE_HUB_BASE}&PPN=P0141"
_CPI_PUBLICATION_CODE = "P0141"

# Direct publication base URL — mirrors _QLFS_PUBLICATION_BASE's /
# _GDP_PUBLICATION_BASE's pattern. **Unconfirmed** against a real Stats SA
# release — see the module docstring's "CPI Excel layout — verification
# status" section.
_CPI_PUBLICATION_BASE = "https://www.statssa.gov.za/publications/P0141/"

# inflation.json — the sole dataset JSON touched by the CPI flow. Shared
# with the (untouched) SARB-owned repo-rate stat and the (deferred)
# annual-cpi-avg stat — see IMPLEMENTATION-SPEC-CPI.md §0.1/§0.2/§7.
_CPI_DATASET_JSON: Path = _DATASETS_DIR / "inflation.json"
_CPI_HEADLINE_STAT_ID = "cpi-headline"
_CPI_FOOD_STAT_ID = "food-inflation"

# The complete set of stat IDs this milestone's code is permitted to
# read/write inside inflation.json. Used both to build
# _transform_inflation()'s rate_map and, more importantly, as the boundary
# _assert_cpi_ownership_boundary() enforces against every OTHER stat in
# the same file (repo-rate, annual-cpi-avg) — see IMPLEMENTATION-SPEC-
# CPI.md §7 and §11 item 5.
_CPI_OWNED_STAT_IDS: frozenset[str] = frozenset(
    {_CPI_HEADLINE_STAT_ID, _CPI_FOOD_STAT_ID}
)

# CPI is year-on-year % change and, unlike QLFS's [0, 100] rates, can in
# principle go negative (deflation) — this document's own judgement call,
# not sourced from dataset-analysis.md or the sourcing plan
# (IMPLEMENTATION-SPEC-CPI.md §11 item 1 / §17 assumption 2). Flagged for
# stakeholder confirmation before this is treated as final.
_CPI_PLAUSIBLE_RANGE: tuple[float, float] = (-5.0, 30.0)

# Narrower than GDP's 5.0 and QLFS's 3.0 default: inflation.json's own
# historical CPI series moves by well under 1pp month-to-month in almost
# every observed case (the +0.9pp April 2026 jump already on file being
# the one exception) — a judgement call, not sourced from the uploaded
# documentation (IMPLEMENTATION-SPEC-CPI.md §11 item 4 / §17 assumption 2).
_CPI_JUMP_WARNING_THRESHOLD = 1.5

# Matches dataset-analysis.md's documented monthly_label rule verbatim.
_MONTHLY_LABEL_RE = re.compile(r"^[A-Z][a-z]{2} \d{4}$")

# CPI month-header cells, e.g. "May 2026" or "May-2026" / "May. 2026".
_MONTH_HEADER_PATTERN = re.compile(
    r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\.?\s+(\d{4})",
    re.IGNORECASE,
)

# Three-letter month name normalisation (case-insensitive input -> the
# canonical "Mon" form already used by inflation.json's series labels).
_MONTH_ABBR_TO_NUM: dict[str, int] = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}
_MONTH_NUM_TO_ABBR: dict[int, str] = {
    num: abbr.capitalize() for abbr, num in _MONTH_ABBR_TO_NUM.items()
}

# Population (MYPE, P0302) release hub — used for ETag/content-hash change
# detection. Same WAF caveats as the QLFS/GDP/CPI hubs above.
_POPULATION_HUB_URL = f"{_RELEASE_HUB_BASE}&PPN=P0302"
_POPULATION_PUBLICATION_CODE = "P0302"

# Direct publication base URL — mirrors _QLFS_PUBLICATION_BASE's /
# _GDP_PUBLICATION_BASE's / _CPI_PUBLICATION_BASE's pattern. **Unconfirmed**
# against a real Stats SA release — see the module docstring's "Population
# Excel layout — verification status" section.
_POPULATION_PUBLICATION_BASE = "https://www.statssa.gov.za/publications/P0302/"

# population.json — the sole dataset JSON touched by the Population flow.
# Shared with the (untouched) Census-2022-sourced population-urban and
# population-median-age stats — see IMPLEMENTATION-SPEC-POPULATION.md
# §0.1/§3/§7.4.
_POPULATION_DATASET_JSON: Path = _DATASETS_DIR / "population.json"
_POPULATION_TOTAL_STAT_ID = "population-total"

# The complete set of stat IDs this milestone's code is permitted to
# read/write inside population.json. Used both by _transform_population()
# and, more importantly, as the boundary
# _assert_population_ownership_boundary() enforces against every OTHER stat
# in the same file (population-urban, population-median-age) — see
# IMPLEMENTATION-SPEC-POPULATION.md §7.4 and §11.
_POPULATION_OWNED_STAT_IDS: frozenset[str] = frozenset({_POPULATION_TOTAL_STAT_ID})

# South Africa's population is well-documented and slow-moving, so this
# band is deliberately wide — this document's own judgement call, not
# sourced from dataset-analysis.md or the sourcing plan
# (IMPLEMENTATION-SPEC-POPULATION.md §10 / §14 item 3). Flagged for
# stakeholder confirmation before this is treated as final.
_POPULATION_PLAUSIBLE_RANGE: tuple[float, float] = (40.0, 90.0)

# A healthy annual population change is typically 1-2% — this document's
# own judgement call, not sourced from dataset-analysis.md or the sourcing
# plan (IMPLEMENTATION-SPEC-POPULATION.md §10 / §14 item 3). Flagged for
# stakeholder confirmation. Deliberately expected to fire on the one-time
# production correction run (§9/§14 item 7) — that is not a parser bug.
_POPULATION_JUMP_WARNING_THRESHOLD = 2.0

# Matches population.json's existing bare-year series label convention
# (e.g. "2024"), not QLFS's quarterly or CPI's monthly label shape.
_ANNUAL_LABEL_RE = re.compile(r"^\d{4}$")

# The direct enforcement mechanism of population.yaml's
# source_guard_required / source_guard_domain — see
# _assert_population_source_guard() and IMPLEMENTATION-SPEC-POPULATION.md
# §7.3. This is the load-bearing constant of this milestone: it is what
# makes the World-Bank-sourced data-integrity bug
# (SA-Data-Hub-Dataset-Sourcing-Plan.md §9) structurally impossible to
# reintroduce via this flow.
_POPULATION_SOURCE_GUARD_DOMAIN = "statssa.gov.za"

# ---------------------------------------------------------------------------
# HTTP client helpers
# ---------------------------------------------------------------------------

# Tier 1 header hardening (IMPLEMENTATION-SPEC-STATSSA-WAF.md §6.1 step 1,
# approach B). Applied only to Stats SA requests via _build_http_client()
# below — no other adapter's HTTP client is affected.
#
# Accept-Encoding is pinned to "identity", NOT a real browser's default
# "gzip, deflate, br": core/http_client.py (out of scope to change here)
# never decompresses a response body, so advertising compression support
# without the ability to decompress would silently corrupt every hub-page
# fetch this adapter relies on for both WAF-marker detection and release-
# period parsing.
_STATSSA_BROWSER_HEADERS: dict[str, str] = {
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-ZA,en;q=0.9",
    "Accept-Encoding": "identity",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
    "Sec-Fetch-User": "?1",
    "Upgrade-Insecure-Requests": "1",
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"
    ),
}


def _build_http_client(source_config: SourceConfig) -> HTTPClient:
    """Create an HTTPClient configured from source_config.

    Tier 1 header hardening (IMPLEMENTATION-SPEC-STATSSA-WAF.md §6.1 step 1)
    ------------------------------------------------------------------------
    Uses an ordinary-browser-equivalent header set — ``Accept-Language``,
    ``Sec-Fetch-*``, and a non-"bot"-labelled ``User-Agent`` — in place of
    the previous, self-identifying
    ``SA-Data-Hub-Automation/0.1 (...; data-automation-bot)`` User-Agent, to
    test whether Incapsula's block is header/UA-driven rather than purely a
    JS/cookie challenge. This changes only the outbound request headers for
    Stats SA calls; the WAF-marker detection logic downstream (§2.2 of the
    spec) is unchanged — if a challenge page is still served, it is still
    detected and still raised as ``WAF_BLOCKED``, never silently accepted
    as a real page.

    ``Accept-Encoding`` is deliberately pinned to ``"identity"`` rather than
    a real browser's ``"gzip, deflate, br"``: ``core/http_client.py`` has no
    response-decompression support and is out of scope to change here
    (spec §7) — advertising compression support without being able to
    decompress the response would corrupt the raw body text this adapter
    both WAF-scans and (via ``_extract_release_period()``) parses.
    """
    return HTTPClient(
        timeout_seconds=source_config.timeout_seconds,
        extra_headers=dict(_STATSSA_BROWSER_HEADERS),
    )


# ---------------------------------------------------------------------------
# Release hub page parsing
# ---------------------------------------------------------------------------


def _fetch_release_hub_html(client: HTTPClient, hub_url: str) -> bytes:
    """
    Fetch the Stats SA release hub page for a given publication.

    Returns the raw HTML bytes. Explicitly checks for Incapsula WAF challenge
    and raises an error if detected.

    Work Item 5 (2026-07-12 implementation spec) finding
    -----------------------------------------------------
    The original open question was whether the WAF challenge page's content
    hash is stable across requests/client-states — if not, hash-based
    "no_change" detection could misreport a WAF block as a genuine release.
    This has NOT been empirically settled: no environment with network
    access to statssa.gov.za was available in any implementation session to
    date, so no dated, request-counted observation exists (the spec's
    acceptance criterion for this item is therefore still open).

    Rather than assume determinism either way, the fix taken here sidesteps
    the question instead of answering it: the response body is scanned for
    the literal Incapsula markers on every fetch, and any match is raised as
    an explicit ``WAF_BLOCKED`` error rather than being hashed and compared
    at all. A WAF challenge page can therefore never be misread as
    "no_change" or as a genuine release, regardless of whether it happens
    to be deterministic. If a future session gets real network access to
    verify determinism directly, that empirical finding should still be
    recorded here for completeness, but it is no longer load-bearing for
    correctness given this guard.

    Raises
    ------
    AutomationHTTPError
        On non-retryable HTTP errors or WAF blocks.
    urllib.error.URLError
        On transient network errors (retried by caller).
    """
    response = client.get(hub_url)
    if not response.body:
        raise ValueError(f"Empty response body from {hub_url}")
        
    body_text = response.body.decode("utf-8", errors="replace")
    if "_Incapsula_Resource" in body_text or "incapsula" in body_text.lower():
        raise AutomationHTTPError(hub_url, 403, "WAF_BLOCKED: Incapsula WAF challenge detected")

    return response.body


def _extract_excel_url(html: bytes, base_url: str = _STATSSA_BASE) -> str | None:
    """
    Scan HTML for a Stats SA Excel workbook link (HTML scraping fallback).

    Tries each pattern in ``_EXCEL_HREF_PATTERNS`` in order, returning the
    first match that resolves to an absolute HTTPS URL.

    This is a fallback for when the release hub page is accessible (not
    WAF-blocked).  The primary discovery strategy is
    ``_probe_qlfs_excel_url()`` which constructs direct URLs.

    Parameters
    ----------
    html:
        Raw HTML bytes of the release hub page.
    base_url:
        Base for resolving relative hrefs.

    Returns
    -------
    str or None
        Absolute URL of the Excel workbook, or None if none found.
    """
    text = html.decode("utf-8", errors="replace")

    for pattern in _EXCEL_HREF_PATTERNS:
        matches = pattern.findall(text)
        for href in matches:
            href = href.strip()
            if not href:
                continue
            # Resolve relative URLs
            if href.startswith("//"):
                href = "https:" + href
            elif href.startswith("/"):
                href = base_url.rstrip("/") + href
            elif not href.startswith("http"):
                href = base_url.rstrip("/") + "/" + href
            # Must be a valid URL
            parsed = urllib.parse.urlparse(href)
            if parsed.scheme in ("http", "https"):
                return href

    return None


def _extract_release_period(html: bytes) -> str:
    """
    Attempt to extract the release period (e.g. 'Q1 2026') from hub HTML.

    Returns an empty string if no match is found.
    """
    text = html.decode("utf-8", errors="replace")
    match = _QUARTER_PATTERN.search(text)
    if match:
        groups = match.groups()
        if groups[0] and groups[1]:
            return f"Q{groups[0]} {groups[1]}"
        if groups[2] and groups[3]:
            return f"Q{groups[2]} {groups[3]}"
    return ""


def _extract_hub_etag_and_hash(
    client: HTTPClient,
    hub_url: str,
) -> tuple[str, str]:
    """
    Return (etag, content_sha256) for the release hub page.

    Used by check_for_updates to detect page changes cheaply.
    """
    try:
        response = client.get(hub_url)
        return response.etag or "", response.content_sha256
    except Exception:
        return "", ""


# ---------------------------------------------------------------------------
# QLFS direct URL construction and probing
# ---------------------------------------------------------------------------


def _build_qlfs_candidate_urls(quarter: int, year: int) -> list[str]:
    """
    Build an ordered list of candidate Excel URLs for the given QLFS quarter.

    Stats SA naming convention (confirmed from live URL probing, 2026-07-12):
      - PDFs:   ``Presentation%20QLFS%20Q{N}%20{YYYY}.pdf``  (confirmed working)
      - Excel:  No confirmed pattern yet — multiple variants tried below

    This list is ordered by likelihood based on observed Stats SA conventions.
    The caller probes each URL in order and uses the first that returns 200
    with a valid Excel file.

    Parameters
    ----------
    quarter : int
        Quarter number (1–4).
    year : int
        Full year (e.g. 2026).
    """
    base = _QLFS_PUBLICATION_BASE
    q = quarter
    y = year
    ord_suffix = _QUARTER_ORDINALS.get(q, f"{q}th")
    month = _QUARTER_MONTH_NAMES.get(q, "")

    # Build URL-encoded equivalents
    # Pattern: Presentation%20QLFS%20Q{N}%20{YYYY}  (spaces → %20)
    pres_prefix = f"Presentation%20QLFS%20Q{q}%20{y}"
    data_prefix1 = f"Data%20tables%20QLFS%20Q{q}%20{y}"
    data_prefix2 = f"Tables%20QLFS%20Q{q}%20{y}"
    data_prefix3 = f"Statistical%20tables%20Q{q}%20{y}"
    data_prefix4 = f"QLFS%20Q{q}%20{y}%20Statistical%20tables"
    data_prefix5 = f"P0211{ord_suffix}Quarter{y}"
    data_prefix6 = f"P0211%20{ord_suffix}%20Quarter%20{y}"

    candidates: list[str] = []
    for prefix in [
        pres_prefix,
        data_prefix1,
        data_prefix2,
        data_prefix3,
        data_prefix4,
        data_prefix5,
        data_prefix6,
    ]:
        for ext in (".xlsx", ".xls", ".pdf"):
            candidates.append(f"{base}{prefix}{ext}")

    return candidates


def _determine_current_qlfs_quarter() -> tuple[int, int]:
    """
    Determine the expected current QLFS release quarter.

    QLFS releases are approximately 6 weeks after quarter-end:
      Q1 (Jan-Mar) → released mid-May
      Q2 (Apr-Jun) → released mid-August
      Q3 (Jul-Sep) → released mid-November
      Q4 (Oct-Dec) → released mid-February

    Returns
    -------
    (quarter, year)
        The quarter and year of the most recently expected QLFS release.
    """
    today = date.today()
    m = today.month
    y = today.year

    # Most recently published quarter (by month)
    # Before May: Q4 of previous year
    # May-Jul: Q1 of current year
    # Aug-Oct: Q2 of current year
    # Nov-Jan: Q3 of current year
    if m < 5:
        return 4, y - 1
    elif m < 8:
        return 1, y
    elif m < 11:
        return 2, y
    else:
        return 3, y


def _probe_qlfs_publication_url(
    client: HTTPClient,
    quarter: int,
    year: int,
) -> str | None:
    """
    Probe candidate URLs for the QLFS quarter and return the first
    that responds with a valid file (HTTP 200, size > 10 KB).

    This is the primary discovery strategy.  It bypasses the
    Incapsula WAF by using direct file URLs rather than scraping the
    release hub HTML listing. It checks for Excel (.xlsx/.xls) first,
    then falls back to the statistical release PDF (.pdf).

    Parameters
    ----------
    client:
        HTTP client to use for probing.
    quarter:
        QLFS quarter (1–4).
    year:
        Year of the release.

    Returns
    -------
    str or None
        Working publication URL, or None if none of the candidates returned a file.
    """
    candidates = _build_qlfs_candidate_urls(quarter, year)
    log.debug(
        "Probing %d candidate URLs for QLFS Q%d %d …",
        len(candidates), quarter, year,
    )
    for url in candidates:
        try:
            # Use HEAD-style check first to avoid downloading a large file
            # multiple times.  Stats SA may not support HEAD, so we catch.
            resp = client.get(url)
            if resp.status == 200 and len(resp.body) > 10_240:  # > 10 KB
                log.info(
                    "QLFS publication found via direct URL probe: %s (%d bytes)",
                    url, len(resp.body),
                )
                return url
            log.debug("Probe %s → %d bytes (too small or error)", url, len(resp.body))
        except AutomationHTTPError as exc:
            if exc.status != 404:
                log.warning("Probe %s → HTTP %s: %s", url, exc.status, exc.reason)
            # 404 is expected for most candidates — skip silently
        except Exception as exc:
            log.debug("Probe %s → %s", url, exc)
    return None


# ---------------------------------------------------------------------------
# QLFS-specific fetch helpers
# ---------------------------------------------------------------------------


def _resolve_publication_url(
    client: HTTPClient,
    hub_url: str,
    probe_func: Callable[[], str | None],
    fallback_period_label: str,
) -> tuple[str | None, str | None, bytes | None]:
    """
    Resolve the publication URL using the ordered discovery architecture:
    Strategy 1: Direct URL probe
    Strategy 2: Release-hub scrape
    """
    file_url = probe_func()
    
    hub_html = None
    release_period = fallback_period_label if file_url else None
    
    try:
        hub_html = with_retry(
            lambda: _fetch_release_hub_html(client, hub_url),
            policy=STATSSA_POLICY,
            label=f"Release hub fetch ({hub_url})",
        )
        hub_period = _extract_release_period(hub_html)
        if hub_period:
            release_period = hub_period
            
        if not file_url:
            file_url = _extract_excel_url(hub_html)
    except AutomationHTTPError as exc:
        if "WAF_BLOCKED" in getattr(exc, "reason", str(exc)):
            log.warning("WAF block encountered during release hub scrape: %s", exc)
        else:
            raise
            
    return file_url, release_period, hub_html


def _discover_qlfs_excel(
    client: HTTPClient,
    *,
    hub_url: str = _QLFS_HUB_URL,
) -> tuple[str | None, str | None, bytes | None]:
    """
    Discover and return the QLFS Excel workbook URL and hub HTML.

    Returns
    -------
    (excel_url, release_period, hub_html)
        excel_url:      Absolute URL of the Excel workbook, or None.
        release_period: Detected quarter label (e.g. 'Q1 2026') or ''.
        hub_html:       Raw HTML bytes of the release hub.
    """
    q, y = _determine_current_qlfs_quarter()
    return _resolve_publication_url(
        client=client,
        hub_url=hub_url,
        probe_func=lambda: _probe_qlfs_publication_url(client, q, y),
        fallback_period_label=f"Q{q} {y}"
    )


def _download_publication(client: HTTPClient, url: str) -> bytes:
    """
    Download the publication file at ``url`` and return its raw bytes.

    Uses a separate HTTPClient call with Accept headers appropriate for
    binary file downloads.

    Raises
    ------
    AutomationHTTPError
        On non-retryable HTTP errors.
    urllib.error.URLError
        On transient network errors.
    """
    dl_client = HTTPClient(
        timeout_seconds=120,  # Files can be large
        extra_headers={
            **dict(_STATSSA_BROWSER_HEADERS),
            "Accept": "application/vnd.ms-excel,application/pdf,*/*",
        },
    )
    response = dl_client.get(url)
    if len(response.body) < 1024:
        raise ValueError(
            f"Downloaded file is suspiciously small ({len(response.body)} bytes) "
            f"from {url} — may be an error page, not a publication file."
        )
    return response.body


# ---------------------------------------------------------------------------
# QLFS Excel parsing (Phase 2)
#
# Pure functions: raw workbook bytes in, a small set of named values out.
# No knowledge of any JSON schema lives here — see IMPLEMENTATION-SPEC-
# STATSSA-PHASE2.md §5. Table location is by header/label text matching,
# not fixed cell coordinates, so a release with reflowed rows/columns
# still parses as long as the same label text appears somewhere in the
# sheet — but a genuinely new layout (renamed indicators, no quarter
# header row, etc.) will fail loudly via ValueError rather than silently
# mis-mapping a value. See the module docstring above for the verification
# status of this parser (no live archived file was available to test
# against in this session).
# ---------------------------------------------------------------------------


@dataclass
class QLFSExtract:
    """Named values extracted from a single QLFS Excel workbook."""

    release_period: str          # e.g. "Q1 2026"
    publication_date: str        # ISO YYYY-MM-DD, best-effort (see below)
    unemployment_rate: float
    youth_unemployment_narrow: float
    youth_unemployment_1524: float
    youth_unemployment_expanded: float
    neet_rate: float
    lfpr_overall: float
    lfpr_female: float


# Label specs used to locate each indicator's row by text match.
# ``include``: all substrings must be present (case-insensitive).
# ``exclude``: none of these substrings may be present.
# Order matters only in that more specific specs (e.g. "expanded") must
# exclude what a more general spec (e.g. plain "unemployment rate") would
# also match, and vice versa, so that no two metrics resolve to the same row.
_QLFS_METRIC_SPECS: dict[str, dict[str, tuple[str, ...]]] = {
    "unemployment_rate": {
        "include": ("unemployment rate",),
        "exclude": ("youth", "expanded", "female", "15"),
    },
    "youth_unemployment_narrow": {
        "include": ("youth", "15", "34"),
        "exclude": ("expanded", "neet"),
    },
    "youth_unemployment_1524": {
        "include": ("youth", "15", "24"),
        "exclude": ("expanded", "neet"),
    },
    "youth_unemployment_expanded": {
        "include": ("expanded", "youth"),
        "exclude": ("neet",),
    },
    "neet_rate": {
        "include": ("neet",),
        "exclude": (),
    },
    "lfpr_overall": {
        "include": ("labour force participation rate",),
        "exclude": ("female",),
    },
    "lfpr_female": {
        "include": ("labour force participation rate", "female"),
        "exclude": (),
    },
}

_MONTH_NAMES_RE = (
    "January|February|March|April|May|June|July|August|September|October|"
    "November|December"
)
_PUB_DATE_RE = re.compile(rf"(\d{{1,2}})\s+({_MONTH_NAMES_RE})\s+(\d{{4}})")


def _find_latest_quarter_column(ws: Any) -> tuple[int, str] | None:
    """
    Scan the first several rows of a worksheet for QLFS quarter-header
    cells (e.g. "Q1 2026") and return the column index and label of the
    chronologically latest one found.

    Returns None if no quarter-header cell is found in this sheet.
    """
    best: tuple[tuple[int, int], int, str] | None = None
    max_header_rows = min(15, ws.max_row or 1)
    for row in ws.iter_rows(min_row=1, max_row=max_header_rows):
        for cell in row:
            val = cell.value
            if not isinstance(val, str):
                continue
            match = _QUARTER_PATTERN.search(val)
            if not match:
                continue
            groups = match.groups()
            if groups[0] and groups[1]:
                q, y = int(groups[0]), int(groups[1])
            elif groups[2] and groups[3]:
                q, y = int(groups[2]), int(groups[3])
            else:
                continue
            key = (y, q)
            if best is None or key > best[0]:
                best = (key, cell.column, f"Q{q} {y}")
    if best is None:
        return None
    return best[1], best[2]


def _find_metric_value(
    ws: Any,
    col_idx: int,
    include: tuple[str, ...],
    exclude: tuple[str, ...] = (),
) -> float | None:
    """
    Search every row of a worksheet for a label cell matching ``include``/
    ``exclude``, then read the numeric value at ``col_idx`` in that same row.

    Returns None if no matching label row is found, or the value at
    ``col_idx`` cannot be interpreted as a number.
    """
    max_row = ws.max_row or 1
    for row in ws.iter_rows(min_row=1, max_row=max_row):
        for cell in row:
            val = cell.value
            if not isinstance(val, str):
                continue
            low = val.lower()
            if not all(term in low for term in include):
                continue
            if any(term in low for term in exclude):
                continue
            value_cell = ws.cell(row=cell.row, column=col_idx)
            v = value_cell.value
            if isinstance(v, bool):
                return None
            if isinstance(v, (int, float)):
                return float(v)
            if isinstance(v, str):
                try:
                    return float(v.strip().rstrip("%"))
                except ValueError:
                    return None
            return None
    return None


def _best_effort_publication_date(wb: Any) -> str | None:
    """
    Best-effort scan for a publication date embedded in the workbook
    (e.g. a title/cover cell reading "Released: 12 May 2026").

    This is metadata, not one of the required named values — a failure to
    find it does NOT raise; callers should fall back to another date
    source (e.g. today's date) and log accordingly.
    """
    for ws in wb.worksheets:
        max_row = min(10, ws.max_row or 1)
        for row in ws.iter_rows(min_row=1, max_row=max_row):
            for cell in row:
                val = cell.value
                if isinstance(val, datetime):
                    return val.date().isoformat()
                if isinstance(val, date):
                    return val.isoformat()
                if isinstance(val, str):
                    match = _PUB_DATE_RE.search(val)
                    if match:
                        try:
                            parsed = datetime.strptime(match.group(0), "%d %B %Y")
                            return parsed.date().isoformat()
                        except ValueError:
                            continue
    return None


def parse_qlfs_workbook(file_bytes: bytes) -> QLFSExtract:
    """
    Parse a QLFS Excel workbook and extract the named values needed by the
    unemployment / youth-unemployment / labour-force transforms.

    This function has no knowledge of any JSON schema — it returns a plain
    :class:`QLFSExtract` of named floats plus the detected release period.

    Raises
    ------
    ValueError
        If the workbook cannot be opened, or if one or more required
        indicators cannot be located by label match in any worksheet. The
        error message names exactly which indicator(s) failed to resolve,
        so a human reviewer can tell at a glance whether this is a layout
        change (see the module docstring's "Excel layout — verification
        status" note) rather than a transient/network problem.
    """
    try:
        wb = openpyxl.load_workbook(
            BytesIO(file_bytes), data_only=True, read_only=True
        )
    except Exception as exc:
        raise ValueError(
            f"Cannot open QLFS file as an Excel workbook — not a valid "
            f".xlsx/.xls file, or the file is corrupted: {exc}"
        ) from exc

    resolved: dict[str, float] = {}
    release_period: str | None = None
    missing: list[str] = []

    for metric_key, spec in _QLFS_METRIC_SPECS.items():
        value: float | None = None
        period_for_value: str | None = None
        for ws in wb.worksheets:
            header = _find_latest_quarter_column(ws)
            if header is None:
                continue
            col_idx, period_label = header
            found = _find_metric_value(
                ws, col_idx, spec["include"], spec.get("exclude", ())
            )
            if found is not None:
                value = found
                period_for_value = period_label
                break

        if value is None:
            missing.append(metric_key)
            continue

        resolved[metric_key] = value
        if release_period is None:
            release_period = period_for_value
        elif period_for_value is not None and period_for_value != release_period:
            log.warning(
                "QLFS parser: metric %s resolved to period %s, which "
                "differs from the period already resolved for other "
                "metrics (%s). Using %s for this run's release_period; "
                "this mismatch itself is exactly the class of cross-file "
                "period drift IMPLEMENTATION-SPEC-STATSSA-PHASE2.md §6 "
                "item 4 asks the run to flag, not silently paper over.",
                metric_key, period_for_value, release_period, release_period,
            )

    if missing or release_period is None:
        raise ValueError(
            "QLFS workbook parse failed — could not locate the following "
            f"required indicator(s) by label match: "
            f"{', '.join(missing) if missing else '(no quarter header found at all)'}. "
            "This most likely means the Stats SA Excel layout for this "
            "release differs from the label-matching rules in "
            "_QLFS_METRIC_SPECS (automation/adapters/statss.py) — per "
            "IMPLEMENTATION-SPEC-STATSSA-PHASE2.md §5, this must fail "
            "loudly rather than guess or fall back to a stale value. "
            "Manual review (Track B) is the correct next step, not a "
            "PDF-parsing fallback (explicitly out of scope for this phase)."
        )

    publication_date = _best_effort_publication_date(wb)
    if publication_date is None:
        publication_date = date.today().isoformat()
        log.warning(
            "QLFS parser: could not find an explicit publication date in "
            "the workbook — using today's date (%s) as a best-effort "
            "fallback for lastUpdated/source.publicationDate fields.",
            publication_date,
        )

    return QLFSExtract(
        release_period=release_period,
        publication_date=publication_date,
        unemployment_rate=resolved["unemployment_rate"],
        youth_unemployment_narrow=resolved["youth_unemployment_narrow"],
        youth_unemployment_1524=resolved["youth_unemployment_1524"],
        youth_unemployment_expanded=resolved["youth_unemployment_expanded"],
        neet_rate=resolved["neet_rate"],
        lfpr_overall=resolved["lfpr_overall"],
        lfpr_female=resolved["lfpr_female"],
    )


# ---------------------------------------------------------------------------
# GDP Excel parsing (Phase 3a)
#
# Unlike parse_qlfs_workbook() — which deliberately reads only the single
# most recent quarter column per metric, correct for QLFS since it does not
# revise prior quarters as a matter of routine practice in this codebase's
# scope — GDP must read every available quarter column in the growth
# table, because Stats SA routinely restates prior quarters' GDP growth
# figures in later releases (see IMPLEMENTATION-SPEC-GDP.md §1 item 3 and
# §5.1). The helpers below generalise _find_latest_quarter_column() /
# _find_metric_value() without modifying either — QLFS keeps using the
# single-column versions, unchanged.
# ---------------------------------------------------------------------------


def _find_all_quarter_columns(ws: Any) -> list[tuple[int, str]]:
    """
    Scan the first several rows of a worksheet for quarter-header cells
    (e.g. "Q1 2026") and return every (column_index, "Qn YYYY") pair found,
    sorted chronologically ascending (oldest first). Deduplicates by
    column index. Returns an empty list if no quarter-header cell is found.

    Generalises _find_latest_quarter_column(), which this function does
    NOT replace or modify — QLFS keeps using the single-column version
    unchanged.
    """
    found: dict[int, tuple[int, int, str]] = {}
    max_header_rows = min(15, ws.max_row or 1)
    for row in ws.iter_rows(min_row=1, max_row=max_header_rows):
        for cell in row:
            val = cell.value
            if not isinstance(val, str):
                continue
            match = _QUARTER_PATTERN.search(val)
            if not match:
                continue
            groups = match.groups()
            if groups[0] and groups[1]:
                q, y = int(groups[0]), int(groups[1])
            elif groups[2] and groups[3]:
                q, y = int(groups[2]), int(groups[3])
            else:
                continue
            found[cell.column] = (y, q, f"Q{q} {y}")

    ordered = sorted(found.items(), key=lambda item: (item[1][0], item[1][1]))
    return [(col_idx, label) for col_idx, (_year, _q, label) in ordered]


def _find_metric_row(
    ws: Any,
    include: tuple[str, ...],
    exclude: tuple[str, ...] = (),
) -> int | None:
    """
    Search every row of a worksheet for a label cell matching include/
    exclude (same case-insensitive substring rules as
    _find_metric_value()'s label matching). Returns the 1-indexed row
    number of the first match, or None.

    Factored out so the caller can read multiple columns from the same
    row (parse_gdp_workbook() needs this); _find_metric_value() itself is
    left unchanged for QLFS's continued single-column use.
    """
    max_row = ws.max_row or 1
    for row in ws.iter_rows(min_row=1, max_row=max_row):
        for cell in row:
            val = cell.value
            if not isinstance(val, str):
                continue
            low = val.lower()
            if not all(term in low for term in include):
                continue
            if any(term in low for term in exclude):
                continue
            return cell.row
    return None


def _read_row_values_at_columns(
    ws: Any,
    row_idx: int,
    columns: list[tuple[int, str]],
) -> list[tuple[str, float]]:
    """
    For each (col_idx, period_label) in `columns`, read the cell at
    (row_idx, col_idx). Uses the exact same numeric-coercion rules as
    _find_metric_value() (int/float pass through; numeric strings with a
    trailing '%' are stripped and parsed; bool and unparseable values are
    skipped, not raised). Columns whose cell is blank or unparseable are
    silently omitted from the result (a worksheet legitimately may not
    print a value for every historical column, e.g. a leading placeholder
    column) — this is not an error condition.

    Returns the list of (period_label, value) pairs actually found, in
    the same chronological order as `columns`.
    """
    results: list[tuple[str, float]] = []
    for col_idx, period_label in columns:
        cell = ws.cell(row=row_idx, column=col_idx)
        v = cell.value
        if isinstance(v, bool):
            continue
        if isinstance(v, (int, float)):
            results.append((period_label, float(v)))
            continue
        if isinstance(v, str):
            try:
                results.append((period_label, float(v.strip().rstrip("%"))))
            except ValueError:
                continue
    return results


@dataclass
class GDPExtract:
    """Named values extracted from a single GDP Excel workbook."""

    release_period: str                      # latest quarter found, e.g. "Q1 2026"
    publication_date: str                    # ISO YYYY-MM-DD, best-effort
    growth_points: list[tuple[str, float]]   # [(period_label, value), ...], chronological


# Label spec used to locate the GDP growth row by text match. Excludes an
# annual-growth row on the same worksheet without requiring knowledge of
# gdp-annual-growth's exact label, consistent with _QLFS_METRIC_SPECS's
# include/exclude convention. Unverified against a real Stats SA P0441
# workbook — see the module docstring's "GDP Excel layout — verification
# status" section. If the real label differs, parse_gdp_workbook() fails
# loudly rather than guessing, and this spec is the first and only place
# that needs correcting.
_GDP_GROWTH_SPEC: dict[str, tuple[str, ...]] = {
    "include": ("gdp growth",),
    "exclude": ("annual",),
}


def parse_gdp_workbook(file_bytes: bytes) -> GDPExtract:
    """
    Parse a GDP Excel workbook and extract every available quarterly GDP
    growth rate point (all columns present in the growth table, not just
    the latest — required for revision handling, see
    IMPLEMENTATION-SPEC-GDP.md §1 item 3).

    Algorithm
    ---------
    1. Open the workbook (openpyxl, data_only=True, read_only=True) — same
       error handling as parse_qlfs_workbook(): any exception opening the
       file is re-raised as ValueError with a clear message.
    2. For each worksheet, call _find_all_quarter_columns(). Skip
       worksheets where this returns an empty list.
    3. On the first worksheet with quarter columns, call _find_metric_row()
       using _GDP_GROWTH_SPEC's include/exclude terms.
    4. If a row is found, call _read_row_values_at_columns() to get every
       (period_label, value) pair present in that row.
    5. Stop at the first worksheet that yields at least one point.

    Raises
    ------
    ValueError
        If no worksheet yields both a quarter-header row AND a matching
        metric row with at least one readable value. The message names
        what specifically could not be found (no quarter-header row found
        at all vs. a quarter-header row was found but no row matched the
        GDP growth label), mirroring parse_qlfs_workbook()'s fail-loudly
        contract and its explicit pointer to manual review (Track B) as
        the correct next step — not a PDF-parsing fallback (explicitly
        out of scope, same as Phase 2).

    release_period is the label of the last (chronologically latest)
    entry in growth_points. publication_date is obtained via the existing,
    already-generic _best_effort_publication_date(wb) — reused as-is, no
    changes needed.
    """
    try:
        wb = openpyxl.load_workbook(
            BytesIO(file_bytes), data_only=True, read_only=True
        )
    except Exception as exc:
        raise ValueError(
            f"Cannot open GDP file as an Excel workbook — not a valid "
            f".xlsx/.xls file, or the file is corrupted: {exc}"
        ) from exc

    growth_points: list[tuple[str, float]] = []
    quarter_header_found = False
    metric_row_found = False

    for ws in wb.worksheets:
        columns = _find_all_quarter_columns(ws)
        if not columns:
            continue
        quarter_header_found = True

        row_idx = _find_metric_row(
            ws, _GDP_GROWTH_SPEC["include"], _GDP_GROWTH_SPEC.get("exclude", ())
        )
        if row_idx is None:
            continue
        metric_row_found = True

        points = _read_row_values_at_columns(ws, row_idx, columns)
        if points:
            growth_points = points
            break

    if not growth_points:
        if not quarter_header_found:
            raise ValueError(
                "GDP workbook parse failed — no quarter-header row (e.g. "
                "'Q1 2026') could be located in any worksheet. This most "
                "likely means the Stats SA P0441 Excel layout differs from "
                "the header-matching rules in parse_gdp_workbook() "
                "(automation/adapters/statss.py) — per "
                "IMPLEMENTATION-SPEC-GDP.md §5, this must fail loudly "
                "rather than guess or fall back to a stale value. Manual "
                "review (Track B) is the correct next step, not a "
                "PDF-parsing fallback (explicitly out of scope for this "
                "phase)."
            )
        if not metric_row_found:
            raise ValueError(
                "GDP workbook parse failed — a quarter-header row was "
                "found, but no row matched the GDP growth label spec "
                f"(include={_GDP_GROWTH_SPEC['include']!r}, "
                f"exclude={_GDP_GROWTH_SPEC.get('exclude', ())!r}). This "
                "most likely means the Stats SA P0441 Excel layout differs "
                "from _GDP_GROWTH_SPEC's label-matching rules "
                "(automation/adapters/statss.py) — per "
                "IMPLEMENTATION-SPEC-GDP.md §5, this must fail loudly "
                "rather than guess. Manual review (Track B) is the correct "
                "next step, not a PDF-parsing fallback (explicitly out of "
                "scope for this phase)."
            )
        raise ValueError(
            "GDP workbook parse failed — a GDP growth row was located but "
            "yielded no readable numeric values in any quarter column."
        )

    release_period = growth_points[-1][0]

    publication_date = _best_effort_publication_date(wb)
    if publication_date is None:
        publication_date = date.today().isoformat()
        log.warning(
            "GDP parser: could not find an explicit publication date in "
            "the workbook — using today's date (%s) as a best-effort "
            "fallback for lastUpdated/source.publicationDate fields.",
            publication_date,
        )

    return GDPExtract(
        release_period=release_period,
        publication_date=publication_date,
        growth_points=growth_points,
    )


# ---------------------------------------------------------------------------
# GDP direct URL construction and probing (Phase 3a)
#
# Mirrors the QLFS discovery pattern exactly, parameterised for P0441.
# **Unconfirmed** against a real Stats SA release — see the module
# docstring's "GDP Excel layout — verification status" section.
# ---------------------------------------------------------------------------


def _build_gdp_candidate_urls(quarter: int, year: int) -> list[str]:
    """
    Build an ordered list of candidate Excel/PDF URLs for the given GDP
    quarter, following _build_qlfs_candidate_urls()'s structure scaled to
    P0441's naming convention. **Unconfirmed** against a real release —
    see the module docstring.
    """
    base = _GDP_PUBLICATION_BASE
    q = quarter
    y = year
    ord_suffix = _QUARTER_ORDINALS.get(q, f"{q}th")

    pres_prefix = f"Presentation%20GDP%20Q{q}%20{y}"
    stat_release_prefix = f"Statistical%20release%20P0441%20Q{q}%20{y}"
    data_prefix1 = f"P0441{ord_suffix}Quarter{y}"
    data_prefix2 = f"GDP%20Q{q}%20{y}%20Statistical%20release"

    candidates: list[str] = []
    for prefix in [pres_prefix, stat_release_prefix, data_prefix1, data_prefix2]:
        for ext in (".xlsx", ".xls", ".pdf"):
            candidates.append(f"{base}{prefix}{ext}")

    return candidates


def _determine_current_gdp_quarter() -> tuple[int, int]:
    """
    Determine the expected most-recently-released GDP quarter, using
    GDP's own release windows (_GDP_RELEASE_WINDOWS: Q1→June, Q2→
    September, Q3→December, Q4→March of the following year) rather than
    QLFS's ~6-week offsets.

    Returns
    -------
    (quarter, year)
        The quarter and year of the most recently expected GDP release.
    """
    today = date.today()
    m = today.month
    y = today.year

    if m < 3:
        return 3, y - 1
    elif m < 6:
        return 4, y - 1
    elif m < 9:
        return 1, y
    elif m < 12:
        return 2, y
    else:
        return 3, y


def _probe_gdp_publication_url(
    client: HTTPClient,
    quarter: int,
    year: int,
) -> str | None:
    """
    Probe candidate URLs for the GDP quarter and return the first that
    responds with a valid file (HTTP 200, size > 10 KB). Structurally
    identical to _probe_qlfs_publication_url().
    """
    candidates = _build_gdp_candidate_urls(quarter, year)
    log.debug(
        "Probing %d candidate URLs for GDP Q%d %d …",
        len(candidates), quarter, year,
    )
    for url in candidates:
        try:
            resp = client.get(url)
            if resp.status == 200 and len(resp.body) > 10_240:
                log.info(
                    "GDP publication found via direct URL probe: %s (%d bytes)",
                    url, len(resp.body),
                )
                return url
            log.debug("Probe %s → %d bytes (too small or error)", url, len(resp.body))
        except AutomationHTTPError as exc:
            if exc.status != 404:
                log.warning("Probe %s → HTTP %s: %s", url, exc.status, exc.reason)
        except Exception as exc:
            log.debug("Probe %s → %s", url, exc)
    return None


def _discover_gdp_excel(
    client: HTTPClient,
    *,
    hub_url: str = _GDP_HUB_URL,
) -> tuple[str | None, str | None, bytes | None]:
    """
    Discover and return the GDP Excel workbook URL and hub HTML.
    Structurally identical to _discover_qlfs_excel(), reusing the fully
    generic _fetch_release_hub_html() / _extract_excel_url() /
    _extract_release_period() unchanged.

    Returns
    -------
    (excel_url, release_period, hub_html)
        excel_url:      Absolute URL of the Excel workbook, or None.
        release_period: Detected quarter label (e.g. 'Q1 2026') or ''.
        hub_html:       Raw HTML bytes of the release hub.
    """
    q, y = _determine_current_gdp_quarter()
    return _resolve_publication_url(
        client=client,
        hub_url=hub_url,
        probe_func=lambda: _probe_gdp_publication_url(client, q, y),
        fallback_period_label=f"Q{q} {y}"
    )


# ---------------------------------------------------------------------------
# GDP validation helpers (Phase 3a)
# ---------------------------------------------------------------------------


def _validate_gdp_growth_rate(value: float, label: str) -> list[str]:
    """
    Validate a GDP growth rate is a plausible percentage in
    _GDP_GROWTH_PLAUSIBLE_RANGE. Unlike QLFS rates, GDP growth can be
    negative and the plausible range is wider than [0, 100] — gdp.json
    itself already contains -6.2% (2020 annual figure) — so this is a
    genuinely new validator, not a reuse of _validate_percentage().
    """
    low, high = _GDP_GROWTH_PLAUSIBLE_RANGE
    if not (low <= value <= high):
        return [
            f"{label} GDP growth value {value} is outside the plausible "
            f"[{low}, {high}] range."
        ]
    return []


def _gdp_growth_values_changed(
    current_doc: dict[str, Any],
    growth_points: list[tuple[str, float]],
    *,
    stat_id: str = _GDP_GROWTH_STAT_ID,
) -> bool:
    """
    Return True if any (label, value) in `growth_points` differs from the
    corresponding existing series point for `stat_id` in `current_doc`
    (or is entirely absent from it), within the same 0.001 tolerance
    _apply_gdp_growth_points() uses for its in-place-revision comparison.

    This mirrors the QLFS flow's per-dataset "did anything actually
    change" check (`dataset_changed`), and — critically — must be
    computed BEFORE _transform_gdp() runs: _transform_gdp() always
    refreshes `_meta.last_verified`/`_meta.automation.updatedAt` to the
    current run, so comparing the full transformed document against the
    current on-disk document would never equal it even when no growth
    value actually changed. Computing the change flag from the raw
    growth_points first (as QLFS does from its extracted rates) avoids
    that trap and lets a genuine no-op run report "no_change" with zero
    side effects.
    """
    existing_data: list[dict[str, Any]] = []
    for stat in current_doc.get("statistics", []):
        if stat.get("id") == stat_id:
            series = stat.get("series") or []
            if series:
                existing_data = series[0].get("data", [])
            break

    existing_by_label = {pt["label"]: pt["value"] for pt in existing_data}
    for label, value in growth_points:
        if label not in existing_by_label:
            return True
        if abs(existing_by_label[label] - value) > 0.001:
            return True
    return False


# ---------------------------------------------------------------------------
# GDP transform helpers (Phase 3a)
#
# _apply_gdp_growth_points() is a deliberate generalisation of
# _apply_qlfs_rate_map() (multiple points instead of one; explicit
# revision-note tracking) rather than a call to it —
# _apply_qlfs_rate_map() is left completely unchanged, still used only by
# the three QLFS transforms.
# ---------------------------------------------------------------------------


def _apply_gdp_growth_points(
    doc: dict[str, Any],
    points: list[tuple[str, float]],
    *,
    stat_id: str = _GDP_GROWTH_STAT_ID,
    publication_date: str,
) -> list[str]:
    """
    Apply every (period_label, value) pair in `points` to the named stat
    in doc["statistics"], in place. Returns a list of human-readable
    revision/anomaly notes (e.g. "Revised Q2 2025: 0.8% -> 0.6%", plus any
    _check_qoq_jump() anomaly warnings) — surfaced in the version-entry
    notes for the human reviewer, since a silent multi-quarter revision is
    exactly the kind of change that benefits from an explicit summary.

    For each point, in the order given (chronological):
      - If the series already has a data point with this label:
          - If the value differs from the existing one by more than
            0.001, overwrite it in place and record a revision note (plus
            an anomaly flag if the swing exceeds
            _GDP_GROWTH_JUMP_WARNING_THRESHOLD). Exact equality is not
            required — reuses the same 0.001 tolerance
            _apply_qlfs_rate_map() already uses.
          - If it doesn't differ, leave it untouched (no note).
      - If the series has no data point with this label, append one,
        flagging an anomaly if the jump from the immediately preceding
        chronological point exceeds the threshold.
      - If the stat has no series at all yet, seed it with the first
        point (mirrors _apply_qlfs_rate_map()'s seed case).

    After all points are applied, update the stat's headline fields
    (value, rawValue, change, changeLabel, trend, lastUpdated,
    source.publicationDate) from ONLY the chronologically last point in
    `points` — mirroring _apply_qlfs_rate_map()'s existing field-update
    logic and formatting (value as f"{rate:.1f}%"; change computed against
    the series' new second-to-last chronological point AFTER revisions
    are applied, not against whatever it was before this run).

    Only `stat_id`'s fields are touched. No other stat in `doc` (e.g.
    gdp-annual-growth, gdp-nominal, gdp-per-capita) is read or modified —
    this is the mechanism that keeps this milestone's blast radius
    limited to gdp-growth.
    """
    notes: list[str] = []

    target_stat: dict[str, Any] | None = None
    for stat in doc.get("statistics", []):
        if stat.get("id") == stat_id:
            target_stat = stat
            break
    if target_stat is None or not points:
        return notes

    if not target_stat.get("series"):
        first_label, first_value = points[0]
        target_stat["series"] = [
            {
                "name": target_stat.get("title", stat_id),
                "unit": target_stat.get("unit", "%"),
                "data": [{"label": first_label, "value": first_value}],
            }
        ]
        log.debug(
            "Seeded first series point %r -> %.1f for %s",
            first_label, first_value, stat_id,
        )
        remaining = points[1:]
    else:
        remaining = points

    series = target_stat["series"]
    data = series[0].setdefault("data", [])

    for label, value in remaining:
        existing = next((pt for pt in data if pt["label"] == label), None)
        if existing is None:
            prev_value = data[-1]["value"] if data else None
            data.append({"label": label, "value": value})
            log.debug("Appended series point %r -> %.1f for %s", label, value, stat_id)
            warning = _check_qoq_jump(
                prev_value, value, label,
                threshold=_GDP_GROWTH_JUMP_WARNING_THRESHOLD,
            )
            if warning:
                notes.append(warning)
        else:
            if abs(existing["value"] - value) > 0.001:
                old_value = existing["value"]
                log.info(
                    "Revising existing series point %r: %.1f -> %.1f for %s",
                    label, old_value, value, stat_id,
                )
                notes.append(f"Revised {label}: {old_value}% -> {value}%")
                existing["value"] = value
                warning = _check_qoq_jump(
                    old_value, value, label,
                    threshold=_GDP_GROWTH_JUMP_WARNING_THRESHOLD,
                )
                if warning:
                    notes.append(warning)

    revision_count = sum(1 for n in notes if n.startswith("Revised "))
    if revision_count > 2:
        log.info(
            "_apply_gdp_growth_points revised %d historical points in a "
            "single run for %s — a normal, documented Stats SA practice, "
            "not a defect (IMPLEMENTATION-SPEC-GDP.md §6 item 5).",
            revision_count, stat_id,
        )

    newest_label, newest_value = points[-1]
    prev_value = None
    prev_label = None
    for i, pt in enumerate(data):
        if pt["label"] == newest_label and i > 0:
            prev_value = data[i - 1]["value"]
            prev_label = data[i - 1]["label"]
            break

    change = 0.0
    if prev_value is not None:
        change = round(newest_value - prev_value, 1)

    target_stat["value"] = f"{newest_value:.1f}%"
    target_stat["rawValue"] = newest_value
    target_stat["change"] = change
    target_stat["changeLabel"] = f"from {prev_label}" if prev_label else newest_label
    if prev_value is None:
        target_stat["trend"] = "stable"
    elif newest_value > prev_value:
        target_stat["trend"] = "up"
    elif newest_value < prev_value:
        target_stat["trend"] = "down"
    else:
        target_stat["trend"] = "stable"
    target_stat["lastUpdated"] = publication_date
    if isinstance(target_stat.get("source"), dict):
        target_stat["source"]["publicationDate"] = publication_date

    return notes


def _update_gdp_meta(
    doc: dict[str, Any],
    *,
    release_period: str,
    publication_date: str,
    source_url: str,
) -> None:
    """Update the shared _meta block, following the SARB/QLFS _meta pattern."""
    if "_meta" not in doc:
        doc["_meta"] = {}
    today_str = date.today().isoformat()
    doc["_meta"]["last_verified"] = today_str
    doc["_meta"]["lastUpdated"] = publication_date
    doc["_meta"]["source_url"] = source_url
    doc["_meta"]["automation"] = {
        "updatedBy": "statssa-adapter/gdp",
        "updatedAt": datetime.now(tz=timezone.utc).isoformat(),
        "releasePeriod": release_period,
        "sourceFile": source_url,
    }


def _transform_gdp(
    doc: dict[str, Any],
    extract: GDPExtract,
    source_url: str = "",
) -> tuple[dict[str, Any], list[str]]:
    """
    Deep-copy `doc`, apply _apply_gdp_growth_points() to the gdp-growth
    stat only, update the shared _meta block (mirroring
    _update_qlfs_meta()'s pattern: last_verified, lastUpdated, source_url,
    and an `automation` sub-block with updatedBy="statssa-adapter/gdp"),
    and return (new_doc, warnings) where warnings is the combined list of
    revision notes and any anomaly flags from _check_qoq_jump() (both
    produced by _apply_gdp_growth_points()).

    Matches _transform_unemployment()'s contract exactly: input document
    is never mutated; the return value is a new dict.
    """
    new_doc = copy.deepcopy(doc)
    warnings = _apply_gdp_growth_points(
        new_doc,
        extract.growth_points,
        publication_date=extract.publication_date,
    )
    _update_gdp_meta(
        new_doc,
        release_period=extract.release_period,
        publication_date=extract.publication_date,
        source_url=source_url,
    )
    return new_doc, warnings


# ---------------------------------------------------------------------------
# CPI Excel parsing (Phase 3b)
#
# CPI, like QLFS, only needs the newest month's value per metric — Stats SA
# does not routinely revise previously published CPI prints the way it
# routinely revises GDP quarters (IMPLEMENTATION-SPEC-CPI.md §10.1,
# assumption flagged in §17 item 1). _find_latest_month_column() is
# therefore a direct parallel to _find_latest_quarter_column() (single
# newest column), not to _find_all_quarter_columns() (every column) — it
# does not replace or modify _find_latest_quarter_column(); QLFS keeps
# using its own version, unchanged.
# ---------------------------------------------------------------------------


def _find_latest_month_column(ws: Any) -> tuple[int, str] | None:
    """
    Scan the first several rows of a worksheet for CPI month-header cells
    (e.g. "May 2026" or "May-2026") and return the column index and
    normalised "Mon YYYY" label of the chronologically latest one found.

    Direct parallel to _find_latest_quarter_column() for month-year
    headers instead of quarter-year headers. Does not replace or modify
    _find_latest_quarter_column() — QLFS keeps using its own version,
    unchanged.

    Returns None if no month-header cell is found in this sheet.
    """
    best: tuple[tuple[int, int], int, str] | None = None
    max_header_rows = min(15, ws.max_row or 1)
    for row in ws.iter_rows(min_row=1, max_row=max_header_rows):
        for cell in row:
            val = cell.value
            if not isinstance(val, str):
                continue
            match = _MONTH_HEADER_PATTERN.search(val)
            if not match:
                continue
            month_str, year_str = match.group(1), match.group(2)
            month_num = _MONTH_ABBR_TO_NUM.get(month_str.lower()[:3])
            if month_num is None:
                continue
            year = int(year_str)
            key = (year, month_num)
            label = f"{_MONTH_NUM_TO_ABBR[month_num]} {year}"
            if best is None or key > best[0]:
                best = (key, cell.column, label)
    if best is None:
        return None
    return best[1], best[2]


@dataclass
class CPIExtract:
    """Named values extracted from a single CPI Excel workbook."""

    release_period: str        # latest month found, e.g. "May 2026"
    publication_date: str      # ISO YYYY-MM-DD, best-effort
    cpi_headline: float
    food_inflation: float


# Label spec used to locate cpi-headline / food-inflation by text match.
# **Unverified** against a real Stats SA P0141 workbook — no session to
# date has had network access to statssa.gov.za (IMPLEMENTATION-SPEC-
# CPI.md §10.3 / §17 assumption 3). If the real labels differ,
# parse_cpi_workbook() fails loudly rather than guessing, and this spec is
# the first and only place that needs correcting.
_CPI_METRIC_SPECS: dict[str, dict[str, tuple[str, ...]]] = {
    "cpi_headline": {
        "include": ("all items",),
        "exclude": ("food",),
    },
    "food_inflation": {
        "include": ("food",),
        "exclude": (),
    },
}


def parse_cpi_workbook(file_bytes: bytes) -> CPIExtract:
    """
    Parse a CPI Excel workbook and extract the latest month's headline and
    food CPI values, by label/header matching (not fixed cell
    coordinates) — same philosophy as parse_qlfs_workbook().

    Raises
    ------
    ValueError
        If the workbook cannot be opened, or if either required indicator
        cannot be located by label match in any worksheet. The message
        names exactly which metric(s) failed to resolve, mirroring
        parse_qlfs_workbook()'s and parse_gdp_workbook()'s fail-loudly
        contract — no PDF fallback, no guessing, no stale-value
        substitution.
    """
    try:
        wb = openpyxl.load_workbook(
            BytesIO(file_bytes), data_only=True, read_only=True
        )
    except Exception as exc:
        raise ValueError(
            f"Cannot open CPI file as an Excel workbook — not a valid "
            f".xlsx/.xls file, or the file is corrupted: {exc}"
        ) from exc

    resolved: dict[str, float] = {}
    release_period: str | None = None
    missing: list[str] = []
    month_header_found = False

    for metric_key, spec in _CPI_METRIC_SPECS.items():
        value: float | None = None
        period_for_value: str | None = None
        for ws in wb.worksheets:
            header = _find_latest_month_column(ws)
            if header is None:
                continue
            month_header_found = True
            col_idx, period_label = header
            found = _find_metric_value(
                ws, col_idx, spec["include"], spec.get("exclude", ())
            )
            if found is not None:
                value = found
                period_for_value = period_label
                break

        if value is None:
            missing.append(metric_key)
            continue

        resolved[metric_key] = value
        if release_period is None:
            release_period = period_for_value
        elif period_for_value is not None and period_for_value != release_period:
            log.warning(
                "CPI parser: metric %s resolved to period %s, which "
                "differs from the period already resolved for other "
                "metrics (%s). Using %s for this run's release_period.",
                metric_key, period_for_value, release_period, release_period,
            )

    if not month_header_found:
        raise ValueError(
            "CPI workbook parse failed — no month-header row (e.g. "
            "'May 2026') could be located in any worksheet. This most "
            "likely means the Stats SA P0141 Excel layout differs from "
            "the header-matching rules in _find_latest_month_column() "
            "(automation/adapters/statss.py) — per IMPLEMENTATION-SPEC-"
            "CPI.md §10, this must fail loudly rather than guess or fall "
            "back to a stale value. Manual review (Track B) is the "
            "correct next step, not a PDF-parsing fallback (explicitly "
            "out of scope for this phase)."
        )

    if missing or release_period is None:
        raise ValueError(
            "CPI workbook parse failed — could not locate the following "
            f"required indicator(s) by label match: {', '.join(missing)}. "
            "This most likely means the Stats SA Excel layout for this "
            "release differs from the label-matching rules in "
            "_CPI_METRIC_SPECS (automation/adapters/statss.py) — per "
            "IMPLEMENTATION-SPEC-CPI.md §10.3, this must fail loudly "
            "rather than guess or fall back to a stale value. Manual "
            "review (Track B) is the correct next step, not a "
            "PDF-parsing fallback (explicitly out of scope for this "
            "phase)."
        )

    publication_date = _best_effort_publication_date(wb)
    if publication_date is None:
        publication_date = date.today().isoformat()
        log.warning(
            "CPI parser: could not find an explicit publication date in "
            "the workbook — using today's date (%s) as a best-effort "
            "fallback for lastUpdated/source.publicationDate fields.",
            publication_date,
        )

    return CPIExtract(
        release_period=release_period,
        publication_date=publication_date,
        cpi_headline=resolved["cpi_headline"],
        food_inflation=resolved["food_inflation"],
    )


# ---------------------------------------------------------------------------
# Population (MYPE, P0302) Excel parsing (Phase 4)
#
# Same header/label-matching philosophy as parse_qlfs_workbook() /
# parse_gdp_workbook() / parse_cpi_workbook(), adapted for a bare 4-digit
# year header row instead of a quarterly or monthly one — see
# IMPLEMENTATION-SPEC-POPULATION.md §7.
# ---------------------------------------------------------------------------

_YEAR_HEADER_PATTERN = re.compile(r"^\s*(\d{4})\s*$")


@dataclass
class PopulationExtract:
    """Named values extracted from a single MYPE Excel workbook."""

    release_period: str          # the estimate year, e.g. "2026"
    publication_date: str        # ISO YYYY-MM-DD, best-effort
    total_population_millions: float   # e.g. 63.1
    # Domain the workbook was downloaded from — feeds the source guard
    # (§7.3). This is the one field with no analogue in
    # QLFSExtract/GDPExtract/CPIExtract: it lets the source guard be
    # enforced with data the parser itself observed, not just the URL the
    # caller happened to pass in.
    source_domain: str


# Label spec used to locate the "total population" row by text match.
# **Unverified** against a real Stats SA P0302 workbook — no session to
# date has had network access to statssa.gov.za (IMPLEMENTATION-SPEC-
# POPULATION.md §14 item 1). If the real labels differ,
# parse_population_workbook() fails loudly rather than guessing, and this
# spec is the first and only place that needs correcting.
_POPULATION_METRIC_SPECS: dict[str, dict[str, tuple[str, ...]]] = {
    "total_population": {
        "include": ("total", "rsa", "south africa"),
        "exclude": (),
    },
}


def _find_latest_year_column(ws: Any) -> tuple[int, str] | None:
    """
    Scan the first several rows of a worksheet for bare 4-digit-year header
    cells (e.g. "2026", not "Q1 2026" or "May 2026") and return the column
    index and label of the chronologically latest one found.

    Parallel to _find_latest_quarter_column() / _find_latest_month_column()
    for an annual header instead of a quarterly/monthly one. Does not
    replace or modify either — QLFS/GDP and CPI keep using their own
    versions, unchanged.

    Returns None if no year-header cell is found in this sheet.
    """
    best: tuple[int, int, str] | None = None
    max_header_rows = min(15, ws.max_row or 1)
    for row in ws.iter_rows(min_row=1, max_row=max_header_rows):
        for cell in row:
            val = cell.value
            year: int | None = None
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                if 1900 <= int(val) <= 2100 and float(val).is_integer():
                    year = int(val)
            elif isinstance(val, str):
                match = _YEAR_HEADER_PATTERN.match(val)
                if match:
                    year = int(match.group(1))
            if year is None:
                continue
            if best is None or year > best[0]:
                best = (year, cell.column, str(year))
    if best is None:
        return None
    return best[1], best[2]


def parse_population_workbook(file_bytes: bytes) -> PopulationExtract:
    """
    Parse a Population (MYPE) Excel workbook and extract the latest year's
    total population figure, by label/header matching (not fixed cell
    coordinates) — same philosophy as parse_qlfs_workbook() /
    parse_gdp_workbook() / parse_cpi_workbook().

    Raises
    ------
    ValueError
        If the workbook cannot be opened, or if no year-header row or no
        matching "total population" row can be located in any worksheet.
        The message names exactly which indicator failed to resolve,
        mirroring the three prior parsers' fail-loudly contract — no PDF
        fallback, no guessing, no stale-value substitution.
    """
    try:
        wb = openpyxl.load_workbook(
            BytesIO(file_bytes), data_only=True, read_only=True
        )
    except Exception as exc:
        raise ValueError(
            f"Cannot open Population file as an Excel workbook — not a "
            f"valid .xlsx/.xls file, or the file is corrupted: {exc}"
        ) from exc

    release_period: str | None = None
    raw_value: float | None = None
    year_header_found = False

    spec = _POPULATION_METRIC_SPECS["total_population"]
    for ws in wb.worksheets:
        header = _find_latest_year_column(ws)
        if header is None:
            continue
        year_header_found = True
        col_idx, period_label = header
        found = _find_metric_value(ws, col_idx, spec["include"], spec.get("exclude", ()))
        if found is not None:
            raw_value = found
            release_period = period_label
            break

    if not year_header_found:
        raise ValueError(
            "Population workbook parse failed — no year-header row (e.g. "
            "'2026') could be located in any worksheet. This most likely "
            "means the Stats SA P0302 Excel layout differs from the "
            "header-matching rules in _find_latest_year_column() "
            "(automation/adapters/statss.py) — per IMPLEMENTATION-SPEC-"
            "POPULATION.md §7.2, this must fail loudly rather than guess "
            "or fall back to a stale value. Manual review is the correct "
            "next step, not a PDF-parsing fallback (explicitly out of "
            "scope for this phase)."
        )

    if raw_value is None or release_period is None:
        raise ValueError(
            "Population workbook parse failed — could not locate the "
            "'total population' row by label match. This most likely "
            "means the Stats SA Excel layout for this release differs "
            "from the label-matching rules in _POPULATION_METRIC_SPECS "
            "(automation/adapters/statss.py) — per IMPLEMENTATION-SPEC-"
            "POPULATION.md §7.2, this must fail loudly rather than guess "
            "or fall back to a stale value. Manual review is the correct "
            "next step, not a PDF-parsing fallback (explicitly out of "
            "scope for this phase)."
        )

    # Raw-headcount-vs-millions heuristic (IMPLEMENTATION-SPEC-POPULATION.md
    # §7.2 / §14 item 2): the workbook may express the total either as a raw
    # headcount (e.g. 63100000) or already in millions (e.g. 63.1). This is
    # genuinely unconfirmed against a real release, so the branch taken is
    # logged rather than silently assumed.
    if raw_value > 1000:
        total_population_millions = raw_value / 1_000_000
        log.info(
            "Population parser: value %s > 1000 — treating as a raw "
            "headcount and dividing by 1,000,000 → %.1fM.",
            raw_value, total_population_millions,
        )
    else:
        total_population_millions = raw_value
        log.info(
            "Population parser: value %s <= 1000 — treating as already "
            "expressed in millions.",
            raw_value,
        )

    publication_date = _best_effort_publication_date(wb)
    if publication_date is None:
        publication_date = date.today().isoformat()
        log.warning(
            "Population parser: could not find an explicit publication "
            "date in the workbook — using today's date (%s) as a "
            "best-effort fallback for lastUpdated/source.publicationDate "
            "fields.",
            publication_date,
        )

    return PopulationExtract(
        release_period=release_period,
        publication_date=publication_date,
        total_population_millions=total_population_millions,
        source_domain="",  # filled in by the caller once the source URL is known
    )


# ---------------------------------------------------------------------------
# CPI direct URL construction and probing (Phase 3b)
#
# Mirrors the QLFS/GDP discovery pattern exactly, parameterised for P0141
# and its monthly cadence. **Unconfirmed** against a real Stats SA release
# — see the module docstring's "CPI Excel layout — verification status"
# section.
# ---------------------------------------------------------------------------

_CPI_MONTH_NAMES: dict[int, str] = {
    1: "January", 2: "February", 3: "March", 4: "April", 5: "May", 6: "June",
    7: "July", 8: "August", 9: "September", 10: "October", 11: "November",
    12: "December",
}


def _build_cpi_candidate_urls(month: int, year: int) -> list[str]:
    """
    Build an ordered list of candidate Excel/PDF URLs for the given CPI
    release month, following _build_qlfs_candidate_urls()'s /
    _build_gdp_candidate_urls()'s structure scaled to P0141's naming
    convention. **Unconfirmed** against a real release — see the module
    docstring.
    """
    base = _CPI_PUBLICATION_BASE
    month_name = _CPI_MONTH_NAMES.get(month, "")

    stat_release_prefix = f"Statistical%20release%20P0141%20{month_name}%20{year}"
    media_release_prefix = f"CPI%20Media%20Release%20{month_name}%20{year}"
    data_prefix = f"P0141{month_name}{year}"

    candidates: list[str] = []
    for prefix in [stat_release_prefix, media_release_prefix, data_prefix]:
        for ext in (".xlsx", ".xls", ".pdf"):
            candidates.append(f"{base}{prefix}{ext}")

    return candidates


def _determine_current_cpi_month() -> tuple[int, int]:
    """
    Determine the most recently expected CPI release month.

    CPI for month M is released ~22nd of month M+1 (dataset-analysis.md;
    SA-Data-Hub-Dataset-Sourcing-Plan.md §4 notes this drifts and is not a
    fixed calendar day). Before the ~22nd of the current month, the most
    recently released figure is for two months prior; from the 22nd
    onward, it is for the previous month.

    Returns
    -------
    (month, year)
        The month and year of the most recently expected CPI release.
    """
    today = date.today()
    m = today.month
    y = today.year

    if today.day < 22:
        m -= 2
    else:
        m -= 1

    while m < 1:
        m += 12
        y -= 1

    return m, y


def _probe_cpi_publication_url(
    client: HTTPClient,
    month: int,
    year: int,
) -> str | None:
    """
    Probe candidate URLs for the CPI release month and return the first
    that responds with a valid file (HTTP 200, size > 10 KB).
    Structurally identical to _probe_qlfs_publication_url() /
    _probe_gdp_publication_url().
    """
    candidates = _build_cpi_candidate_urls(month, year)
    log.debug(
        "Probing %d candidate URLs for CPI %s %d …",
        len(candidates), _CPI_MONTH_NAMES.get(month, month), year,
    )
    for url in candidates:
        try:
            resp = client.get(url)
            if resp.status == 200 and len(resp.body) > 10_240:
                log.info(
                    "CPI publication found via direct URL probe: %s (%d bytes)",
                    url, len(resp.body),
                )
                return url
            log.debug("Probe %s → %d bytes (too small or error)", url, len(resp.body))
        except AutomationHTTPError as exc:
            if exc.status != 404:
                log.warning("Probe %s → HTTP %s: %s", url, exc.status, exc.reason)
        except Exception as exc:
            log.debug("Probe %s → %s", url, exc)
    return None


def _discover_cpi_excel(
    client: HTTPClient,
    *,
    hub_url: str = _CPI_HUB_URL,
) -> tuple[str | None, str | None, bytes | None]:
    """
    Discover and return the CPI Excel workbook URL and hub HTML.
    Structurally identical to _discover_qlfs_excel() /
    _discover_gdp_excel(), reusing the fully generic
    _fetch_release_hub_html() / _extract_excel_url() /
    _extract_release_period() unchanged.

    Returns
    -------
    (excel_url, release_period, hub_html)
        excel_url:      Absolute URL of the Excel workbook, or None.
        release_period: Detected month label (e.g. 'May 2026') or ''.
        hub_html:       Raw HTML bytes of the release hub.
    """
    m, y = _determine_current_cpi_month()
    month_name = _CPI_MONTH_NAMES.get(m, str(m))
    return _resolve_publication_url(
        client=client,
        hub_url=hub_url,
        probe_func=lambda: _probe_cpi_publication_url(client, m, y),
        fallback_period_label=f"{month_name} {y}"
    )


# ---------------------------------------------------------------------------
# Population direct URL construction and probing (Phase 4)
#
# Mirrors the QLFS/GDP/CPI discovery pattern exactly, parameterised for
# P0302's annual cadence (one release per year, not per quarter/month).
# **Unconfirmed** against a real Stats SA release — see the module
# docstring's "Population Excel layout — verification status" section.
# ---------------------------------------------------------------------------


def _build_population_candidate_urls(year: int) -> list[str]:
    """
    Build an ordered list of candidate Excel/PDF URLs for the given MYPE
    release year, following _build_qlfs_candidate_urls()'s /
    _build_gdp_candidate_urls()'s / _build_cpi_candidate_urls()'s structure
    scaled to P0302's naming convention. **Unconfirmed** against a real
    release — see the module docstring.
    """
    base = _POPULATION_PUBLICATION_BASE
    stat_release_prefix = f"Statistical%20release%20P0302%20{year}"
    media_release_prefix = f"MYPE%20Media%20Release%20{year}"
    data_prefix = f"P0302{year}"
    candidates: list[str] = []
    for prefix in [stat_release_prefix, media_release_prefix, data_prefix]:
        for ext in (".xlsx", ".xls", ".pdf"):
            candidates.append(f"{base}{prefix}{ext}")
    return candidates


def _determine_current_population_year() -> int:
    """
    Determine the most recently expected MYPE release year.

    MYPE for year Y is released in late July of year Y. Before ~28 July of
    the current year, the most recently expected release is for the
    *prior* year; from ~28 July onward, the current year's release is
    expected. Simpler than CPI's day-of-month cutoff since there is
    exactly one release window per year, not twelve.
    """
    today = date.today()
    return today.year if today.month > 7 or (today.month == 7 and today.day >= 28) else today.year - 1


def _probe_population_publication_url(
    client: HTTPClient,
    year: int,
) -> str | None:
    """
    Probe candidate URLs for the MYPE release year and return the first
    that responds with a valid file (HTTP 200, size > 10 KB).
    Structurally identical to _probe_qlfs_publication_url() /
    _probe_gdp_publication_url() / _probe_cpi_publication_url().
    """
    candidates = _build_population_candidate_urls(year)
    log.debug(
        "Probing %d candidate URLs for Population MYPE %d …",
        len(candidates), year,
    )
    for url in candidates:
        try:
            resp = client.get(url)
            if resp.status == 200 and len(resp.body) > 10_240:
                log.info(
                    "Population publication found via direct URL probe: %s (%d bytes)",
                    url, len(resp.body),
                )
                return url
            log.debug("Probe %s → %d bytes (too small or error)", url, len(resp.body))
        except AutomationHTTPError as exc:
            if exc.status != 404:
                log.warning("Probe %s → HTTP %s: %s", url, exc.status, exc.reason)
        except Exception as exc:
            log.debug("Probe %s → %s", url, exc)
    return None


def _discover_population_excel(
    client: HTTPClient,
    *,
    hub_url: str = _POPULATION_HUB_URL,
) -> tuple[str | None, str | None, bytes | None]:
    """
    Discover and return the Population Excel workbook URL and hub HTML.
    Structurally identical to _discover_qlfs_excel() / _discover_gdp_excel()
    / _discover_cpi_excel(), reusing the fully generic
    _fetch_release_hub_html() / _extract_excel_url() /
    _extract_release_period() unchanged.

    Returns
    -------
    (excel_url, release_period, hub_html)
        excel_url:      Absolute URL of the Excel workbook, or None.
        release_period: Detected year label (e.g. '2026') or ''.
        hub_html:       Raw HTML bytes of the release hub.
    """
    y = _determine_current_population_year()
    return _resolve_publication_url(
        client=client,
        hub_url=hub_url,
        probe_func=lambda: _probe_population_publication_url(client, y),
        fallback_period_label=str(y)
    )


# ---------------------------------------------------------------------------
# CPI validation helpers (Phase 3b)
# ---------------------------------------------------------------------------


def _validate_cpi_rate(value: float, label: str) -> list[str]:
    """
    Validate a CPI year-on-year rate is a plausible percentage in
    _CPI_PLAUSIBLE_RANGE. Unlike QLFS rates, CPI can in principle be
    negative (deflation), so this is a genuinely new validator, not a
    reuse of _validate_percentage()'s [0, 100] range.
    """
    low, high = _CPI_PLAUSIBLE_RANGE
    if not (low <= value <= high):
        return [
            f"{label} CPI value {value} is outside the plausible "
            f"[{low}, {high}] range."
        ]
    return []


def _validate_monthly_label(label: str) -> list[str]:
    """Validate a period label matches the monthly format (dataset-analysis.md RULES)."""
    if not _MONTHLY_LABEL_RE.match(label):
        return [f"Release period {label!r} does not match the expected 'Mon YYYY' format."]
    return []


# ---------------------------------------------------------------------------
# CPI ownership boundary (Phase 3b)
#
# The single most load-bearing new function in this milestone
# (IMPLEMENTATION-SPEC-CPI.md §7 / §11 item 5). Stricter than
# check_protected_fields(): it hard-fails on ANY difference — not just a
# protected-field change — in a stat this milestone does not own.
# ---------------------------------------------------------------------------


def _assert_cpi_ownership_boundary(
    previous_doc: dict[str, Any],
    proposed_doc: dict[str, Any],
) -> list[str]:
    """
    Deep-compare every statistics[] entry in `proposed_doc` whose id is
    NOT in _CPI_OWNED_STAT_IDS against the corresponding entry in
    `previous_doc` (matched by id). Returns a list of violation messages —
    empty if repo-rate and annual-cpi-avg (and any other non-owned stat)
    are byte-for-byte identical between the two documents, and the set of
    stat IDs present is unchanged.
    """
    violations: list[str] = []

    previous_by_id = {s.get("id"): s for s in previous_doc.get("statistics", [])}
    proposed_by_id = {s.get("id"): s for s in proposed_doc.get("statistics", [])}

    if set(previous_by_id.keys()) != set(proposed_by_id.keys()):
        violations.append(
            "CPI ownership boundary violation: the set of stat IDs in "
            "inflation.json changed (added/removed a stat) — previous="
            f"{sorted(k for k in previous_by_id if k is not None)!r}, "
            f"proposed={sorted(k for k in proposed_by_id if k is not None)!r}."
        )

    for stat_id, prev_stat in previous_by_id.items():
        if stat_id in _CPI_OWNED_STAT_IDS:
            continue
        proposed_stat = proposed_by_id.get(stat_id)
        if proposed_stat != prev_stat:
            violations.append(
                f"CPI ownership boundary violation: non-owned stat "
                f"{stat_id!r} changed. This milestone's code MUST NOT "
                f"modify repo-rate or annual-cpi-avg (IMPLEMENTATION-"
                f"SPEC-CPI.md §0.1/§0.2/§7)."
            )

    return violations


# ---------------------------------------------------------------------------
# CPI transform helpers (Phase 3b)
#
# _apply_qlfs_rate_map() (§5, unchanged) is reused directly for CPI — see
# IMPLEMENTATION-SPEC-CPI.md §12.1: despite its name, it has no
# QLFS-specific logic and only mutates stats whose id is a key in the
# rate_map it's given, which is exactly the scoping mechanism this
# milestone's ownership boundary depends on.
# ---------------------------------------------------------------------------


def _update_cpi_meta(
    doc: dict[str, Any],
    *,
    release_period: str,
    publication_date: str,
) -> None:
    """
    Update ONLY doc["_meta"]["last_verified"] and doc["_meta"]["automation"].

    Deliberately does NOT touch _meta["source"], _meta["source_url"],
    _meta["update_frequency"], or _meta["notes"] — unlike
    _update_qlfs_meta()/_update_gdp_meta(), which overwrite source_url
    unconditionally. inflation.json's _meta block is shared prose
    describing BOTH the Stats SA CPI component and the SARB repo-rate
    component (its "notes" field explicitly mentions MPC cadence).
    Rewriting those fields is a documentation/copy decision that belongs
    to a human editing the file deliberately, not something an automated
    CPI-only write path should do as a side effect of updating two
    numbers (IMPLEMENTATION-SPEC-CPI.md §12.3).
    """
    if "_meta" not in doc:
        doc["_meta"] = {}
    doc["_meta"]["last_verified"] = date.today().isoformat()
    doc["_meta"]["automation"] = {
        "updatedBy": "statssa-adapter/cpi",
        "updatedAt": datetime.now(tz=timezone.utc).isoformat(),
        "releasePeriod": release_period,
        "sourceFile": publication_date,
    }


def _transform_inflation(
    current_doc: dict[str, Any],
    extract: CPIExtract,
    source_url: str = "",
) -> dict[str, Any]:
    """
    Apply CPI values to the existing inflation.json document shape.
    Touches cpi-headline and food-inflation only. repo-rate and
    annual-cpi-avg are never read or written by this function — the
    deep-copy at the top preserves them exactly as they were in
    current_doc, and _apply_qlfs_rate_map() only mutates stats whose id
    is a key in rate_map.
    """
    doc = copy.deepcopy(current_doc)
    rate_map = {
        _CPI_HEADLINE_STAT_ID: extract.cpi_headline,
        _CPI_FOOD_STAT_ID: extract.food_inflation,
    }
    # Defensive assertion (IMPLEMENTATION-SPEC-CPI.md §11 item 6): a
    # rate_map key outside _CPI_OWNED_STAT_IDS would be a programming
    # error in this function itself, not a data problem — fail loudly
    # here rather than relying solely on _assert_cpi_ownership_boundary()
    # to catch it downstream.
    assert set(rate_map.keys()) <= _CPI_OWNED_STAT_IDS, (
        "_transform_inflation() built a rate_map with a key outside "
        "_CPI_OWNED_STAT_IDS — this is a programming error, not a data "
        "problem (IMPLEMENTATION-SPEC-CPI.md §11 item 6)."
    )
    _apply_qlfs_rate_map(
        doc, rate_map,
        release_period=extract.release_period,
        publication_date=extract.publication_date,
    )
    _update_cpi_meta(
        doc,
        release_period=extract.release_period,
        publication_date=extract.publication_date,
    )
    return doc


def _cpi_values_changed(
    current_doc: dict[str, Any],
    extract: CPIExtract,
) -> bool:
    """
    Return True if either cpi_headline or food_inflation differs from the
    current on-disk rawValue for that stat (or the stat has no rawValue
    yet), mirroring the QLFS/GDP flows' pre-transform "did anything
    actually change" check. Must be computed BEFORE _transform_inflation()
    runs, since that function always refreshes _meta["last_verified"].
    """
    current_headline = _get_current_stat_rate(current_doc, _CPI_HEADLINE_STAT_ID)
    current_food = _get_current_stat_rate(current_doc, _CPI_FOOD_STAT_ID)

    if current_headline is None or abs(current_headline - extract.cpi_headline) > 0.001:
        return True
    if current_food is None or abs(current_food - extract.food_inflation) > 0.001:
        return True
    return False


# ---------------------------------------------------------------------------
# Population validation helpers (Phase 4)
# ---------------------------------------------------------------------------


def _validate_population_total(value: float, label: str) -> list[str]:
    """
    Validate a total-population figure (in millions) is a plausible
    magnitude in _POPULATION_PLAUSIBLE_RANGE. South Africa's population is
    well-documented and slow-moving, so this band is deliberately wide —
    IMPLEMENTATION-SPEC-POPULATION.md §10 / §14 item 3.
    """
    low, high = _POPULATION_PLAUSIBLE_RANGE
    if not (low <= value <= high):
        return [
            f"{label} population value {value}M is outside the plausible "
            f"[{low}, {high}] million range."
        ]
    return []


def _validate_annual_label(label: str) -> list[str]:
    """Validate a period label matches the bare-year format (dataset-analysis.md-style RULES)."""
    if not _ANNUAL_LABEL_RE.match(label):
        return [f"Release period {label!r} does not match the expected 'YYYY' format."]
    return []


def _check_yoy_jump(
    current: float | None,
    new: float,
    stat_label: str,
    threshold: float = _POPULATION_JUMP_WARNING_THRESHOLD,
) -> str | None:
    """
    Return a human-readable anomaly warning if the year-over-year percentage
    change for ``stat_label`` exceeds ``threshold`` percentage points.
    Parallel to _check_qoq_jump(), scaled for an annual rather than
    quarterly/monthly cadence — IMPLEMENTATION-SPEC-POPULATION.md §10.

    This is a review aid, not a hard failure. It is deliberately expected
    to fire on the one-time production correction run (§9/§14 item 7) —
    the human reviewer should expect and accept that flag on that specific
    run, not treat it as a parser bug.
    """
    if current is None or current == 0:
        return None
    delta_pct = (new - current) / current * 100
    if abs(delta_pct) > threshold:
        sign = "+" if delta_pct > 0 else ""
        return (
            f"ANOMALY: {stat_label} moved {sign}{delta_pct:.1f}% year-over-year "
            f"({current:.1f} → {new:.1f}), exceeding the {threshold:.1f}% "
            "review threshold. Verify against the official release before approving."
        )
    return None


# ---------------------------------------------------------------------------
# Population source guard (Phase 4) — this milestone's genuinely new
# complication.
#
# The direct enforcement of population.yaml's source_guard_required /
# source_guard_domain — the mechanism that makes the World-Bank-sourced
# data-integrity bug (SA-Data-Hub-Dataset-Sourcing-Plan.md §9) structurally
# impossible to reintroduce via this flow. See IMPLEMENTATION-SPEC-
# POPULATION.md §7.3.
# ---------------------------------------------------------------------------


def _assert_population_source_guard(url: str) -> list[str]:
    """
    Hard-fail check: ``url`` must resolve to the statssa.gov.za domain (or
    a documented Stats SA subdomain).

    This is the direct enforcement of population.yaml's
    source_guard_required — the mechanism that makes the World-Bank-
    sourced data-integrity bug (SA-Data-Hub-Dataset-Sourcing-Plan.md §9)
    structurally impossible to reintroduce via this flow. Checked twice
    (URL-level, before download, and extract-level, before staging) —
    deliberately redundant rather than relying on a single check
    (IMPLEMENTATION-SPEC-POPULATION.md §7.3).
    """
    parsed = urllib.parse.urlparse(url)
    domain = parsed.netloc.lower()
    if not (domain == _POPULATION_SOURCE_GUARD_DOMAIN
            or domain.endswith(f".{_POPULATION_SOURCE_GUARD_DOMAIN}")):
        return [
            f"Population source guard violation: resolved publication URL "
            f"{url!r} does not originate from "
            f"{_POPULATION_SOURCE_GUARD_DOMAIN!r} (population.yaml's "
            f"source_guard_domain). This is the exact class of error that "
            f"produced the current data-integrity bug in population.json "
            f"(SA-Data-Hub-Dataset-Sourcing-Plan.md §9) and must hard-fail, "
            f"not warn."
        ]
    return []


# ---------------------------------------------------------------------------
# Population ownership boundary (Phase 4)
#
# Structurally identical to _assert_cpi_ownership_boundary(), scoped to
# population.json's non-owned stats (population-urban,
# population-median-age) — no new design, reuse of the proven pattern.
# See IMPLEMENTATION-SPEC-POPULATION.md §7.4.
# ---------------------------------------------------------------------------


def _assert_population_ownership_boundary(
    previous_doc: dict[str, Any],
    proposed_doc: dict[str, Any],
) -> list[str]:
    """
    Deep-compare every statistics[] entry in ``proposed_doc`` whose id is
    NOT in _POPULATION_OWNED_STAT_IDS against the corresponding entry in
    ``previous_doc`` (matched by id). Returns a list of violation messages
    — empty if population-urban and population-median-age (and any other
    non-owned stat) are byte-for-byte identical between the two documents,
    and the set of stat IDs present is unchanged.
    """
    violations: list[str] = []

    previous_by_id = {s.get("id"): s for s in previous_doc.get("statistics", [])}
    proposed_by_id = {s.get("id"): s for s in proposed_doc.get("statistics", [])}

    if set(previous_by_id.keys()) != set(proposed_by_id.keys()):
        violations.append(
            "Population ownership boundary violation: the set of stat "
            "IDs in population.json changed (added/removed a stat) — "
            f"previous={sorted(k for k in previous_by_id if k is not None)!r}, "
            f"proposed={sorted(k for k in proposed_by_id if k is not None)!r}."
        )

    for stat_id, prev_stat in previous_by_id.items():
        if stat_id in _POPULATION_OWNED_STAT_IDS:
            continue
        proposed_stat = proposed_by_id.get(stat_id)
        if proposed_stat != prev_stat:
            violations.append(
                f"Population ownership boundary violation: non-owned stat "
                f"{stat_id!r} changed. This milestone's code MUST NOT "
                f"modify population-urban or population-median-age "
                f"(IMPLEMENTATION-SPEC-POPULATION.md §3/§7.4)."
            )

    return violations


# ---------------------------------------------------------------------------
# Population transform helpers (Phase 4)
#
# _apply_qlfs_rate_map() is NOT reused here (unlike CPI) — population-total
# has a materially different display shape (bare-year labels, millions
# magnitude with a dual millions/raw-headcount convention) — see
# IMPLEMENTATION-SPEC-POPULATION.md §8.2.
# ---------------------------------------------------------------------------


def _apply_population_total_point(
    doc: dict[str, Any],
    new_value_millions: float,
    *,
    release_period: str,
    publication_date: str,
) -> None:
    """
    Seed-or-append-or-revise the population-total stat's series, following
    the same structural pattern as _apply_qlfs_rate_map() and
    _apply_gdp_growth_points(), but formatted for a millions-of-people
    magnitude with a bare-year label instead of a percentage rate with a
    quarterly/monthly label.

    Note: ``rawValue`` is stored as a raw headcount (e.g. 64000000),
    matching the existing on-disk convention in population.json, while the
    series stores millions (e.g. 64.0) — this preserves the existing, if
    slightly inconsistent, dual convention already present in the
    production file rather than "fixing" it as an uninvited refactor
    (IMPLEMENTATION-SPEC-POPULATION.md §8.2).
    """
    for stat in doc.get("statistics", []):
        if stat.get("id") != _POPULATION_TOTAL_STAT_ID:
            continue
        prev_value = stat.get("rawValue")
        change_pct = None
        if isinstance(prev_value, (int, float)) and prev_value:
            change_pct = round(
                (new_value_millions * 1_000_000 - prev_value) / prev_value * 100, 1
            )
        stat["value"] = f"{new_value_millions:.1f}M"
        stat["rawValue"] = round(new_value_millions * 1_000_000)
        if change_pct is not None:
            stat["change"] = change_pct
        stat["changeLabel"] = f"from {int(release_period) - 1}"
        stat["trend"] = "up" if (change_pct or 0) >= 0 else "down"
        stat["lastUpdated"] = publication_date
        if isinstance(stat.get("source"), dict):
            stat["source"]["publicationDate"] = publication_date
            stat["source"]["publicationName"] = f"Mid-Year Population Estimates {release_period}"

        # Series: same seed-or-append-or-revise pattern as
        # _apply_qlfs_rate_map(), storing the value in millions (matching
        # the existing series' unit) not the raw headcount. Handles both
        # a missing "series" key AND an existing-but-empty series list —
        # setdefault() alone only covers the former.
        series_list = stat.setdefault("series", [])
        if not series_list:
            series_list.append(
                {"name": "Population (millions)", "unit": "million", "data": []}
            )
        series = series_list[0]
        existing_labels = {pt["label"] for pt in series.get("data", [])}
        if release_period not in existing_labels:
            series.setdefault("data", []).append(
                {"label": release_period, "value": round(new_value_millions, 1)}
            )
        else:
            for pt in series["data"]:
                if pt["label"] == release_period:
                    pt["value"] = round(new_value_millions, 1)
                    break
    return


def _update_population_meta(
    doc: dict[str, Any],
    *,
    release_period: str,
    publication_date: str,
    source_url: str = "",
) -> None:
    """
    Update _meta.last_verified, _meta.lastUpdated, _meta.source_url, and
    _meta.automation. Follows _update_qlfs_meta()'s pattern (overwrite,
    not narrow-scope like _update_cpi_meta()) because — unlike
    inflation.json — population.json's _meta block describes only Stats SA
    content; there is no SARB-owned prose sharing the file that a blanket
    _meta rewrite would clobber (IMPLEMENTATION-SPEC-POPULATION.md §8.3).
    """
    if "_meta" not in doc:
        doc["_meta"] = {}
    doc["_meta"]["last_verified"] = date.today().isoformat()
    doc["_meta"]["lastUpdated"] = publication_date
    if source_url:
        doc["_meta"]["source_url"] = source_url
    doc["_meta"]["automation"] = {
        "updatedBy": "statssa-adapter/population",
        "updatedAt": datetime.now(tz=timezone.utc).isoformat(),
        "releasePeriod": release_period,
        "sourceFile": publication_date,
    }


def _transform_population(
    current_doc: dict[str, Any],
    extract: PopulationExtract,
    source_url: str = "",
) -> dict[str, Any]:
    """
    Apply the MYPE total-population value to the existing population.json
    document shape. Touches population-total only. population-urban and
    population-median-age are never read or written — the deep-copy at
    the top preserves them exactly.
    """
    doc = copy.deepcopy(current_doc)
    _apply_population_total_point(
        doc, extract.total_population_millions,
        release_period=extract.release_period,
        publication_date=extract.publication_date,
    )
    _update_population_meta(
        doc,
        release_period=extract.release_period,
        publication_date=extract.publication_date,
        source_url=source_url,
    )
    return doc


def _population_value_changed(current_doc: dict[str, Any], extract: PopulationExtract) -> bool:
    """
    Mirrors _cpi_values_changed()'s pre-transform check, computed before
    _transform_population() runs (which always refreshes
    _meta.last_verified).
    """
    for stat in current_doc.get("statistics", []):
        if stat.get("id") == _POPULATION_TOTAL_STAT_ID:
            current_raw = stat.get("rawValue")
            new_raw = round(extract.total_population_millions * 1_000_000)
            return current_raw is None or abs(current_raw - new_raw) > 1000
    return True


# ---------------------------------------------------------------------------
# QLFS validation helpers (Phase 2)
# ---------------------------------------------------------------------------

_QUARTERLY_LABEL_RE = re.compile(r"^Q[1-4] \d{4}$")


def _validate_percentage(value: float, label: str) -> list[str]:
    """Validate a rate is a plausible percentage in [0, 100]."""
    if not (0.0 <= value <= 100.0):
        return [f"{label} value {value} is outside the plausible [0, 100] range."]
    return []


def _validate_quarterly_label(label: str) -> list[str]:
    """Validate a period label matches the quarterly format (dataset-analysis.md RULES)."""
    if not _QUARTERLY_LABEL_RE.match(label):
        return [f"Release period {label!r} does not match the expected 'Q[1-4] YYYY' format."]
    return []


def _check_qoq_jump(
    current: float | None,
    new: float,
    stat_label: str,
    threshold: float = _QOQ_JUMP_WARNING_THRESHOLD,
) -> str | None:
    """
    Return a human-readable anomaly warning if the quarter-over-quarter
    jump for ``stat_label`` exceeds ``threshold`` percentage points.

    This is a review aid, not a hard failure — per
    IMPLEMENTATION-SPEC-STATSSA-PHASE2.md §6 item 3, an anomaly flag is
    for the human approver's attention, not an automatic rejection.
    """
    if current is None:
        return None
    delta = new - current
    if abs(delta) > threshold:
        sign = "+" if delta > 0 else ""
        return (
            f"ANOMALY: {stat_label} moved {sign}{delta:.1f}pp "
            f"({current:.1f}% → {new:.1f}%), exceeding the {threshold:.1f}pp "
            "review threshold. Verify against the official release before approving."
        )
    return None


# ---------------------------------------------------------------------------
# QLFS transform helpers (Phase 2)
#
# Each transform follows the exact pattern SARBAdapter._transform_interest_rates()
# already established: deep-copy the current document, update only
# rate-bearing fields (value, rawValue, change, changeLabel, trend,
# lastUpdated, source.publicationDate), seed-or-append the series history,
# and update the shared _meta block. Structural/protected fields (id,
# categoryId, unit, description, etc.) are never touched here.
# ---------------------------------------------------------------------------


def _read_current_dataset_json(path: Path) -> dict[str, Any]:
    """Read an existing dataset JSON file. Returns {} if absent/unreadable."""
    if not path.exists():
        log.warning("%s not found — treating as empty.", path)
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError) as exc:
        log.error("Cannot read %s: %s", path, exc)
        return {}


def _get_current_stat_rate(doc: dict[str, Any], stat_id: str) -> float | None:
    """Return the rawValue for a given stat_id from a dataset document."""
    for stat in doc.get("statistics", []):
        if stat.get("id") == stat_id:
            try:
                return float(stat["rawValue"])
            except (KeyError, TypeError, ValueError):
                return None
    return None


def _determine_qlfs_trend(current_rate: float | None, new_rate: float) -> str:
    """Determine trend direction from previous to new rate."""
    if current_rate is None:
        return "stable"
    if new_rate > current_rate:
        return "up"
    if new_rate < current_rate:
        return "down"
    return "stable"


def _apply_qlfs_rate_map(
    doc: dict[str, Any],
    rate_map: dict[str, float],
    *,
    release_period: str,
    publication_date: str,
) -> None:
    """
    Apply ``rate_map`` (stat_id -> new rawValue) to ``doc["statistics"]``
    in place, following the shared seed-or-append/in-place-revision series
    logic already established for SARB.

    Only stat IDs present in ``rate_map`` are touched; every other stat in
    the document (and every other field on a touched stat) is left as-is.
    """
    for stat in doc.get("statistics", []):
        stat_id = stat.get("id")
        if stat_id not in rate_map:
            continue

        new_rate = rate_map[stat_id]
        prev_rate = _get_current_stat_rate({"statistics": [stat]}, stat_id)

        change = 0.0
        if prev_rate is not None:
            change = round(new_rate - prev_rate, 1)

        # Previous label, for the changeLabel text — the last existing
        # series point before we mutate it below (if any).
        prev_label = None
        existing_series = stat.get("series") or []
        if existing_series and existing_series[0].get("data"):
            prev_label = existing_series[0]["data"][-1].get("label")

        stat["value"] = f"{new_rate:.1f}%"
        stat["rawValue"] = new_rate
        stat["change"] = change
        stat["changeLabel"] = f"from {prev_label}" if prev_label else release_period
        stat["trend"] = _determine_qlfs_trend(prev_rate, new_rate)
        stat["lastUpdated"] = publication_date
        if isinstance(stat.get("source"), dict):
            stat["source"]["publicationDate"] = publication_date

        # Series: seed-or-append-or-revise, exactly as
        # SARBAdapter._transform_interest_rates() does.
        if not stat.get("series"):
            stat["series"] = [
                {
                    "name": stat.get("title", stat_id),
                    "unit": stat.get("unit", "%"),
                    "data": [{"label": release_period, "value": new_rate}],
                }
            ]
            log.debug(
                "Seeded first series point %r -> %.1f for %s",
                release_period, new_rate, stat_id,
            )
            continue

        for series in stat.get("series", []):
            existing_labels = {pt["label"] for pt in series.get("data", [])}
            if release_period not in existing_labels:
                series.setdefault("data", []).append(
                    {"label": release_period, "value": new_rate}
                )
                log.debug(
                    "Appended series point %r -> %.1f for %s",
                    release_period, new_rate, stat_id,
                )
            else:
                for pt in series["data"]:
                    if pt["label"] == release_period:
                        if abs(pt["value"] - new_rate) > 0.001:
                            log.info(
                                "Updating existing series point %r: %.1f -> %.1f for %s",
                                release_period, pt["value"], new_rate, stat_id,
                            )
                            pt["value"] = new_rate
                        break


def _update_qlfs_meta(
    doc: dict[str, Any],
    *,
    release_period: str,
    publication_date: str,
    source_url: str,
) -> None:
    """Update the shared _meta block, following the SARB _meta pattern."""
    if "_meta" not in doc:
        doc["_meta"] = {}
    today_str = date.today().isoformat()
    doc["_meta"]["last_verified"] = today_str
    doc["_meta"]["lastUpdated"] = publication_date
    doc["_meta"]["source_url"] = source_url
    doc["_meta"]["automation"] = {
        "updatedBy": "statssa-adapter/qlfs",
        "updatedAt": datetime.now(tz=timezone.utc).isoformat(),
        "releasePeriod": release_period,
        "sourceFile": source_url,
    }


def _transform_unemployment(
    current_doc: dict[str, Any],
    extract: QLFSExtract,
    source_url: str,
) -> dict[str, Any]:
    """Apply QLFS values to the existing unemployment.json document shape."""
    doc = copy.deepcopy(current_doc)
    rate_map = {"unemployment-national": extract.unemployment_rate}
    _apply_qlfs_rate_map(
        doc, rate_map,
        release_period=extract.release_period,
        publication_date=extract.publication_date,
    )
    _update_qlfs_meta(
        doc,
        release_period=extract.release_period,
        publication_date=extract.publication_date,
        source_url=source_url,
    )
    return doc


def _transform_youth_unemployment(
    current_doc: dict[str, Any],
    extract: QLFSExtract,
    source_url: str,
) -> dict[str, Any]:
    """Apply QLFS values to the existing youth-unemployment.json document shape."""
    doc = copy.deepcopy(current_doc)
    rate_map = {
        "youth-unemployment-narrow": extract.youth_unemployment_narrow,
        "youth-unemployment-1524": extract.youth_unemployment_1524,
        "youth-unemployment-expanded": extract.youth_unemployment_expanded,
        # NEET's existing series uses annual labels (2019..2025) in the
        # current on-disk data, but the live QLFS release reports this
        # figure alongside the other youth indicators each quarter (per
        # SA-Data-Hub-Dataset-Sourcing-Plan.md's Q1 2026 findings). Going
        # forward this stat is updated on the same quarterly cadence as
        # the rest of the youth family, labeled by quarter like the others
        # — flagged here rather than silently assumed, since it's a change
        # from the historical annual-label convention in this one series.
        "youth-neet-rate": extract.neet_rate,
    }
    _apply_qlfs_rate_map(
        doc, rate_map,
        release_period=extract.release_period,
        publication_date=extract.publication_date,
    )
    _update_qlfs_meta(
        doc,
        release_period=extract.release_period,
        publication_date=extract.publication_date,
        source_url=source_url,
    )
    return doc


def _transform_labour_force(
    current_doc: dict[str, Any],
    extract: QLFSExtract,
    source_url: str,
) -> dict[str, Any]:
    """Apply QLFS values to the existing labour-force.json document shape."""
    doc = copy.deepcopy(current_doc)
    rate_map = {
        "lfpr-overall": extract.lfpr_overall,
        "female-labour-participation": extract.lfpr_female,
    }
    _apply_qlfs_rate_map(
        doc, rate_map,
        release_period=extract.release_period,
        publication_date=extract.publication_date,
    )
    _update_qlfs_meta(
        doc,
        release_period=extract.release_period,
        publication_date=extract.publication_date,
        source_url=source_url,
    )
    return doc


_QLFS_TRANSFORMS = {
    "unemployment": _transform_unemployment,
    "youth-unemployment": _transform_youth_unemployment,
    "labour-force": _transform_labour_force,
}

# stat_id -> dataset_id, used for per-dataset "did anything actually
# change" comparisons against the current on-disk JSON.
_QLFS_STAT_TO_DATASET: dict[str, str] = {
    "unemployment-national": "unemployment",
    "youth-unemployment-narrow": "youth-unemployment",
    "youth-unemployment-1524": "youth-unemployment",
    "youth-unemployment-expanded": "youth-unemployment",
    "youth-neet-rate": "youth-unemployment",
    "lfpr-overall": "labour-force",
    "female-labour-participation": "labour-force",
}


# ---------------------------------------------------------------------------
# Adapter class
# ---------------------------------------------------------------------------


class StatsSAAdapter(BaseAdapter):
    """
    Adapter for Statistics South Africa (statssa.gov.za).

    Covers QLFS, GDP, CPI, MYPE, GHS, Census, and the Municipal Fact Sheet.

    Phase 1 implements:
      - Real release detection for the QLFS family (ETag/content-hash watch)
      - ``fetch_and_apply()`` for the QLFS family:
          1. Discover latest QLFS release
          2. Locate the official publication (Excel or PDF)
          3. Download it
          4. Archive the raw file
          5. Record version metadata
          6. Generate a run report

    All other datasets remain Phase A stubs in ``check_for_updates``.
    """

    source_id = "statssa"
    display_name = "Statistics South Africa"
    priority = 10   # Run first — largest number of datasets
    version = "0.6.0"  # 0.1.0 Phase A stub -> 0.2.0 Phase 1 (download) -> 0.3.0 Phase 2 (QLFS parse/transform/stage) -> 0.4.0 Phase 3a (GDP parse/transform/stage) -> 0.4.1 WAF Tier 1 (header hardening + direct-URL fallback probe) -> 0.5.0 Phase 3b (CPI parse/transform/stage + ownership boundary) -> 0.6.0 Phase 4 (Population parse/transform/stage + source guard)

    def __init__(
        self,
        config: Any,
        source_config: Any = None,
    ) -> None:
        super().__init__(config, source_config)
        # Run-level cache: QLFS hub check is shared across all three QLFS datasets.
        # Avoids three identical HTTP fetches per runner invocation.
        self._qlfs_check_cache: DatasetCheckResult | None = None
        # Run-level cache for the GDP hub check (parallels _qlfs_check_cache).
        self._gdp_check_cache: DatasetCheckResult | None = None
        # Run-level cache for the CPI hub check (parallels _gdp_check_cache).
        self._cpi_check_cache: DatasetCheckResult | None = None
        # Run-level cache for the Population hub check (parallels _cpi_check_cache).
        self._population_check_cache: DatasetCheckResult | None = None

    def validate_config(self) -> list[str]:
        """
        Validate Stats SA adapter configuration.

        Stats SA requires no API key or credentials — all data is publicly
        available via the release hub.  We validate that:
        - The automation raw archive directory is writable (Phase B dependency)
        - The source config, if present, has the expected base_url
        """
        errors: list[str] = []

        # Check that the archive directory can be created
        archive_root = self.config.raw_archive_dir
        try:
            archive_root.mkdir(parents=True, exist_ok=True)
        except OSError as exc:
            errors.append(
                f"Cannot create raw archive directory {archive_root}: {exc}"
            )

        # Validate base URL if configured
        if self.source_config.base_url and not self.source_config.base_url.startswith(
            "https://"
        ):
            errors.append(
                f"source_config.base_url should start with 'https://', "
                f"got: {self.source_config.base_url!r}"
            )

        return errors

    def datasets(self) -> list[str]:
        """Return all datasets this adapter is responsible for."""
        return _STATSSA_DATASETS

    # ------------------------------------------------------------------
    # check_for_updates — Phase 1: real ETag watch for QLFS; stubs for rest
    # ------------------------------------------------------------------

    def check_for_updates(
        self,
        dataset_id: str,
        dataset_config: DatasetConfig | None,
    ) -> DatasetCheckResult:
        """
        Check whether a new release is available for ``dataset_id``.

        Phase 1 implements a real ETag/content-hash watch for the QLFS family.
        All other datasets retain the Phase A stub behaviour.
        """
        # QLFS family — Phase 1: real release hub check
        # All three QLFS datasets share one hub fetch per run (cached)
        # to avoid 3 identical HTTP calls and ensure consistent status.
        if dataset_id in _QLFS_FAMILY:
            if self._qlfs_check_cache is None:
                self._qlfs_check_cache = self._check_qlfs(dataset_id, dataset_config)
            # Return a copy with the correct dataset_id (status/message shared)
            cached = self._qlfs_check_cache
            return DatasetCheckResult(
                dataset_id=dataset_id,
                status=cached.status,
                message=cached.message,
                latest_period=cached.latest_period,
                current_period=cached.current_period,
                source_url=cached.source_url,
                notes=cached.notes,
            )

        # GDP — Phase 3a: real ETag/content-hash check against the P0441
        # release hub, mirroring the QLFS caching pattern above.
        if dataset_id == "gdp":
            if self._gdp_check_cache is None:
                self._gdp_check_cache = self._check_gdp(dataset_id, dataset_config)
            return self._gdp_check_cache

        # ---------- Phase A stubs for remaining datasets ----------

        # Static datasets — Census and municipalities
        if dataset_id in _STATIC_DATASETS:
            return DatasetCheckResult(
                dataset_id=dataset_id,
                status="unknown",
                message=(
                    f"[Phase A] {dataset_id} is a static/erratum-watch dataset. "
                    "Phase B will implement a lightweight page-hash check on the "
                    "Stats SA release hub."
                ),
                notes=(
                    "No action needed unless Stats SA publishes an erratum. "
                    "municipalities.json is current (verified 2026-06-04). "
                    "census.json is correct until ~2032."
                ),
            )

        # CPI — Phase 3b: real ETag/content-hash check against the P0141
        # release hub, mirroring the QLFS/GDP caching pattern above. Repo
        # rate (SARB-owned, same inflation.json file) is never touched by
        # this check — see IMPLEMENTATION-SPEC-CPI.md §0.1/§7.
        if dataset_id == "inflation":
            if self._cpi_check_cache is None:
                self._cpi_check_cache = self._check_cpi(dataset_id, dataset_config)
            return self._cpi_check_cache

        # Population — Phase 4: real ETag/content-hash check against the
        # P0302 release hub, mirroring the QLFS/GDP/CPI caching pattern
        # above. This replaces the Phase A advisory stub entirely — the
        # stub's warning content (data-integrity bug, source guard
        # requirement) is now enforced in code by
        # _assert_population_source_guard(), not left as dead advisory
        # text (IMPLEMENTATION-SPEC-POPULATION.md §12 item 13).
        if dataset_id == "population":
            if self._population_check_cache is None:
                self._population_check_cache = self._check_population(dataset_id, dataset_config)
            return self._population_check_cache

        # Housing (GHS component)
        if dataset_id == "housing":
            return DatasetCheckResult(
                dataset_id=dataset_id,
                status="unknown",
                message=(
                    "[Phase A] Housing P0318 (GHS component). "
                    "Phase B will implement annual GHS Excel download. "
                    "Census-baseline component is static. "
                    "Pending: confirm whether GHS 2024/2025 has updated the "
                    "three tracked indicators (piped water, electricity, formal dwellings)."
                ),
                source_url="https://www.statssa.gov.za/?page_id=1854&PPN=P0318",
            )

        # Fallback
        return DatasetCheckResult(
            dataset_id=dataset_id,
            status="unknown",
            message=f"[Phase A] No check implemented yet for {dataset_id}.",
        )

    # ------------------------------------------------------------------
    # QLFS-specific check (Phase 1)
    # ------------------------------------------------------------------

    def _check_qlfs(
        self,
        dataset_id: str,
        dataset_config: DatasetConfig | None,
    ) -> DatasetCheckResult:
        """
        Real release detection for the QLFS family.

        Performs an ETag/content-hash check against the P0211 release hub.
        Returns ``update_available`` if the page has changed since the last
        known hash, ``up_to_date`` if it hasn't, or ``error`` on failure.

        The *same* check result is appropriate for all three QLFS datasets
        (unemployment, youth-unemployment, labour-force) — they all come
        from the same release.
        """
        client = _build_http_client(self.source_config)
        previous_hash = self._load_qlfs_previous_hash()

        self._log.info(
            "Checking QLFS release hub: %s (previous_hash=%s…)",
            _QLFS_HUB_URL,
            previous_hash[:8] if previous_hash else "none",
        )

        try:
            changed, response = with_retry(
                lambda: client.etag_check(
                    _QLFS_HUB_URL,
                    previous_sha256=previous_hash,
                ),
                policy=WATCH_POLICY,
                label="QLFS release hub ETag check",
            )
        except AutomationHTTPError as exc:
            return DatasetCheckResult(
                dataset_id=dataset_id,
                status="error",
                message=f"QLFS release hub returned HTTP {exc.status}: {exc.reason}",
                source_url=_QLFS_HUB_URL,
            )
        except Exception as exc:
            return DatasetCheckResult(
                dataset_id=dataset_id,
                status="error",
                message=f"Failed to check QLFS release hub: {exc}",
                source_url=_QLFS_HUB_URL,
            )
            
        # WAF check
        if response.body:
            body_text = response.body.decode("utf-8", errors="replace")
            if "_Incapsula_Resource" in body_text or "incapsula" in body_text.lower():
                self._log.error("WAF challenge detected on QLFS release hub")
                # Tier 1 fallback (IMPLEMENTATION-SPEC-STATSSA-WAF.md §6.1
                # step 2): the hub being WAF-blocked does not mean the
                # direct publication-URL path is also blocked. Probe it
                # before giving up. This is additive only: it does not
                # change the WAF-marker detection above, and does not
                # touch fetch_and_apply()'s own, separate use of the same
                # probing functions.
                q, y = _determine_current_qlfs_quarter()
                try:
                    fallback_url = _probe_qlfs_publication_url(client, q, y)
                except Exception as probe_exc:
                    self._log.warning(
                        "QLFS direct-URL fallback probe raised: %s", probe_exc
                    )
                    fallback_url = None
                if fallback_url:
                    self._log.info(
                        "QLFS hub WAF-blocked, but a direct publication URL "
                        "is reachable: %s", fallback_url,
                    )
                    return DatasetCheckResult(
                        dataset_id=dataset_id,
                        status="unknown",
                        message=(
                            "WAF_BLOCKED: Incapsula WAF challenge detected on "
                            "the QLFS release hub, but a direct publication "
                            "URL is reachable. This is a probe-based signal, "
                            "not a hub-diff signal — it confirms a candidate "
                            "file can be fetched, not that a new release "
                            "exists. Run fetch_and_apply() to attempt "
                            "download."
                        ),
                        source_url=_QLFS_HUB_URL,
                        notes=f"Direct-URL fallback probe found: {fallback_url}",
                    )
                return DatasetCheckResult(
                    dataset_id=dataset_id,
                    status="error",
                    message="WAF_BLOCKED: Incapsula WAF challenge detected. Cannot check for updates.",
                    source_url=_QLFS_HUB_URL,
                )

        # Parse release period from the page even if unchanged (for context)
        release_period = _extract_release_period(response.body)

        if not changed:
            self._log.info(
                "QLFS release hub unchanged (sha256=%s…)", previous_hash[:8]
            )
            return DatasetCheckResult(
                dataset_id=dataset_id,
                status="up_to_date",
                message=(
                    "QLFS P0211 release hub page is unchanged since last check. "
                    "No new publication detected."
                ),
                latest_period=release_period or "unknown",
                source_url=_QLFS_HUB_URL,
            )

        # Page has changed — save new hash and report update available
        self._save_qlfs_hash(response.content_sha256)
        self._log.info(
            "QLFS release hub changed — new sha256=%s… — update likely available",
            response.content_sha256[:8],
        )

        # Attempt to find an Excel link for richer reporting
        excel_url = _extract_excel_url(response.body)

        return DatasetCheckResult(
            dataset_id=dataset_id,
            status="update_available",
            message=(
                "QLFS P0211 release hub page has changed. "
                "A new QLFS publication is likely available."
            ),
            latest_period=release_period or "unknown (check release hub)",
            current_period="Q4 2025",
            source_url=_QLFS_HUB_URL,
            notes=(
                f"Excel workbook URL detected: {excel_url or 'not found — check hub manually'}. "
                "Run fetch_and_apply() to download and archive the workbook."
            ),
        )

    # ------------------------------------------------------------------
    # Hash persistence helpers (lightweight state for ETag check)
    # ------------------------------------------------------------------

    def _qlfs_hash_path(self) -> Path:
        """Return the path to the stored QLFS hub content hash."""
        return self.config.report_dir / "versions" / "qlfs_hub.sha256"

    def _load_qlfs_previous_hash(self) -> str:
        """Return the last-known SHA-256 of the QLFS hub page, or ''."""
        p = self._qlfs_hash_path()
        if not p.exists():
            return ""
        try:
            return p.read_text(encoding="utf-8").strip()
        except OSError:
            return ""

    def _save_qlfs_hash(self, sha256: str) -> None:
        """Persist the new SHA-256 of the QLFS hub page."""
        p = self._qlfs_hash_path()
        try:
            p.parent.mkdir(parents=True, exist_ok=True)
            p.write_text(sha256, encoding="utf-8")
        except OSError as exc:
            self._log.warning("Cannot save QLFS hub hash to %s: %s", p, exc)

    # ------------------------------------------------------------------
    # GDP-specific check (Phase 3a)
    # ------------------------------------------------------------------

    def _check_gdp(
        self,
        dataset_id: str,
        dataset_config: DatasetConfig | None,
    ) -> DatasetCheckResult:
        """
        Real release detection for GDP (P0441).

        Performs an ETag/content-hash check against the P0441 release hub,
        mirroring _check_qlfs() structurally exactly.
        """
        client = _build_http_client(self.source_config)
        previous_hash = self._load_gdp_previous_hash()

        self._log.info(
            "Checking GDP release hub: %s (previous_hash=%s…)",
            _GDP_HUB_URL,
            previous_hash[:8] if previous_hash else "none",
        )

        try:
            changed, response = with_retry(
                lambda: client.etag_check(
                    _GDP_HUB_URL,
                    previous_sha256=previous_hash,
                ),
                policy=WATCH_POLICY,
                label="GDP release hub ETag check",
            )
        except AutomationHTTPError as exc:
            return DatasetCheckResult(
                dataset_id=dataset_id,
                status="error",
                message=f"GDP release hub returned HTTP {exc.status}: {exc.reason}",
                source_url=_GDP_HUB_URL,
            )
        except Exception as exc:
            return DatasetCheckResult(
                dataset_id=dataset_id,
                status="error",
                message=f"Failed to check GDP release hub: {exc}",
                source_url=_GDP_HUB_URL,
            )

        # WAF check — same guard as _check_qlfs(), copied rather than
        # factored into a shared helper (see IMPLEMENTATION-SPEC-GDP.md §8:
        # consolidating this is a future refactor, out of scope here since
        # it would touch _check_qlfs()'s code path).
        if response.body:
            body_text = response.body.decode("utf-8", errors="replace")
            if "_Incapsula_Resource" in body_text or "incapsula" in body_text.lower():
                self._log.error("WAF challenge detected on GDP release hub")
                # Tier 1 fallback (IMPLEMENTATION-SPEC-STATSSA-WAF.md §6.1
                # step 2), mirroring _check_qlfs()'s fallback exactly —
                # copied rather than shared, for the same reason as the WAF
                # scan above.
                q, y = _determine_current_gdp_quarter()
                try:
                    fallback_url = _probe_gdp_publication_url(client, q, y)
                except Exception as probe_exc:
                    self._log.warning(
                        "GDP direct-URL fallback probe raised: %s", probe_exc
                    )
                    fallback_url = None
                if fallback_url:
                    self._log.info(
                        "GDP hub WAF-blocked, but a direct publication URL "
                        "is reachable: %s", fallback_url,
                    )
                    return DatasetCheckResult(
                        dataset_id=dataset_id,
                        status="unknown",
                        message=(
                            "WAF_BLOCKED: Incapsula WAF challenge detected on "
                            "the GDP release hub, but a direct publication "
                            "URL is reachable. This is a probe-based signal, "
                            "not a hub-diff signal — it confirms a candidate "
                            "file can be fetched, not that a new release "
                            "exists. Run fetch_and_apply() to attempt "
                            "download."
                        ),
                        source_url=_GDP_HUB_URL,
                        notes=f"Direct-URL fallback probe found: {fallback_url}",
                    )
                return DatasetCheckResult(
                    dataset_id=dataset_id,
                    status="error",
                    message="WAF_BLOCKED: Incapsula WAF challenge detected. Cannot check for updates.",
                    source_url=_GDP_HUB_URL,
                )

        release_period = _extract_release_period(response.body)

        if not changed:
            self._log.info(
                "GDP release hub unchanged (sha256=%s…)", previous_hash[:8]
            )
            return DatasetCheckResult(
                dataset_id=dataset_id,
                status="up_to_date",
                message=(
                    "GDP P0441 release hub page is unchanged since last check. "
                    "No new publication detected."
                ),
                latest_period=release_period or "unknown",
                source_url=_GDP_HUB_URL,
            )

        self._save_gdp_hash(response.content_sha256)
        self._log.info(
            "GDP release hub changed — new sha256=%s… — update likely available",
            response.content_sha256[:8],
        )

        excel_url = _extract_excel_url(response.body)

        return DatasetCheckResult(
            dataset_id=dataset_id,
            status="update_available",
            message=(
                "GDP P0441 release hub page has changed. "
                "A new GDP publication is likely available."
            ),
            latest_period=release_period or "unknown (check release hub)",
            source_url=_GDP_HUB_URL,
            notes=(
                f"Excel workbook URL detected: {excel_url or 'not found — check hub manually'}. "
                "Run fetch_and_apply() to download and archive the workbook."
            ),
        )

    # ------------------------------------------------------------------
    # GDP hash persistence helpers (parallel to the QLFS ones above)
    # ------------------------------------------------------------------

    def _gdp_hash_path(self) -> Path:
        """Return the path to the stored GDP hub content hash."""
        return self.config.report_dir / "versions" / "gdp_hub.sha256"

    def _load_gdp_previous_hash(self) -> str:
        """Return the last-known SHA-256 of the GDP hub page, or ''."""
        p = self._gdp_hash_path()
        if not p.exists():
            return ""
        try:
            return p.read_text(encoding="utf-8").strip()
        except OSError:
            return ""

    def _save_gdp_hash(self, sha256: str) -> None:
        """Persist the new SHA-256 of the GDP hub page."""
        p = self._gdp_hash_path()
        try:
            p.parent.mkdir(parents=True, exist_ok=True)
            p.write_text(sha256, encoding="utf-8")
        except OSError as exc:
            self._log.warning("Cannot save GDP hub hash to %s: %s", p, exc)

    # ------------------------------------------------------------------
    # CPI-specific check (Phase 3b)
    # ------------------------------------------------------------------

    def _check_cpi(
        self,
        dataset_id: str,
        dataset_config: DatasetConfig | None,
    ) -> DatasetCheckResult:
        """
        Real release detection for CPI (P0141).

        Performs an ETag/content-hash check against the P0141 release hub,
        mirroring _check_qlfs() / _check_gdp() structurally exactly. Never
        reads or writes repo-rate (the SARB-owned stat in the same
        inflation.json file) — this method only ever inspects the P0141
        release hub page.
        """
        client = _build_http_client(self.source_config)
        previous_hash = self._load_cpi_previous_hash()

        self._log.info(
            "Checking CPI release hub: %s (previous_hash=%s…)",
            _CPI_HUB_URL,
            previous_hash[:8] if previous_hash else "none",
        )

        try:
            changed, response = with_retry(
                lambda: client.etag_check(
                    _CPI_HUB_URL,
                    previous_sha256=previous_hash,
                ),
                policy=WATCH_POLICY,
                label="CPI release hub ETag check",
            )
        except AutomationHTTPError as exc:
            return DatasetCheckResult(
                dataset_id=dataset_id,
                status="error",
                message=f"CPI release hub returned HTTP {exc.status}: {exc.reason}",
                source_url=_CPI_HUB_URL,
            )
        except Exception as exc:
            return DatasetCheckResult(
                dataset_id=dataset_id,
                status="error",
                message=f"Failed to check CPI release hub: {exc}",
                source_url=_CPI_HUB_URL,
            )

        # WAF check — same guard as _check_qlfs()/_check_gdp(), copied
        # rather than factored into a shared helper (same reasoning as
        # _check_gdp()'s copy of _check_qlfs()'s guard — consolidating
        # this is a future refactor, out of scope here).
        if response.body:
            body_text = response.body.decode("utf-8", errors="replace")
            if "_Incapsula_Resource" in body_text or "incapsula" in body_text.lower():
                self._log.error("WAF challenge detected on CPI release hub")
                # Tier 1 fallback (IMPLEMENTATION-SPEC-STATSSA-WAF.md §6.1
                # step 2), mirroring _check_qlfs()'s/_check_gdp()'s
                # fallback exactly — copied rather than shared, for the
                # same reason as the WAF scan above.
                m, y = _determine_current_cpi_month()
                try:
                    fallback_url = _probe_cpi_publication_url(client, m, y)
                except Exception as probe_exc:
                    self._log.warning(
                        "CPI direct-URL fallback probe raised: %s", probe_exc
                    )
                    fallback_url = None
                if fallback_url:
                    self._log.info(
                        "CPI hub WAF-blocked, but a direct publication URL "
                        "is reachable: %s", fallback_url,
                    )
                    return DatasetCheckResult(
                        dataset_id=dataset_id,
                        status="unknown",
                        message=(
                            "WAF_BLOCKED: Incapsula WAF challenge detected on "
                            "the CPI release hub, but a direct publication "
                            "URL is reachable. This is a probe-based signal, "
                            "not a hub-diff signal — it confirms a candidate "
                            "file can be fetched, not that a new release "
                            "exists. Run fetch_and_apply() to attempt "
                            "download."
                        ),
                        source_url=_CPI_HUB_URL,
                        notes=f"Direct-URL fallback probe found: {fallback_url}",
                    )
                return DatasetCheckResult(
                    dataset_id=dataset_id,
                    status="error",
                    message="WAF_BLOCKED: Incapsula WAF challenge detected. Cannot check for updates.",
                    source_url=_CPI_HUB_URL,
                )

        release_period = _extract_release_period(response.body)

        if not changed:
            self._log.info(
                "CPI release hub unchanged (sha256=%s…)", previous_hash[:8]
            )
            return DatasetCheckResult(
                dataset_id=dataset_id,
                status="up_to_date",
                message=(
                    "CPI P0141 release hub page is unchanged since last check. "
                    "No new publication detected."
                ),
                latest_period=release_period or "unknown",
                source_url=_CPI_HUB_URL,
            )

        self._save_cpi_hash(response.content_sha256)
        self._log.info(
            "CPI release hub changed — new sha256=%s… — update likely available",
            response.content_sha256[:8],
        )

        excel_url = _extract_excel_url(response.body)

        return DatasetCheckResult(
            dataset_id=dataset_id,
            status="update_available",
            message=(
                "CPI P0141 release hub page has changed. "
                "A new CPI publication is likely available."
            ),
            latest_period=release_period or "unknown (check release hub)",
            source_url=_CPI_HUB_URL,
            notes=(
                f"Excel workbook URL detected: {excel_url or 'not found — check hub manually'}. "
                "Run fetch_and_apply() to download and archive the workbook."
            ),
        )

    # ------------------------------------------------------------------
    # CPI hash persistence helpers (parallel to the QLFS/GDP ones above)
    # ------------------------------------------------------------------

    def _cpi_hash_path(self) -> Path:
        """Return the path to the stored CPI hub content hash."""
        return self.config.report_dir / "versions" / "cpi_hub.sha256"

    def _load_cpi_previous_hash(self) -> str:
        """Return the last-known SHA-256 of the CPI hub page, or ''."""
        p = self._cpi_hash_path()
        if not p.exists():
            return ""
        try:
            return p.read_text(encoding="utf-8").strip()
        except OSError:
            return ""

    def _save_cpi_hash(self, sha256: str) -> None:
        """Persist the new SHA-256 of the CPI hub page."""
        p = self._cpi_hash_path()
        try:
            p.parent.mkdir(parents=True, exist_ok=True)
            p.write_text(sha256, encoding="utf-8")
        except OSError as exc:
            self._log.warning("Cannot save CPI hub hash to %s: %s", p, exc)

    # ------------------------------------------------------------------
    # Population-specific check (Phase 4)
    # ------------------------------------------------------------------

    def _check_population(
        self,
        dataset_id: str,
        dataset_config: DatasetConfig | None,
    ) -> DatasetCheckResult:
        """
        Real release detection for Population (MYPE, P0302).

        Performs an ETag/content-hash check against the P0302 release hub,
        mirroring _check_qlfs() / _check_gdp() / _check_cpi() structurally
        exactly. Never reads or writes population-urban or
        population-median-age (the Census-2022-owned stats in the same
        population.json file) — this method only ever inspects the P0302
        release hub page.
        """
        client = _build_http_client(self.source_config)
        previous_hash = self._load_population_previous_hash()

        self._log.info(
            "Checking Population release hub: %s (previous_hash=%s…)",
            _POPULATION_HUB_URL,
            previous_hash[:8] if previous_hash else "none",
        )

        try:
            changed, response = with_retry(
                lambda: client.etag_check(
                    _POPULATION_HUB_URL,
                    previous_sha256=previous_hash,
                ),
                policy=WATCH_POLICY,
                label="Population release hub ETag check",
            )
        except AutomationHTTPError as exc:
            return DatasetCheckResult(
                dataset_id=dataset_id,
                status="error",
                message=f"Population release hub returned HTTP {exc.status}: {exc.reason}",
                source_url=_POPULATION_HUB_URL,
            )
        except Exception as exc:
            return DatasetCheckResult(
                dataset_id=dataset_id,
                status="error",
                message=f"Failed to check Population release hub: {exc}",
                source_url=_POPULATION_HUB_URL,
            )

        # WAF check — same guard as _check_qlfs()/_check_gdp()/_check_cpi(),
        # copied rather than factored into a shared helper (same reasoning
        # as _check_cpi()'s copy of _check_gdp()'s guard — consolidating
        # this is a future refactor, out of scope here).
        if response.body:
            body_text = response.body.decode("utf-8", errors="replace")
            if "_Incapsula_Resource" in body_text or "incapsula" in body_text.lower():
                self._log.error("WAF challenge detected on Population release hub")
                # Tier 1 fallback (IMPLEMENTATION-SPEC-STATSSA-WAF.md §6.1
                # step 2), mirroring _check_qlfs()'s/_check_gdp()'s/
                # _check_cpi()'s fallback exactly — copied rather than
                # shared, for the same reason as the WAF scan above.
                y = _determine_current_population_year()
                try:
                    fallback_url = _probe_population_publication_url(client, y)
                except Exception as probe_exc:
                    self._log.warning(
                        "Population direct-URL fallback probe raised: %s", probe_exc
                    )
                    fallback_url = None
                if fallback_url:
                    self._log.info(
                        "Population hub WAF-blocked, but a direct publication "
                        "URL is reachable: %s", fallback_url,
                    )
                    return DatasetCheckResult(
                        dataset_id=dataset_id,
                        status="unknown",
                        message=(
                            "WAF_BLOCKED: Incapsula WAF challenge detected on "
                            "the Population release hub, but a direct "
                            "publication URL is reachable. This is a "
                            "probe-based signal, not a hub-diff signal — it "
                            "confirms a candidate file can be fetched, not "
                            "that a new release exists. Run fetch_and_apply() "
                            "to attempt download."
                        ),
                        source_url=_POPULATION_HUB_URL,
                        notes=f"Direct-URL fallback probe found: {fallback_url}",
                    )
                return DatasetCheckResult(
                    dataset_id=dataset_id,
                    status="error",
                    message="WAF_BLOCKED: Incapsula WAF challenge detected. Cannot check for updates.",
                    source_url=_POPULATION_HUB_URL,
                )

        release_period = _extract_release_period(response.body)

        if not changed:
            self._log.info(
                "Population release hub unchanged (sha256=%s…)", previous_hash[:8]
            )
            return DatasetCheckResult(
                dataset_id=dataset_id,
                status="up_to_date",
                message=(
                    "Population P0302 release hub page is unchanged since "
                    "last check. No new publication detected."
                ),
                latest_period=release_period or "unknown",
                source_url=_POPULATION_HUB_URL,
            )

        self._save_population_hash(response.content_sha256)
        self._log.info(
            "Population release hub changed — new sha256=%s… — update likely available",
            response.content_sha256[:8],
        )

        excel_url = _extract_excel_url(response.body)

        return DatasetCheckResult(
            dataset_id=dataset_id,
            status="update_available",
            message=(
                "Population P0302 release hub page has changed. "
                "A new MYPE publication is likely available."
            ),
            latest_period=release_period or "unknown (check release hub)",
            source_url=_POPULATION_HUB_URL,
            notes=(
                f"Excel workbook URL detected: {excel_url or 'not found — check hub manually'}. "
                "Run fetch_and_apply() to download and archive the workbook."
            ),
        )

    # ------------------------------------------------------------------
    # Population hash persistence helpers (parallel to the CPI ones above)
    # ------------------------------------------------------------------

    def _population_hash_path(self) -> Path:
        """Return the path to the stored Population hub content hash."""
        return self.config.report_dir / "versions" / "population_hub.sha256"

    def _load_population_previous_hash(self) -> str:
        """Return the last-known SHA-256 of the Population hub page, or ''."""
        p = self._population_hash_path()
        if not p.exists():
            return ""
        try:
            return p.read_text(encoding="utf-8").strip()
        except OSError:
            return ""

    def _save_population_hash(self, sha256: str) -> None:
        """Persist the new SHA-256 of the Population hub page."""
        p = self._population_hash_path()
        try:
            p.parent.mkdir(parents=True, exist_ok=True)
            p.write_text(sha256, encoding="utf-8")
        except OSError as exc:
            self._log.warning("Cannot save Population hub hash to %s: %s", p, exc)

    # ------------------------------------------------------------------
    # fetch_and_apply — Phase 1: QLFS discovery + download + archive
    # ------------------------------------------------------------------

    def fetch_and_apply(
        self,
        *,
        dry_run: bool = False,
        run_id: str = "",
    ) -> dict[str, Any]:
        """
        Manually-invoked utility, reachable via ``python -m automation.runner
        --apply`` (any adapter defining ``fetch_and_apply()`` is invoked by
        ``runner.py`` — see ``automation/docs/developer-guide.md``).

        Phase 2: Discover, download, archive, parse, transform, validate,
        and **stage** the latest QLFS publication for all three QLFS-family
        datasets. This method never writes directly to any dataset JSON —
        every candidate document is written to the staging area and
        recorded as a ``pending`` version entry; reaching production
        requires a separate, explicit ``--approve`` then ``--promote`` per
        dataset, exactly as already enforced for ``interest-rates.json``
        (Work Item 4, ``automation/core/staging.py`` / ``promote.py``).

        Steps
        -----
        1. Fetch the QLFS P0211 release hub page (with retry).
        2. Locate the official publication link (Excel or PDF) by direct URL probing.
        3. Download the publication file (with retry, large-file timeout).
        4. Archive the raw file with checksum + manifest.
        5. Parse the workbook (``parse_qlfs_workbook()``); fail loudly (no
           staging, no version entries) if the file isn't an Excel
           workbook or the expected indicators can't be located.
        6. For each of the three QLFS datasets whose values actually
           changed: transform, validate (range + label format +
           protected-field diff), flag (not hard-fail) large
           quarter-over-quarter jumps, then stage the candidate document
           and record one ``pending`` version entry.
        7. Return an execution result dict for the caller / report system.

        If none of the three datasets' values differ from what's already
        on disk, this returns ``status="no_change"`` with no staging and
        no new version entries — the same no-op-on-no-change behaviour
        already established for SARB.

        Parameters
        ----------
        dry_run:
            When True, downloads and parses the file but does NOT write to
            disk (no archive write, no staging, no version entries). All
            other steps execute normally (for verification purposes).
        run_id:
            Correlation ID from the runner, injected into version entries.

        Returns
        -------
        dict
            Keys:
              status            — "ok" | "no_change" | "no_publication_found" | "error"
              hub_url           — QLFS release hub URL checked
              file_url          — discovered publication URL, or None
              release_period    — detected quarter label, e.g. "Q1 2026"
              archive_path      — path where file was saved, or None
              sha256            — checksum of the downloaded file
              file_size_bytes   — size of the downloaded file
              version_ids       — list of version IDs staged (one per changed dataset)
              dry_run           — echo of the dry_run flag
              notes             — human-readable summary
              errors            — list of error strings
              gdp               — (Phase 3a, additive) nested dict describing the
                                   independent GDP flow's own outcome: status
                                   ("ok" | "no_change" | "no_publication_found" |
                                   "error"), hub_url, file_url, release_period,
                                   archive_path, sha256, file_size_bytes,
                                   version_id (singular — gdp.json is one
                                   dataset, not three), notes, errors. A GDP
                                   failure never changes the top-level `status`
                                   above, which continues to describe the QLFS
                                   run only (see IMPLEMENTATION-SPEC-GDP.md §9).
              cpi               — (Phase 3b, additive) nested dict describing
                                   the independent CPI flow's own outcome,
                                   same shape as `gdp` above (status,
                                   hub_url, file_url, release_period,
                                   archive_path, sha256, file_size_bytes,
                                   version_id — singular, inflation.json is
                                   one dataset — notes, errors). Touches
                                   cpi-headline and food-inflation only;
                                   repo-rate and annual-cpi-avg are never
                                   read or written. A CPI failure never
                                   changes the top-level `status` above or
                                   `result["gdp"]` (see
                                   IMPLEMENTATION-SPEC-CPI.md §14).
              population        — (Phase 4, additive) nested dict describing
                                   the independent Population flow's own
                                   outcome, same shape as `gdp`/`cpi` above
                                   (status, hub_url, file_url,
                                   release_period, archive_path, sha256,
                                   file_size_bytes, version_id — singular,
                                   population.json is one dataset — notes,
                                   errors). Touches population-total only;
                                   population-urban and
                                   population-median-age are never read or
                                   written. A Population failure never
                                   changes the top-level `status` above or
                                   `result["gdp"]`/`result["cpi"]` (see
                                   IMPLEMENTATION-SPEC-POPULATION.md §11).
        """
        result: dict[str, Any] = {
            "status": "error",
            "hub_url": _QLFS_HUB_URL,
            "file_url": None,
            "release_period": "",
            "archive_path": None,
            "sha256": None,
            "file_size_bytes": None,
            "version_ids": [],
            "dry_run": dry_run,
            "notes": "",
            "errors": [],
        }
        # GDP (Phase 3a) — additive key, see §9.1 of IMPLEMENTATION-SPEC-GDP.md.
        # Every key above keeps its existing meaning, describing the QLFS run
        # only. This nested dict is the sole place GDP's own outcome lives.
        result["gdp"] = {
            "status": "error",
            "hub_url": _GDP_HUB_URL,
            "file_url": None,
            "release_period": "",
            "archive_path": None,
            "sha256": None,
            "file_size_bytes": None,
            "version_id": None,
            "notes": "",
            "errors": [],
        }
        # CPI (Phase 3b) — additive key, mirroring result["gdp"]'s shape
        # exactly. Every key above keeps its existing meaning, describing
        # the QLFS run only. This nested dict is the sole place CPI's own
        # outcome lives. See IMPLEMENTATION-SPEC-CPI.md §14.
        result["cpi"] = {
            "status": "error",
            "hub_url": _CPI_HUB_URL,
            "file_url": None,
            "release_period": "",
            "archive_path": None,
            "sha256": None,
            "file_size_bytes": None,
            "version_id": None,
            "notes": "",
            "errors": [],
        }
        # Population (Phase 4) — additive key, mirroring result["cpi"]'s
        # shape exactly. Every key above keeps its existing meaning,
        # describing the QLFS run only. This nested dict is the sole place
        # Population's own outcome lives. See IMPLEMENTATION-SPEC-
        # POPULATION.md §11.
        result["population"] = {
            "status": "error",
            "hub_url": _POPULATION_HUB_URL,
            "file_url": None,
            "release_period": "",
            "archive_path": None,
            "sha256": None,
            "file_size_bytes": None,
            "version_id": None,
            "notes": "",
            "errors": [],
        }

        client = _build_http_client(self.source_config)

        # ----------------------------------------------------------
        # Step 1 & 2: Discover publication link
        # ----------------------------------------------------------
        self._log.info("Fetching QLFS release hub: %s", _QLFS_HUB_URL)
        try:
            file_url, release_period, hub_html = _discover_qlfs_excel(client)
        except AutomationHTTPError as exc:
            msg = f"QLFS release hub returned HTTP {exc.status}: {exc.reason}"
            result["errors"].append(msg)
            self._log.error(msg)
            return result
        except Exception as exc:
            msg = f"Failed to fetch QLFS release hub: {exc}"
            result["errors"].append(msg)
            self._log.error(msg)
            return result
            
        result["release_period"] = release_period or ""
        self._log.info(
            "QLFS release hub fetched — detected period: %r",
            release_period or "(not detected)",
        )

        result["file_url"] = file_url

        if file_url is None:
            msg = (
                "No publication link found for QLFS. "
                "Direct probes failed (likely no standard naming) and HTML scrape "
                "failed (likely WAF blocked). "
                f"Hub URL: {_QLFS_HUB_URL}"
            )
            self._log.warning(msg)
            result["status"] = "no_publication_found"
            result["notes"] = (
                "Could not locate the QLFS publication file. "
                "The Stats SA hub is WAF-protected, preventing scraping, and direct "
                "URL probing did not find a standard filename. Manual check required."
            )
            # Save the hub HTML to archive for manual inspection
            if not dry_run:
                try:
                    hub_dest, hub_sha = save_to_archive(
                        self.config.raw_archive_dir,
                        hub_html,
                        dataset_id="unemployment",  # canonical QLFS dataset
                        source_id=self.source_id,
                        suffix="_hub.html",
                    )
                    self._log.info(
                        "Hub HTML archived for inspection → %s", hub_dest
                    )
                except Exception as archive_exc:
                    self._log.warning(
                        "Could not archive hub HTML: %s", archive_exc
                    )
            return result

        self._log.info("Located publication: %s", file_url)

        # ----------------------------------------------------------
        # Step 3: Download the publication
        # ----------------------------------------------------------
        self._log.info("Downloading QLFS publication from %s …", file_url)
        try:
            file_bytes = with_retry(
                lambda: _download_publication(client, file_url),  # type: ignore[arg-type]
                policy=STATSSA_POLICY,
                label=f"QLFS file download ({file_url})",
            )
        except AutomationHTTPError as exc:
            msg = (
                f"QLFS file download failed — HTTP {exc.status}: {exc.reason} "
                f"(URL: {file_url})"
            )
            result["errors"].append(msg)
            self._log.error(msg)
            return result
        except Exception as exc:
            msg = f"QLFS file download failed: {exc}"
            result["errors"].append(msg)
            self._log.error(msg)
            return result

        result["file_size_bytes"] = len(file_bytes)
        self._log.info(
            "Downloaded QLFS publication: %d bytes", len(file_bytes)
        )

        # ----------------------------------------------------------
        # Step 4: Archive the raw file
        # ----------------------------------------------------------
        file_ext = Path(urllib.parse.urlparse(file_url).path).suffix.lower() or ".bin"
        if not dry_run:
            try:
                archive_dest, sha256 = save_to_archive(
                    self.config.raw_archive_dir,
                    file_bytes,
                    dataset_id="unemployment",  # canonical QLFS dataset slug
                    source_id=self.source_id,
                    suffix=file_ext,
                )
                result["archive_path"] = portable_archive_path(
                    self.config.raw_archive_dir, archive_dest
                )
                result["sha256"] = sha256
                self._log.info(
                    "QLFS file archived → %s (sha256=%s…, %d bytes)",
                    archive_dest,
                    sha256[:8],
                    len(file_bytes),
                )
            except Exception as exc:
                msg = f"Archive write failed: {exc}"
                result["errors"].append(msg)
                self._log.error(msg)
                # Non-fatal — continue to version recording
        else:
            from automation.core.files import sha256_of_bytes
            sha256 = sha256_of_bytes(file_bytes)
            result["sha256"] = sha256
            self._log.info(
                "[DRY RUN] Would archive %d bytes (sha256=%s…)",
                len(file_bytes),
                sha256[:8],
            )

        sha256 = result["sha256"] or ""

        # ----------------------------------------------------------
        # Step 4.5: Parse the workbook (Phase 2)
        #
        # PDF parsing is explicitly out of scope for this phase (see
        # IMPLEMENTATION-SPEC-STATSSA-PHASE2.md §2/§5). If the URL probe in
        # Step 2 fell back to the statistical-release PDF (no Excel
        # candidate responded), the correct behaviour is the same
        # status="error" path used for every other unparseable-file case
        # here — not a best-effort PDF scrape.
        # ----------------------------------------------------------
        if file_ext not in (".xlsx", ".xls"):
            msg = (
                f"Downloaded QLFS publication is not an Excel workbook "
                f"(extension {file_ext!r}, url={file_url}). PDF parsing is "
                f"explicitly out of scope for this phase — falling back to "
                f"the manual-review path (Track B) rather than guessing at "
                f"PDF-extracted values."
            )
            result["errors"].append(msg)
            result["status"] = "error"
            self._log.error(msg)
            return result

        try:
            extract = parse_qlfs_workbook(file_bytes)
        except Exception as exc:
            msg = f"QLFS workbook parse failed: {exc}"
            result["errors"].append(msg)
            result["status"] = "error"
            self._log.error(msg)
            return result

        result["release_period"] = extract.release_period
        self._log.info(
            "QLFS workbook parsed OK — release period %s, "
            "unemployment=%.1f%%, youth-narrow=%.1f%%, youth-1524=%.1f%%, "
            "youth-expanded=%.1f%%, NEET=%.1f%%, LFPR=%.1f%%, LFPR-female=%.1f%%",
            extract.release_period,
            extract.unemployment_rate,
            extract.youth_unemployment_narrow,
            extract.youth_unemployment_1524,
            extract.youth_unemployment_expanded,
            extract.neet_rate,
            extract.lfpr_overall,
            extract.lfpr_female,
        )

        # ----------------------------------------------------------
        # Step 5: Transform, validate, diff, stage — one pass per QLFS
        # output dataset. No dataset JSON is ever written directly; the
        # only outputs of this step are staged candidates + pending
        # version entries, exactly as SARBAdapter.fetch_and_apply() does
        # for interest-rates.json.
        # ----------------------------------------------------------
        current_docs = {
            ds_id: _read_current_dataset_json(path)
            for ds_id, path in _QLFS_DATASET_JSON.items()
        }

        new_values = {
            "unemployment-national": extract.unemployment_rate,
            "youth-unemployment-narrow": extract.youth_unemployment_narrow,
            "youth-unemployment-1524": extract.youth_unemployment_1524,
            "youth-unemployment-expanded": extract.youth_unemployment_expanded,
            "youth-neet-rate": extract.neet_rate,
            "lfpr-overall": extract.lfpr_overall,
            "female-labour-participation": extract.lfpr_female,
        }

        # Per-dataset "did anything actually change" flags, mirroring
        # SARBAdapter's repo_changed/prime_changed pattern — and, per
        # IMPLEMENTATION-SPEC-STATSSA-PHASE2.md §10 acceptance criterion 6,
        # the basis for producing status="no_change" with zero side
        # effects when nothing has moved.
        dataset_changed: dict[str, bool] = {ds_id: False for ds_id in _QLFS_FAMILY}
        for stat_id, new_val in new_values.items():
            ds_id = _QLFS_STAT_TO_DATASET[stat_id]
            current_val = _get_current_stat_rate(current_docs[ds_id], stat_id)
            if current_val is None or abs(new_val - current_val) > 0.05:
                dataset_changed[ds_id] = True

        if not any(dataset_changed.values()):
            result["status"] = "no_change"
            result["notes"] = (
                f"No change: {extract.release_period} values already match "
                f"unemployment.json / youth-unemployment.json / labour-force.json."
            )
            self._log.info("No change detected — all three QLFS JSON files are already current.")
            # Still worth refreshing the hub hash baseline below.
        else:
            version_ids: list[str] = []
            for ds_id in sorted(_QLFS_FAMILY):
                if not dataset_changed[ds_id]:
                    continue  # this dataset's own values are unchanged; leave it alone

                current_doc = current_docs[ds_id]
                updated_doc = _QLFS_TRANSFORMS[ds_id](current_doc, extract, file_url)

                # --- schema/range validation ---
                validation_errors: list[str] = []
                validation_errors += _validate_quarterly_label(extract.release_period)
                for stat_id, new_val in new_values.items():
                    if _QLFS_STAT_TO_DATASET[stat_id] != ds_id:
                        continue
                    validation_errors += _validate_percentage(new_val, stat_id)

                if validation_errors:
                    msg = f"Validation failed for {ds_id}: {'; '.join(validation_errors)}"
                    result["errors"].append(msg)
                    self._log.error(msg)
                    continue  # do not stage this dataset

                # --- protected-field check (reused unchanged) ---
                if current_doc:
                    violations = check_protected_fields(current_doc, updated_doc)
                    if violations:
                        msg = f"Protected field violation(s) for {ds_id}: {violations}"
                        result["errors"].append(msg)
                        result.setdefault("protected_field_violations", {})[ds_id] = violations
                        self._log.error(msg)
                        continue  # abort staging for this document only

                # --- anomaly / plausibility check (warning, not a hard fail) ---
                anomaly_notes: list[str] = []
                for stat_id, new_val in new_values.items():
                    if _QLFS_STAT_TO_DATASET[stat_id] != ds_id:
                        continue
                    current_val = _get_current_stat_rate(current_doc, stat_id)
                    warning = _check_qoq_jump(current_val, new_val, stat_id)
                    if warning:
                        anomaly_notes.append(warning)
                        self._log.warning(warning)

                # --- stage + version entry ---
                entry = new_version_entry(
                    dataset_id=ds_id,
                    source_id=self.source_id,
                    source_url=file_url,
                    sha256=sha256,
                    archive_path=result["archive_path"] or "",
                    adapter_version=self.version,
                    notes=(
                        f"QLFS {extract.release_period} parsed and staged. "
                        f"Hub URL: {_QLFS_HUB_URL}. "
                        + ("; ".join(anomaly_notes) + ". " if anomaly_notes else "")
                        + "Manual approval required before promotion."
                    ),
                    run_id=run_id,
                )

                if not dry_run:
                    try:
                        write_staged_dataset(
                            self.config.report_dir,
                            dataset_id=ds_id,
                            version_id=entry.version_id,
                            document=updated_doc,
                        )
                        save_version_entry(self.config.report_dir, entry)
                        self._log.info(
                            "Staged %s — version %s (status=pending)",
                            ds_id, entry.version_id,
                        )
                    except Exception as exc:
                        msg = f"Staging failed for {ds_id}: {exc}"
                        result["errors"].append(msg)
                        self._log.error(msg)
                        continue
                else:
                    self._log.info(
                        "[DRY RUN] Would stage %s — version %s", ds_id, entry.version_id,
                    )

                version_ids.append(entry.version_id)

            result["version_ids"] = version_ids

            if version_ids:
                result["status"] = "ok"
            else:
                # A change was detected but every candidate dataset failed
                # validation/protected-field checks — nothing was staged.
                result["status"] = "error"

        # ----------------------------------------------------------
        # Step 6: Compose result and update QLFS hub hash
        # ----------------------------------------------------------
        if not result.get("notes"):
            result["notes"] = (
                f"QLFS Phase 2 complete. "
                f"Release period: {extract.release_period}. "
                f"Publication file: {len(file_bytes):,} bytes. "
                f"Archive: {result['archive_path'] or '(dry-run)'}. "
                f"Version entries staged: {', '.join(result.get('version_ids', [])) or 'none'}. "
                f"No dataset JSON was written directly — staged candidates require "
                f"`--approve` then `--promote` per dataset."
            )

        # Update hub hash so the next check_for_updates call sees the new baseline
        if not dry_run:
            try:
                # Re-fetch to get a fresh hash (we already have the bytes)
                from automation.core.files import sha256_of_bytes
                hub_hash = sha256_of_bytes(hub_html)
                self._save_qlfs_hash(hub_hash)
                self._log.debug("QLFS hub hash updated: %s…", hub_hash[:8])
            except Exception as exc:
                self._log.warning("Could not update QLFS hub hash: %s", exc)

        # ============================================================
        # GDP flow (Phase 3a) — Excel discovery + download + archive +
        # parse + transform + validate + stage for gdp-growth only.
        # Runs after the QLFS flow above and is fully independent of it:
        # a GDP-specific failure is recorded only in result["gdp"] (and
        # appended to the top-level errors list for visibility) and never
        # changes the top-level `status`, which continues to reflect the
        # QLFS flow only (IMPLEMENTATION-SPEC-GDP.md §9.2). GDP writes no
        # dataset JSON directly — only gdp.json's staging area.
        # ============================================================
        gdp_client = _build_http_client(self.source_config)
        gdp_hub_html: bytes | None = None

        try:
            self._log.info("Fetching GDP release hub: %s", _GDP_HUB_URL)
            excel_url, release_period, gdp_hub_html = _discover_gdp_excel(gdp_client)
            result["gdp"]["release_period"] = release_period or ""
            result["gdp"]["file_url"] = excel_url

            if excel_url is None:
                msg = (
                    "No publication link found for GDP. Direct probes "
                    "failed (likely no standard naming) and HTML scrape "
                    f"failed (likely WAF blocked). Hub URL: {_GDP_HUB_URL}"
                )
                self._log.warning(msg)
                result["gdp"]["status"] = "no_publication_found"
                result["gdp"]["errors"].append(msg)
            else:
                self._log.info("Located GDP publication: %s", excel_url)
                gdp_file_bytes = with_retry(
                    lambda: _download_publication(gdp_client, excel_url),  # type: ignore[arg-type]
                    policy=STATSSA_POLICY,
                    label=f"GDP file download ({excel_url})",
                )
                result["gdp"]["file_size_bytes"] = len(gdp_file_bytes)

                gdp_file_ext = (
                    Path(urllib.parse.urlparse(excel_url).path).suffix.lower() or ".bin"
                )
                if not dry_run:
                    try:
                        gdp_archive_dest, gdp_sha256 = save_to_archive(
                            self.config.raw_archive_dir,
                            gdp_file_bytes,
                            dataset_id="gdp",
                            source_id=self.source_id,
                            suffix=gdp_file_ext,
                        )
                        result["gdp"]["archive_path"] = portable_archive_path(
                            self.config.raw_archive_dir, gdp_archive_dest
                        )
                        result["gdp"]["sha256"] = gdp_sha256
                        self._log.info(
                            "GDP file archived → %s (sha256=%s…, %d bytes)",
                            gdp_archive_dest, gdp_sha256[:8], len(gdp_file_bytes),
                        )
                    except Exception as archive_exc:
                        self._log.warning("GDP archive write failed: %s", archive_exc)
                else:
                    from automation.core.files import sha256_of_bytes
                    result["gdp"]["sha256"] = sha256_of_bytes(gdp_file_bytes)
                    self._log.info(
                        "[DRY RUN] Would archive GDP file (%d bytes, sha256=%s…)",
                        len(gdp_file_bytes), result["gdp"]["sha256"][:8],
                    )

                if gdp_file_ext not in (".xlsx", ".xls"):
                    msg = (
                        f"Downloaded GDP publication is not an Excel "
                        f"workbook (extension {gdp_file_ext!r}, "
                        f"url={excel_url}). PDF parsing is explicitly out "
                        f"of scope — falling back to the manual-review "
                        f"path (Track B) rather than guessing at "
                        f"PDF-extracted values."
                    )
                    result["gdp"]["status"] = "error"
                    result["gdp"]["errors"].append(msg)
                    self._log.error(msg)
                else:
                    extract = parse_gdp_workbook(gdp_file_bytes)
                    result["gdp"]["release_period"] = extract.release_period
                    self._log.info(
                        "GDP workbook parsed OK — release period %s, "
                        "%d growth point(s) found",
                        extract.release_period, len(extract.growth_points),
                    )

                    current_gdp_doc = _read_current_dataset_json(_GDP_DATASET_JSON)

                    range_errors: list[str] = []
                    for label, value in extract.growth_points:
                        range_errors += _validate_gdp_growth_rate(value, label)
                        range_errors += _validate_quarterly_label(label)

                    if range_errors:
                        result["gdp"]["status"] = "error"
                        result["gdp"]["errors"].extend(range_errors)
                        self._log.error(
                            "GDP validation failed: %s", "; ".join(range_errors)
                        )
                    elif not _gdp_growth_values_changed(
                        current_gdp_doc, extract.growth_points
                    ):
                        # Computed from the raw extracted points against
                        # the current on-disk document — BEFORE calling
                        # _transform_gdp() — since _transform_gdp() always
                        # refreshes _meta's timestamps, which would make a
                        # full-document equality check never match even on
                        # a genuine no-op run. Mirrors the QLFS flow's
                        # pre-transform dataset_changed check.
                        result["gdp"]["status"] = "no_change"
                        result["gdp"]["notes"] = (
                            f"No change: {extract.release_period} GDP "
                            f"growth value(s) already match gdp.json."
                        )
                        self._log.info(
                            "No change detected — gdp.json is already current."
                        )
                    else:
                        new_gdp_doc, gdp_warnings = _transform_gdp(
                            current_gdp_doc, extract, excel_url
                        )
                        protected_violations = check_protected_fields(
                            current_gdp_doc, new_gdp_doc
                        )
                        if protected_violations:
                            msg = f"Protected field violation: {protected_violations}"
                            result["gdp"]["status"] = "error"
                            result["gdp"]["errors"].append(msg)
                            self._log.error(msg)
                        else:
                            gdp_entry = new_version_entry(
                                dataset_id="gdp",
                                source_id=self.source_id,
                                source_url=excel_url,
                                sha256=result["gdp"]["sha256"] or "",
                                archive_path=result["gdp"]["archive_path"] or "",
                                adapter_version=self.version,
                                notes=(
                                    f"GDP {extract.release_period} parsed and "
                                    f"staged. Hub URL: {_GDP_HUB_URL}. "
                                    + ("; ".join(gdp_warnings) + ". " if gdp_warnings else "")
                                    + "Manual approval required before promotion."
                                ),
                                run_id=run_id,
                            )

                            if not dry_run:
                                try:
                                    write_staged_dataset(
                                        self.config.report_dir,
                                        dataset_id="gdp",
                                        version_id=gdp_entry.version_id,
                                        document=new_gdp_doc,
                                    )
                                    save_version_entry(self.config.report_dir, gdp_entry)
                                    self._log.info(
                                        "Staged gdp — version %s (status=pending)",
                                        gdp_entry.version_id,
                                    )
                                    result["gdp"]["version_id"] = gdp_entry.version_id
                                    result["version_ids"].append(gdp_entry.version_id)
                                    result["gdp"]["status"] = "ok"
                                    result["gdp"]["notes"] = (
                                        "; ".join(gdp_warnings) if gdp_warnings else ""
                                    )
                                except Exception as exc:
                                    msg = f"Staging failed for gdp: {exc}"
                                    result["gdp"]["status"] = "error"
                                    result["gdp"]["errors"].append(msg)
                                    self._log.error(msg)
                            else:
                                self._log.info(
                                    "[DRY RUN] Would stage gdp — version %s",
                                    gdp_entry.version_id,
                                )
                                result["gdp"]["version_id"] = gdp_entry.version_id
                                result["version_ids"].append(gdp_entry.version_id)
                                result["gdp"]["status"] = "ok"
                                result["gdp"]["notes"] = (
                                    "; ".join(gdp_warnings) if gdp_warnings else ""
                                )
        except ValueError as exc:
            msg = f"GDP workbook parse failed: {exc}"
            result["gdp"]["status"] = "error"
            result["gdp"]["errors"].append(msg)
            self._log.error(msg)
        except AutomationHTTPError as exc:
            msg = f"GDP release hub/file HTTP error: {exc.status} {exc.reason}"
            result["gdp"]["status"] = "error"
            result["gdp"]["errors"].append(msg)
            self._log.error(msg)
        except Exception as exc:
            msg = f"GDP fetch_and_apply failed: {exc}"
            result["gdp"]["status"] = "error"
            result["gdp"]["errors"].append(msg)
            self._log.error(msg)

        # Deliberate deviation from IMPLEMENTATION-SPEC-GDP.md §9.1's
        # illustrative pseudocode (explicitly non-literal): GDP errors are
        # NOT merged into the top-level `errors` list. §9.1's own
        # constraint is that every existing key "keeps its existing
        # meaning... required so that none of the six existing
        # fetch_and_apply-based tests need to change" — those six tests
        # assert on `result["errors"]` describing the QLFS run only (e.g.
        # `assert not result["errors"]` on a successful QLFS-only stage).
        # Merging GDP's own errors into that shared list would silently
        # change its meaning for callers who have relied on it describing
        # QLFS alone, and would break those six tests the moment GDP's
        # discovery doesn't find a publication (its default outcome
        # whenever a caller/test only mocks QLFS's discovery functions).
        # GDP's errors remain fully visible in result["gdp"]["errors"].
        #
        # Update GDP hub hash so the next check_for_updates call sees the
        # new baseline (mirrors the QLFS hash-update step above).
        if not dry_run and gdp_hub_html is not None:
            try:
                from automation.core.files import sha256_of_bytes
                gdp_hub_hash = sha256_of_bytes(gdp_hub_html)
                self._save_gdp_hash(gdp_hub_hash)
                self._log.debug("GDP hub hash updated: %s…", gdp_hub_hash[:8])
            except Exception as exc:
                self._log.warning("Could not update GDP hub hash: %s", exc)

        # ============================================================
        # CPI flow (Phase 3b) — Excel discovery + download + archive +
        # parse + transform + validate + stage for cpi-headline and
        # food-inflation only. Runs after the QLFS and GDP flows above and
        # is fully independent of both: a CPI-specific failure is recorded
        # only in result["cpi"] and never changes the top-level `status`,
        # which continues to reflect the QLFS flow only, and never affects
        # result["gdp"] (IMPLEMENTATION-SPEC-CPI.md §14). CPI writes no
        # dataset JSON directly — only inflation.json's staging area — and
        # never reads or writes repo-rate/annual-cpi-avg (§0.1/§0.2/§7).
        # ============================================================
        cpi_client = _build_http_client(self.source_config)
        cpi_hub_html: bytes | None = None

        try:
            self._log.info("Fetching CPI release hub: %s", _CPI_HUB_URL)
            excel_url, release_period, cpi_hub_html = _discover_cpi_excel(cpi_client)
            result["cpi"]["release_period"] = release_period or ""
            result["cpi"]["file_url"] = excel_url

            if excel_url is None:
                msg = (
                    "No publication link found for CPI. Direct probes "
                    "failed (likely no standard naming) and HTML scrape "
                    f"failed (likely WAF blocked). Hub URL: {_CPI_HUB_URL}"
                )
                self._log.warning(msg)
                result["cpi"]["status"] = "no_publication_found"
                result["cpi"]["errors"].append(msg)
            else:
                self._log.info("Located CPI publication: %s", excel_url)
                cpi_file_bytes = with_retry(
                    lambda: _download_publication(cpi_client, excel_url),  # type: ignore[arg-type]
                    policy=STATSSA_POLICY,
                    label=f"CPI file download ({excel_url})",
                )
                result["cpi"]["file_size_bytes"] = len(cpi_file_bytes)

                cpi_file_ext = (
                    Path(urllib.parse.urlparse(excel_url).path).suffix.lower() or ".bin"
                )
                if not dry_run:
                    try:
                        cpi_archive_dest, cpi_sha256 = save_to_archive(
                            self.config.raw_archive_dir,
                            cpi_file_bytes,
                            dataset_id="inflation",
                            source_id=self.source_id,
                            suffix=cpi_file_ext,
                        )
                        result["cpi"]["archive_path"] = portable_archive_path(
                            self.config.raw_archive_dir, cpi_archive_dest
                        )
                        result["cpi"]["sha256"] = cpi_sha256
                        self._log.info(
                            "CPI file archived → %s (sha256=%s…, %d bytes)",
                            cpi_archive_dest, cpi_sha256[:8], len(cpi_file_bytes),
                        )
                    except Exception as archive_exc:
                        self._log.warning("CPI archive write failed: %s", archive_exc)
                else:
                    from automation.core.files import sha256_of_bytes
                    result["cpi"]["sha256"] = sha256_of_bytes(cpi_file_bytes)
                    self._log.info(
                        "[DRY RUN] Would archive CPI file (%d bytes, sha256=%s…)",
                        len(cpi_file_bytes), result["cpi"]["sha256"][:8],
                    )

                if cpi_file_ext not in (".xlsx", ".xls"):
                    msg = (
                        f"Downloaded CPI publication is not an Excel "
                        f"workbook (extension {cpi_file_ext!r}, "
                        f"url={excel_url}). PDF parsing is explicitly out "
                        f"of scope — falling back to the manual-review "
                        f"path (Track B) rather than guessing at "
                        f"PDF-extracted values."
                    )
                    result["cpi"]["status"] = "error"
                    result["cpi"]["errors"].append(msg)
                    self._log.error(msg)
                else:
                    extract = parse_cpi_workbook(cpi_file_bytes)
                    result["cpi"]["release_period"] = extract.release_period
                    self._log.info(
                        "CPI workbook parsed OK — release period %s "
                        "(cpi_headline=%.1f, food_inflation=%.1f)",
                        extract.release_period,
                        extract.cpi_headline, extract.food_inflation,
                    )

                    current_cpi_doc = _read_current_dataset_json(_CPI_DATASET_JSON)

                    range_errors: list[str] = []
                    range_errors += _validate_cpi_rate(
                        extract.cpi_headline, _CPI_HEADLINE_STAT_ID
                    )
                    range_errors += _validate_cpi_rate(
                        extract.food_inflation, _CPI_FOOD_STAT_ID
                    )
                    range_errors += _validate_monthly_label(extract.release_period)

                    if range_errors:
                        result["cpi"]["status"] = "error"
                        result["cpi"]["errors"].extend(range_errors)
                        self._log.error(
                            "CPI validation failed: %s", "; ".join(range_errors)
                        )
                    elif not _cpi_values_changed(current_cpi_doc, extract):
                        # Computed from the raw extracted values against the
                        # current on-disk document — BEFORE calling
                        # _transform_inflation() — since that function
                        # always refreshes _meta["last_verified"], which
                        # would make a full-document equality check never
                        # match even on a genuine no-op run. Mirrors the
                        # QLFS/GDP flows' pre-transform dataset_changed
                        # check.
                        result["cpi"]["status"] = "no_change"
                        result["cpi"]["notes"] = (
                            f"No change: {extract.release_period} CPI "
                            f"value(s) already match inflation.json."
                        )
                        self._log.info(
                            "No change detected — inflation.json is already current."
                        )
                    else:
                        new_cpi_doc = _transform_inflation(
                            current_cpi_doc, extract, excel_url
                        )
                        protected_violations = check_protected_fields(
                            current_cpi_doc, new_cpi_doc
                        )
                        # The ownership boundary (IMPLEMENTATION-SPEC-CPI.md
                        # §7/§11 item 5) — the load-bearing check this
                        # milestone exists to add. Stricter than
                        # check_protected_fields(): it hard-fails on ANY
                        # difference (not just a protected-field change) in
                        # repo-rate or annual-cpi-avg, and on the stat-ID
                        # set changing at all.
                        ownership_violations = _assert_cpi_ownership_boundary(
                            current_cpi_doc, new_cpi_doc
                        )
                        all_violations = protected_violations + ownership_violations
                        if all_violations:
                            msg = f"Protected field violation: {all_violations}"
                            result["cpi"]["status"] = "error"
                            result["cpi"]["errors"].append(msg)
                            self._log.error(msg)
                        else:
                            cpi_jump_warnings: list[str] = []
                            prev_headline = _get_current_stat_rate(
                                current_cpi_doc, _CPI_HEADLINE_STAT_ID
                            )
                            prev_food = _get_current_stat_rate(
                                current_cpi_doc, _CPI_FOOD_STAT_ID
                            )
                            for prev_val, new_val, stat_label in (
                                (prev_headline, extract.cpi_headline, _CPI_HEADLINE_STAT_ID),
                                (prev_food, extract.food_inflation, _CPI_FOOD_STAT_ID),
                            ):
                                warning = _check_qoq_jump(
                                    prev_val, new_val, stat_label,
                                    threshold=_CPI_JUMP_WARNING_THRESHOLD,
                                )
                                if warning:
                                    cpi_jump_warnings.append(warning)

                            cpi_entry = new_version_entry(
                                dataset_id="inflation",
                                source_id=self.source_id,
                                source_url=excel_url,
                                sha256=result["cpi"]["sha256"] or "",
                                archive_path=result["cpi"]["archive_path"] or "",
                                adapter_version=self.version,
                                notes=(
                                    f"CPI {extract.release_period} parsed and "
                                    f"staged (cpi-headline, food-inflation "
                                    f"only). Hub URL: {_CPI_HUB_URL}. "
                                    + ("; ".join(cpi_jump_warnings) + ". " if cpi_jump_warnings else "")
                                    + "Manual approval required before promotion."
                                ),
                                run_id=run_id,
                            )

                            if not dry_run:
                                try:
                                    write_staged_dataset(
                                        self.config.report_dir,
                                        dataset_id="inflation",
                                        version_id=cpi_entry.version_id,
                                        document=new_cpi_doc,
                                    )
                                    save_version_entry(self.config.report_dir, cpi_entry)
                                    self._log.info(
                                        "Staged inflation — version %s (status=pending)",
                                        cpi_entry.version_id,
                                    )
                                    result["cpi"]["version_id"] = cpi_entry.version_id
                                    result["version_ids"].append(cpi_entry.version_id)
                                    result["cpi"]["status"] = "ok"
                                    result["cpi"]["notes"] = (
                                        "; ".join(cpi_jump_warnings) if cpi_jump_warnings else ""
                                    )
                                except Exception as exc:
                                    msg = f"Staging failed for inflation: {exc}"
                                    result["cpi"]["status"] = "error"
                                    result["cpi"]["errors"].append(msg)
                                    self._log.error(msg)
                            else:
                                self._log.info(
                                    "[DRY RUN] Would stage inflation — version %s",
                                    cpi_entry.version_id,
                                )
                                result["cpi"]["version_id"] = cpi_entry.version_id
                                result["version_ids"].append(cpi_entry.version_id)
                                result["cpi"]["status"] = "ok"
                                result["cpi"]["notes"] = (
                                    "; ".join(cpi_jump_warnings) if cpi_jump_warnings else ""
                                )
        except ValueError as exc:
            msg = f"CPI workbook parse failed: {exc}"
            result["cpi"]["status"] = "error"
            result["cpi"]["errors"].append(msg)
            self._log.error(msg)
        except AutomationHTTPError as exc:
            msg = f"CPI release hub/file HTTP error: {exc.status} {exc.reason}"
            result["cpi"]["status"] = "error"
            result["cpi"]["errors"].append(msg)
            self._log.error(msg)
        except Exception as exc:
            msg = f"CPI fetch_and_apply failed: {exc}"
            result["cpi"]["status"] = "error"
            result["cpi"]["errors"].append(msg)
            self._log.error(msg)

        # Deliberate deviation, identical in kind to the GDP flow's own
        # deviation from IMPLEMENTATION-SPEC-GDP.md §9.1 above: CPI errors
        # are NOT merged into the top-level `errors` list, for the same
        # reason (preserving every existing key's meaning for the six
        # pre-existing QLFS-only tests, and now also the GDP tests, which
        # assert on `result["errors"]` describing the QLFS run only). CPI's
        # errors remain fully visible in result["cpi"]["errors"].
        #
        # Update CPI hub hash so the next check_for_updates call sees the
        # new baseline (mirrors the QLFS/GDP hash-update steps above).
        if not dry_run and cpi_hub_html is not None:
            try:
                from automation.core.files import sha256_of_bytes
                cpi_hub_hash = sha256_of_bytes(cpi_hub_html)
                self._save_cpi_hash(cpi_hub_hash)
                self._log.debug("CPI hub hash updated: %s…", cpi_hub_hash[:8])
            except Exception as exc:
                self._log.warning("Could not update CPI hub hash: %s", exc)

        # ============================================================
        # Population flow (Phase 4) — Excel discovery + download + archive
        # + parse + transform + validate + stage for population-total
        # only. Runs after the QLFS/GDP/CPI flows above and is fully
        # independent of all three: a Population-specific failure is
        # recorded only in result["population"] and never changes the
        # top-level `status`, which continues to reflect the QLFS flow
        # only, and never affects result["gdp"]/result["cpi"]
        # (IMPLEMENTATION-SPEC-POPULATION.md §11). Population writes no
        # dataset JSON directly — only population.json's staging area —
        # and never reads or writes population-urban/population-median-age
        # (§3/§7.4). Enforces the source guard (§7.3) at two independent
        # points, deliberately redundant rather than relying on a single
        # check.
        # ============================================================
        population_client = _build_http_client(self.source_config)
        population_hub_html: bytes | None = None

        try:
            self._log.info("Fetching Population release hub: %s", _POPULATION_HUB_URL)
            excel_url, release_period, population_hub_html = _discover_population_excel(
                population_client
            )
            result["population"]["release_period"] = release_period or ""
            result["population"]["file_url"] = excel_url

            if excel_url is None:
                msg = (
                    "No publication link found for Population. Direct "
                    "probes failed (likely no standard naming) and HTML "
                    "scrape failed (likely WAF blocked). Hub URL: "
                    f"{_POPULATION_HUB_URL}"
                )
                self._log.warning(msg)
                result["population"]["status"] = "no_publication_found"
                result["population"]["errors"].append(msg)
            else:
                # Source guard, step 1 (URL-level) — checked immediately
                # after discovery, before download. Hard-fails the flow if
                # the resolved publication URL does not originate from
                # statssa.gov.za (IMPLEMENTATION-SPEC-POPULATION.md §7.3
                # step 1). This is the exact class of error that produced
                # the current data-integrity bug in population.json.
                guard_violations = _assert_population_source_guard(excel_url)
                if guard_violations:
                    result["population"]["status"] = "error"
                    result["population"]["errors"].extend(guard_violations)
                    self._log.error(
                        "Population source guard failed: %s",
                        "; ".join(guard_violations),
                    )
                else:
                    self._log.info("Located Population publication: %s", excel_url)
                    population_file_bytes = with_retry(
                        lambda: _download_publication(population_client, excel_url),  # type: ignore[arg-type]
                        policy=STATSSA_POLICY,
                        label=f"Population file download ({excel_url})",
                    )
                    result["population"]["file_size_bytes"] = len(population_file_bytes)

                    population_file_ext = (
                        Path(urllib.parse.urlparse(excel_url).path).suffix.lower() or ".bin"
                    )
                    if not dry_run:
                        try:
                            population_archive_dest, population_sha256 = save_to_archive(
                                self.config.raw_archive_dir,
                                population_file_bytes,
                                dataset_id="population",
                                source_id=self.source_id,
                                suffix=population_file_ext,
                            )
                            result["population"]["archive_path"] = portable_archive_path(
                                self.config.raw_archive_dir, population_archive_dest
                            )
                            result["population"]["sha256"] = population_sha256
                            self._log.info(
                                "Population file archived → %s (sha256=%s…, %d bytes)",
                                population_archive_dest, population_sha256[:8],
                                len(population_file_bytes),
                            )
                        except Exception as archive_exc:
                            self._log.warning(
                                "Population archive write failed: %s", archive_exc
                            )
                    else:
                        from automation.core.files import sha256_of_bytes
                        result["population"]["sha256"] = sha256_of_bytes(population_file_bytes)
                        self._log.info(
                            "[DRY RUN] Would archive Population file (%d bytes, sha256=%s…)",
                            len(population_file_bytes), result["population"]["sha256"][:8],
                        )

                    if population_file_ext not in (".xlsx", ".xls"):
                        msg = (
                            f"Downloaded Population publication is not an "
                            f"Excel workbook (extension {population_file_ext!r}, "
                            f"url={excel_url}). PDF parsing is explicitly "
                            f"out of scope — falling back to the "
                            f"manual-review path rather than guessing at "
                            f"PDF-extracted values."
                        )
                        result["population"]["status"] = "error"
                        result["population"]["errors"].append(msg)
                        self._log.error(msg)
                    else:
                        extract = parse_population_workbook(population_file_bytes)
                        # source_domain is set here, from the same URL the
                        # parser was actually given — the extract-level
                        # guard (step 2 below) re-derives from this field,
                        # not from excel_url directly, catching a wiring
                        # bug in this flow's own code rather than a data
                        # problem (§7.3 step 2).
                        extract.source_domain = urllib.parse.urlparse(excel_url).netloc.lower()
                        result["population"]["release_period"] = extract.release_period
                        self._log.info(
                            "Population workbook parsed OK — release period %s "
                            "(total_population_millions=%.1f)",
                            extract.release_period,
                            extract.total_population_millions,
                        )

                        # Source guard, step 2 (extract-level) — a
                        # defensive, redundant check, in the same spirit as
                        # _transform_inflation()'s internal rate_map
                        # assertion: catching a wiring bug in this flow's
                        # own code before it can ever reach staging
                        # (§7.3 step 2).
                        extract_guard_violations = _assert_population_source_guard(
                            f"https://{extract.source_domain}/"
                        )
                        if extract_guard_violations:
                            result["population"]["status"] = "error"
                            result["population"]["errors"].extend(extract_guard_violations)
                            self._log.error(
                                "Population extract-level source guard failed: %s",
                                "; ".join(extract_guard_violations),
                            )
                        else:
                            current_population_doc = _read_current_dataset_json(
                                _POPULATION_DATASET_JSON
                            )

                            range_errors: list[str] = []
                            range_errors += _validate_population_total(
                                extract.total_population_millions,
                                _POPULATION_TOTAL_STAT_ID,
                            )
                            range_errors += _validate_annual_label(extract.release_period)

                            if range_errors:
                                result["population"]["status"] = "error"
                                result["population"]["errors"].extend(range_errors)
                                self._log.error(
                                    "Population validation failed: %s",
                                    "; ".join(range_errors),
                                )
                            elif not _population_value_changed(current_population_doc, extract):
                                # Computed from the raw extracted value
                                # against the current on-disk document —
                                # BEFORE calling _transform_population() —
                                # since that function always refreshes
                                # _meta["last_verified"], which would make
                                # a full-document equality check never
                                # match even on a genuine no-op run.
                                # Mirrors the QLFS/GDP/CPI flows'
                                # pre-transform dataset_changed check.
                                result["population"]["status"] = "no_change"
                                result["population"]["notes"] = (
                                    f"No change: {extract.release_period} "
                                    f"Population value already matches "
                                    f"population.json."
                                )
                                self._log.info(
                                    "No change detected — population.json is already current."
                                )
                            else:
                                new_population_doc = _transform_population(
                                    current_population_doc, extract, excel_url
                                )
                                protected_violations = check_protected_fields(
                                    current_population_doc, new_population_doc
                                )
                                # The ownership boundary
                                # (IMPLEMENTATION-SPEC-POPULATION.md
                                # §7.4/§11) — hard-fails on ANY difference
                                # (not just a protected-field change) in
                                # population-urban or
                                # population-median-age, and on the
                                # stat-ID set changing at all.
                                ownership_violations = _assert_population_ownership_boundary(
                                    current_population_doc, new_population_doc
                                )
                                all_violations = protected_violations + ownership_violations
                                if all_violations:
                                    msg = f"Protected field violation: {all_violations}"
                                    result["population"]["status"] = "error"
                                    result["population"]["errors"].append(msg)
                                    self._log.error(msg)
                                else:
                                    population_jump_warnings: list[str] = []
                                    # _get_current_stat_rate() returns the
                                    # on-disk rawValue, which for
                                    # population-total is a raw headcount
                                    # (e.g. 64000000), not millions — unlike
                                    # CPI's rawValue, which already is the
                                    # rate. Convert to millions before
                                    # comparing against
                                    # extract.total_population_millions.
                                    prev_total_raw = _get_current_stat_rate(
                                        current_population_doc, _POPULATION_TOTAL_STAT_ID
                                    )
                                    prev_total_millions = (
                                        prev_total_raw / 1_000_000
                                        if prev_total_raw is not None else None
                                    )
                                    warning = _check_yoy_jump(
                                        prev_total_millions,
                                        extract.total_population_millions,
                                        _POPULATION_TOTAL_STAT_ID,
                                        threshold=_POPULATION_JUMP_WARNING_THRESHOLD,
                                    )
                                    if warning:
                                        population_jump_warnings.append(warning)

                                    population_entry = new_version_entry(
                                        dataset_id="population",
                                        source_id=self.source_id,
                                        source_url=excel_url,
                                        sha256=result["population"]["sha256"] or "",
                                        archive_path=result["population"]["archive_path"] or "",
                                        adapter_version=self.version,
                                        notes=(
                                            f"Population {extract.release_period} "
                                            f"parsed and staged (population-total "
                                            f"only). Hub URL: {_POPULATION_HUB_URL}. "
                                            + ("; ".join(population_jump_warnings) + ". "
                                               if population_jump_warnings else "")
                                            + "Manual approval required before promotion."
                                        ),
                                        run_id=run_id,
                                    )

                                    if not dry_run:
                                        try:
                                            write_staged_dataset(
                                                self.config.report_dir,
                                                dataset_id="population",
                                                version_id=population_entry.version_id,
                                                document=new_population_doc,
                                            )
                                            save_version_entry(
                                                self.config.report_dir, population_entry
                                            )
                                            self._log.info(
                                                "Staged population — version %s (status=pending)",
                                                population_entry.version_id,
                                            )
                                            result["population"]["version_id"] = population_entry.version_id
                                            result["version_ids"].append(population_entry.version_id)
                                            result["population"]["status"] = "ok"
                                            result["population"]["notes"] = (
                                                "; ".join(population_jump_warnings)
                                                if population_jump_warnings else ""
                                            )
                                        except Exception as exc:
                                            msg = f"Staging failed for population: {exc}"
                                            result["population"]["status"] = "error"
                                            result["population"]["errors"].append(msg)
                                            self._log.error(msg)
                                    else:
                                        self._log.info(
                                            "[DRY RUN] Would stage population — version %s",
                                            population_entry.version_id,
                                        )
                                        result["population"]["version_id"] = population_entry.version_id
                                        result["version_ids"].append(population_entry.version_id)
                                        result["population"]["status"] = "ok"
                                        result["population"]["notes"] = (
                                            "; ".join(population_jump_warnings)
                                            if population_jump_warnings else ""
                                        )
        except ValueError as exc:
            msg = f"Population workbook parse failed: {exc}"
            result["population"]["status"] = "error"
            result["population"]["errors"].append(msg)
            self._log.error(msg)
        except AutomationHTTPError as exc:
            msg = f"Population release hub/file HTTP error: {exc.status} {exc.reason}"
            result["population"]["status"] = "error"
            result["population"]["errors"].append(msg)
            self._log.error(msg)
        except Exception as exc:
            msg = f"Population fetch_and_apply failed: {exc}"
            result["population"]["status"] = "error"
            result["population"]["errors"].append(msg)
            self._log.error(msg)

        # Deliberate deviation, identical in kind to the GDP/CPI flows' own
        # deviation above: Population errors are NOT merged into the
        # top-level `errors` list, for the same reason (preserving every
        # existing key's meaning for the pre-existing QLFS/GDP/CPI tests,
        # which assert on `result["errors"]` describing the QLFS run
        # only). Population's errors remain fully visible in
        # result["population"]["errors"].
        #
        # Update Population hub hash so the next check_for_updates call
        # sees the new baseline (mirrors the QLFS/GDP/CPI hash-update
        # steps above).
        if not dry_run and population_hub_html is not None:
            try:
                from automation.core.files import sha256_of_bytes
                population_hub_hash = sha256_of_bytes(population_hub_html)
                self._save_population_hash(population_hub_hash)
                self._log.debug(
                    "Population hub hash updated: %s…", population_hub_hash[:8]
                )
            except Exception as exc:
                self._log.warning("Could not update Population hub hash: %s", exc)

        return result

    # ------------------------------------------------------------------
    # describe()
    # ------------------------------------------------------------------

    def describe(self) -> dict[str, Any]:
        return {
            "source_id": self.source_id,
            "display_name": self.display_name,
            "base_url": _STATSSA_BASE,
            "release_hub": _RELEASE_HUB_BASE,
            "api_available": False,
            "datasets": _STATSSA_DATASETS,
            "qlfs_family": sorted(_QLFS_FAMILY),
            "static_datasets": sorted(_STATIC_DATASETS),
            "automation_levels": {
                "unemployment": "hybrid",
                "youth-unemployment": "hybrid",
                "labour-force": "hybrid",
                "gdp": "hybrid",
                "inflation": "hybrid (cpi-headline, food-inflation only; repo-rate via SARB, untouched; annual-cpi-avg deferred)",
                "population": "hybrid (population-total only; source-guarded to statssa.gov.za; population-urban/population-median-age untouched)",
                "housing": "hybrid (pending GHS source confirmation)",
                "census": "static (erratum watch only)",
                "municipalities": "static (erratum watch only)",
            },
            "qlfs_release_windows": _QLFS_RELEASE_WINDOWS,
            "gdp_release_windows": _GDP_RELEASE_WINDOWS,
            "retry_policy": "STATSSA_POLICY (up to 2 h on release day)",
            "phase_1_status": (
                "QLFS: real ETag/hash detection + Excel download + archive. "
                "All other datasets: Phase A stubs."
            ),
            "phase_2_status": (
                "QLFS: parse (label-matched, not fixed cell ranges) + transform "
                "+ validate + stage implemented in "
                "automation/adapters/statss.py (parse_qlfs_workbook() and the "
                "_transform_* family) — reuses the existing staging/approval/"
                "promote pipeline; fetch_and_apply() stages up to 3 pending "
                "version entries (one per QLFS output dataset) and writes no "
                "dataset JSON directly. Housing/census/municipalities "
                "remain Phase A stubs — out of scope for this build. "
                "CPI is now implemented — see phase_3b_status. Population "
                "is now implemented — see phase_4_status."
            ),
            "phase_3a_status": (
                "GDP (P0441): real ETag/hash detection (_check_gdp(), mirroring "
                "_check_qlfs()) + Excel discovery/download/archive + parse "
                "(parse_gdp_workbook(), label-matched, reads EVERY available "
                "quarter column — not just the latest — to support revisions) "
                "+ transform (_transform_gdp() / _apply_gdp_growth_points(), "
                "overwrites revised historical points in place per gdp.yaml's "
                "overwrites_historical_points: true) + validate + stage, "
                "reusing the same staging/approval/promote pipeline as SARB "
                "and QLFS. fetch_and_apply() stages at most one pending "
                "version entry for gdp-growth only; gdp-annual-growth, "
                "gdp-nominal, and gdp-per-capita are never read or modified — "
                "out of scope for this build. The P0441 URL-naming convention "
                "and _GDP_GROWTH_SPEC's label match are unverified against a "
                "real workbook, same caveat class as QLFS's Excel layout."
            ),
            "phase_3b_status": (
                "CPI (P0141): real ETag/hash detection (_check_cpi(), "
                "mirroring _check_qlfs()/_check_gdp()) + Excel discovery/"
                "download/archive + parse (parse_cpi_workbook(), "
                "label-matched, reads only the latest month's column per "
                "metric — CPI does not routinely revise prior months the "
                "way GDP revises prior quarters) + transform "
                "(_transform_inflation(), reusing _apply_qlfs_rate_map() "
                "unchanged) + validate + stage, reusing the same staging/"
                "approval/promote pipeline as SARB, QLFS, and GDP. "
                "fetch_and_apply() stages at most one pending version "
                "entry for cpi-headline and food-inflation only; repo-rate "
                "(SARB-owned) and annual-cpi-avg (deferred) are never read "
                "or modified — enforced by a dedicated, tested "
                "_assert_cpi_ownership_boundary() hard-fail check that runs "
                "before staging, in addition to check_protected_fields(). "
                "The P0141 URL-naming convention, _CPI_METRIC_SPECS's label "
                "match, and the _CPI_PLAUSIBLE_RANGE/_CPI_JUMP_WARNING_"
                "THRESHOLD judgement calls are unverified against a real "
                "workbook, same caveat class as QLFS's and GDP's Excel "
                "layouts."
            ),
            "phase_4_status": (
                "Population (P0302, MYPE): real ETag/hash detection "
                "(_check_population(), mirroring _check_qlfs()/_check_gdp()/"
                "_check_cpi()) + Excel discovery/download/archive + parse "
                "(parse_population_workbook(), label-matched, reads only "
                "the latest year's column) + transform "
                "(_transform_population()/_apply_population_total_point(), "
                "a new small helper — not a reuse of _apply_qlfs_rate_map(), "
                "since population-total's display shape is materially "
                "different) + validate + stage, reusing the same staging/"
                "approval/promote pipeline as SARB, QLFS, GDP, and CPI. "
                "fetch_and_apply() stages at most one pending version "
                "entry for population-total only; population-urban and "
                "population-median-age are never read or modified — "
                "enforced by a dedicated, tested "
                "_assert_population_ownership_boundary() hard-fail check "
                "that runs before staging, in addition to "
                "check_protected_fields(). This milestone's genuinely new "
                "complication is the source guard "
                "(_assert_population_source_guard()), checked twice "
                "(URL-level before download, extract-level before "
                "staging): it hard-fails the flow if the resolved "
                "publication does not originate from statssa.gov.za — the "
                "direct fix for the World-Bank-sourced data-integrity bug "
                "documented in SA-Data-Hub-Dataset-Sourcing-Plan.md §9. "
                "The P0302 URL-naming convention, _POPULATION_METRIC_"
                "SPECS's label match, the raw-headcount-vs-millions "
                "heuristic, and the _POPULATION_PLAUSIBLE_RANGE/"
                "_POPULATION_JUMP_WARNING_THRESHOLD judgement calls are "
                "unverified against a real workbook, same caveat class as "
                "QLFS's/GDP's/CPI's Excel layouts."
            ),
            "waf_access_status": (
                "IMPLEMENTATION-SPEC-STATSSA-WAF.md Tier 1 implemented: "
                "_build_http_client() now sends an ordinary-browser-"
                "equivalent header set (non-bot User-Agent, Accept-Language, "
                "Sec-Fetch-*) for Stats SA requests; _check_qlfs()/"
                "_check_gdp() now fall through to the existing direct-"
                "publication-URL probe on a WAF_BLOCKED hub response, "
                "returning status='unknown' (probe-based signal) if a "
                "candidate file is reachable, or the prior status='error' "
                "unchanged if it is not. WAF-marker detection semantics are "
                "unchanged. Tier 2 (browser automation) is not adopted."
            ),
            "notes": (
                "One release, one job: QLFS family uses a single extractor. "
                "Population source guard is implemented "
                "(_assert_population_source_guard(), Phase 4) — hard-fails "
                "on any non-statssa.gov.za origin. "
                "GDP ETL overwrites historical points (revisions), not append — "
                "implemented in Phase 3a via _apply_gdp_growth_points()."
            ),
        }


# ---------------------------------------------------------------------------
# Register the adapter
# ---------------------------------------------------------------------------

register(StatsSAAdapter)
