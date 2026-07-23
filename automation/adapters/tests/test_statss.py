"""
Tests for automation.adapters.statss — Phase 2 (QLFS parse/transform/stage)
and Phase 3a (GDP parse/transform/stage).

Covers the acceptance criteria in IMPLEMENTATION-SPEC-STATSSA-PHASE2.md §10
(QLFS, tests 1-21 below) and IMPLEMENTATION-SPEC-GDP.md §10 (GDP, appended
at the end of this file):
  1. Successful parse of a representative QLFS workbook, verified against
     every one of the seven named values.
  2. Correct application of each transform (unemployment / youth /
     labour-force) to a realistic current-document fixture.
  3. A stat with an empty/missing series list is correctly seeded.
  4. A protected-field violation correctly aborts staging (for that
     dataset only — the other two QLFS outputs still stage normally).
  5. A quarter-over-quarter jump beyond the plausibility threshold is
     flagged, not hard-failed.
  6. Running the pipeline against unchanged source data produces
     status="no_change" with no staging / no version entries.
  7. A missing/unparseable/non-Excel source file produces status="error".

No archived QLFS .xlsx file was available in this environment to test
against (see the module docstring in automation/adapters/statss.py); the
fixture workbook below is a synthetic stand-in built to the documented
Stats SA convention (quarter-label header row + labeled indicator rows).

The GDP tests at the end of this file cover parse_gdp_workbook() (multi-
quarter extraction, fail-loudly paths, blank-column skipping),
_validate_gdp_growth_rate(), _apply_gdp_growth_points() (append, in-place
revision — the "single most important test" per IMPLEMENTATION-SPEC-
GDP.md §10 item 7 — and empty-series seeding), _transform_gdp()'s scope
boundary (gdp-annual-growth / gdp-nominal / gdp-per-capita untouched),
_check_gdp()'s hub-change detection, and fetch_and_apply()'s combined
QLFS+GDP flow (staging without direct write, no-change, protected-field
isolation between the two flows, and the approve→promote end-to-end
proof). Exactly as with QLFS, no archived GDP .xlsx file was available in
this session — the GDP fixture workbook below is likewise a synthetic
stand-in built to the documented Stats SA convention.
"""

from __future__ import annotations

import copy
import json
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest

import automation.adapters.statss as statss_mod
from automation.adapters.statss import (
    CPIExtract,
    GDPExtract,
    PopulationExtract,
    QLFSExtract,
    StatsSAAdapter,
    _apply_gdp_growth_points,
    _apply_population_total_point,
    _assert_cpi_ownership_boundary,
    _assert_population_ownership_boundary,
    _assert_population_source_guard,
    _check_qoq_jump,
    _check_yoy_jump,
    _CPI_HEADLINE_STAT_ID,
    _CPI_FOOD_STAT_ID,
    _CPI_JUMP_WARNING_THRESHOLD,
    _CPI_OWNED_STAT_IDS,
    _GDP_GROWTH_JUMP_WARNING_THRESHOLD,
    _GDP_GROWTH_PLAUSIBLE_RANGE,
    _POPULATION_JUMP_WARNING_THRESHOLD,
    _POPULATION_OWNED_STAT_IDS,
    _POPULATION_TOTAL_STAT_ID,
    _transform_gdp,
    _transform_inflation,
    _transform_labour_force,
    _transform_population,
    _transform_unemployment,
    _transform_youth_unemployment,
    _update_cpi_meta,
    _validate_annual_label,
    _validate_cpi_rate,
    _validate_gdp_growth_rate,
    _validate_monthly_label,
    _validate_percentage,
    _validate_population_total,
    _validate_quarterly_label,
    parse_cpi_workbook,
    parse_gdp_workbook,
    parse_population_workbook,
    parse_qlfs_workbook,
)
from automation.core.config import AutomationConfig, SourceConfig
from automation.core.http_client import HTTPResponse
from automation.core.metadata import check_protected_fields
from automation.core.promote import promote_version
from automation.core.staging import read_staged_dataset
from automation.core.version import approve_version, pending_versions

# ---------------------------------------------------------------------------
# Fixture workbook builder
# ---------------------------------------------------------------------------

_FIXTURE_LABELS = {
    "unemployment_rate": "Official unemployment rate",
    "youth_unemployment_narrow": "Unemployment rate: Youth (15-34)",
    "youth_unemployment_1524": "Unemployment rate: Youth (15-24)",
    "youth_unemployment_expanded": "Expanded unemployment rate: Youth",
    "neet_rate": "NEET rate (15-24)",
    "lfpr_overall": "Labour force participation rate",
    "lfpr_female": "Labour force participation rate: Female",
}

_DEFAULT_VALUES = {
    "unemployment_rate": (31.9, 31.4, 32.7),
    "youth_unemployment_narrow": (45.9, 44.8, 46.3),
    "youth_unemployment_1524": (61.9, 61.4, 60.9),
    "youth_unemployment_expanded": (57.9, 56.8, 58.1),
    "neet_rate": (37.9, 37.2, 37.6),
    "lfpr_overall": (60.9, 60.6, 60.5),
    "lfpr_female": (55.9, 55.2, 55.0),
}

_DEFAULT_HEADERS = ("Q3 2025", "Q4 2025", "Q1 2026")


def _build_fixture_workbook(
    values: dict[str, tuple[float, float, float]] = _DEFAULT_VALUES,
    headers: tuple[str, ...] = _DEFAULT_HEADERS,
) -> bytes:
    """Build a minimal, representative QLFS-style workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QLFS Time Series"
    ws.cell(row=1, column=1, value="Indicator")
    for i, h in enumerate(headers, start=2):
        ws.cell(row=1, column=i, value=h)

    row = 2
    for key, label in _FIXTURE_LABELS.items():
        ws.cell(row=row, column=1, value=label)
        for c, v in enumerate(values[key], start=2):
            ws.cell(row=row, column=c, value=v)
        row += 1

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _extract(**overrides) -> QLFSExtract:
    base = dict(
        release_period="Q1 2026",
        publication_date="2026-05-12",
        unemployment_rate=32.7,
        youth_unemployment_narrow=46.3,
        youth_unemployment_1524=60.9,
        youth_unemployment_expanded=58.1,
        neet_rate=37.6,
        lfpr_overall=60.5,
        lfpr_female=55.0,
    )
    base.update(overrides)
    return QLFSExtract(**base)


# ---------------------------------------------------------------------------
# 1. Parser
# ---------------------------------------------------------------------------


def test_parse_qlfs_workbook_extracts_all_named_values():
    data = _build_fixture_workbook()
    extract = parse_qlfs_workbook(data)

    assert extract.release_period == "Q1 2026"
    assert extract.unemployment_rate == 32.7
    assert extract.youth_unemployment_narrow == 46.3
    assert extract.youth_unemployment_1524 == 60.9
    assert extract.youth_unemployment_expanded == 58.1
    assert extract.neet_rate == 37.6
    assert extract.lfpr_overall == 60.5
    assert extract.lfpr_female == 55.0


def test_parse_qlfs_workbook_missing_indicator_fails_loudly():
    data = _build_fixture_workbook()
    wb = openpyxl.load_workbook(BytesIO(data))
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "NEET rate (15-24)":
                cell.value = "Some unrelated row"
    buf = BytesIO()
    wb.save(buf)

    with pytest.raises(ValueError, match="neet_rate"):
        parse_qlfs_workbook(buf.getvalue())


def test_parse_qlfs_workbook_not_an_excel_file_fails_loudly():
    with pytest.raises(ValueError, match="Cannot open"):
        parse_qlfs_workbook(b"this is not a valid xlsx file at all")


# ---------------------------------------------------------------------------
# 2. Validation / anomaly helpers
# ---------------------------------------------------------------------------


def test_validate_percentage_in_range():
    assert _validate_percentage(32.7, "unemployment_rate") == []


def test_validate_percentage_out_of_range():
    errors = _validate_percentage(150.0, "unemployment_rate")
    assert len(errors) == 1
    assert "unemployment_rate" in errors[0]


def test_validate_quarterly_label_ok():
    assert _validate_quarterly_label("Q1 2026") == []


def test_validate_quarterly_label_bad_format():
    assert len(_validate_quarterly_label("2026-Q1")) == 1


def test_check_qoq_jump_within_threshold_no_warning():
    assert _check_qoq_jump(31.4, 32.0, "unemployment-national") is None


def test_check_qoq_jump_beyond_threshold_flags_anomaly_not_error():
    warning = _check_qoq_jump(31.4, 40.0, "unemployment-national")
    assert warning is not None
    assert "ANOMALY" in warning


def test_check_qoq_jump_no_current_value_no_warning():
    assert _check_qoq_jump(None, 32.7, "unemployment-national") is None


# ---------------------------------------------------------------------------
# 3. Transform functions
# ---------------------------------------------------------------------------


def _unemployment_doc():
    return {
        "_meta": {"source": "Stats SA", "last_verified": "2026-05-01"},
        "statistics": [
            {
                "id": "unemployment-national",
                "value": "31.4%",
                "rawValue": 31.4,
                "change": -0.5,
                "changeLabel": "from Q3 2025 (31.9%)",
                "trend": "down",
                "source": {"name": "Stats SA", "publicationDate": "2026-02-17"},
                "lastUpdated": "2026-02-17",
                "series": [
                    {
                        "name": "Unemployment rate",
                        "unit": "%",
                        "data": [
                            {"label": "Q3 2025", "value": 31.9},
                            {"label": "Q4 2025", "value": 31.4},
                        ],
                    }
                ],
            },
            {
                "id": "labour-force-participation",
                "rawValue": 42.7,
                "series": [],
            },
        ],
    }


def _youth_doc():
    return {
        "_meta": {},
        "statistics": [
            {
                "id": "youth-unemployment-narrow",
                "rawValue": 45.5,
                "series": [{"data": [{"label": "Q4 2025", "value": 45.5}]}],
            },
            {
                "id": "youth-unemployment-1524",
                "rawValue": 61.2,
                "series": [{"data": [{"label": "Q4 2025", "value": 61.2}]}],
            },
            {
                "id": "youth-unemployment-expanded",
                "rawValue": 56.8,
                "series": [{"data": [{"label": "Q4 2025", "value": 56.8}]}],
            },
            {
                "id": "youth-neet-rate",
                "rawValue": 37.2,
                "series": [
                    {
                        "data": [
                            {"label": "2024", "value": 37.9},
                            {"label": "2025", "value": 37.2},
                        ]
                    }
                ],
            },
        ],
    }


def _labour_force_doc():
    return {
        "_meta": {},
        "statistics": [
            {
                "id": "lfpr-overall",
                "rawValue": 60.6,
                "series": [{"data": [{"label": "Q4 2025", "value": 60.6}]}],
            },
            {
                "id": "female-labour-participation",
                "rawValue": 55.2,
                "series": [{"data": [{"label": "Q4 2025", "value": 55.2}]}],
            },
        ],
    }


def test_transform_unemployment_updates_only_rate_bearing_fields():
    current = _unemployment_doc()
    updated = _transform_unemployment(current, _extract(), "https://statssa.gov.za/x.xlsx")

    stat = {s["id"]: s for s in updated["statistics"]}["unemployment-national"]
    assert stat["rawValue"] == 32.7
    assert stat["value"] == "32.7%"
    assert stat["change"] == 1.3
    assert stat["trend"] == "up"
    assert stat["lastUpdated"] == "2026-05-12"
    assert stat["source"]["publicationDate"] == "2026-05-12"
    assert stat["source"]["name"] == "Stats SA"  # untouched non-rate-bearing field
    assert stat["series"][0]["data"][-1] == {"label": "Q1 2026", "value": 32.7}
    assert len(stat["series"][0]["data"]) == 3

    untouched = {s["id"]: s for s in updated["statistics"]}["labour-force-participation"]
    assert untouched["rawValue"] == 42.7  # not part of this transform's scope

    # Deep copy — original input document is not mutated.
    assert current["statistics"][0]["rawValue"] == 31.4


def test_transform_unemployment_seeds_empty_series():
    current = _unemployment_doc()
    current["statistics"][0]["series"] = []
    del current["statistics"][0]["rawValue"]  # true "first-ever update": nothing to diff against
    updated = _transform_unemployment(current, _extract(), "https://statssa.gov.za/x.xlsx")

    stat = {s["id"]: s for s in updated["statistics"]}["unemployment-national"]
    assert stat["series"][0]["data"] == [{"label": "Q1 2026", "value": 32.7}]
    assert stat["change"] == 0.0
    assert stat["trend"] == "stable"


def test_transform_youth_unemployment_maps_all_four_stats():
    current = _youth_doc()
    updated = _transform_youth_unemployment(current, _extract(), "https://statssa.gov.za/x.xlsx")
    stats = {s["id"]: s for s in updated["statistics"]}

    assert stats["youth-unemployment-narrow"]["rawValue"] == 46.3
    assert stats["youth-unemployment-1524"]["rawValue"] == 60.9
    assert stats["youth-unemployment-expanded"]["rawValue"] == 58.1
    assert stats["youth-neet-rate"]["rawValue"] == 37.6

    neet_labels = [pt["label"] for pt in stats["youth-neet-rate"]["series"][0]["data"]]
    assert neet_labels == ["2024", "2025", "Q1 2026"]


def test_transform_labour_force_maps_both_stats():
    current = _labour_force_doc()
    updated = _transform_labour_force(current, _extract(), "https://statssa.gov.za/x.xlsx")
    stats = {s["id"]: s for s in updated["statistics"]}

    assert stats["lfpr-overall"]["rawValue"] == 60.5
    assert stats["female-labour-participation"]["rawValue"] == 55.0


def test_transform_output_id_tampering_is_caught_by_check_protected_fields():
    current = _unemployment_doc()
    updated = _transform_unemployment(current, _extract(), "https://statssa.gov.za/x.xlsx")
    updated["statistics"][0]["id"] = "unemployment-national-renamed"

    violations = check_protected_fields(current, updated)
    assert violations


# ---------------------------------------------------------------------------
# 4. fetch_and_apply() integration — network layer mocked, no real I/O
# ---------------------------------------------------------------------------


def _patch_network(monkeypatch, file_bytes: bytes, file_url: str) -> None:
    monkeypatch.setattr(
        statss_mod, "_fetch_release_hub_html",
        lambda client, url: b"<html>Q1 2026 QLFS release</html>",
    )
    monkeypatch.setattr(statss_mod, "_extract_release_period", lambda html: "Q1 2026")
    monkeypatch.setattr(statss_mod, "_determine_current_qlfs_quarter", lambda: (1, 2026))
    monkeypatch.setattr(
        statss_mod, "_probe_qlfs_publication_url",
        lambda client, q, y: file_url,
    )
    monkeypatch.setattr(statss_mod, "_download_publication", lambda client, url: file_bytes)


def _make_adapter(tmp_path: Path) -> StatsSAAdapter:
    config = AutomationConfig(
        report_dir=tmp_path / "reports",
        raw_archive_dir=tmp_path / "raw_archive",
    )
    return StatsSAAdapter(
        config, SourceConfig(source_id="statssa", display_name="Statistics South Africa")
    )


def test_fetch_and_apply_stages_all_three_datasets_without_direct_write(tmp_path, monkeypatch):
    file_bytes = _build_fixture_workbook()
    file_url = "https://www.statssa.gov.za/publications/P0211/fixture.xlsx"
    _patch_network(monkeypatch, file_bytes, file_url)

    prod_dir = tmp_path / "production"
    prod_dir.mkdir()
    stale_docs = {
        "unemployment": {
            "_meta": {},
            "statistics": [
                {"id": "unemployment-national", "rawValue": 31.4,
                 "series": [{"data": [{"label": "Q4 2025", "value": 31.4}]}]}
            ],
        },
        "youth-unemployment": _youth_doc(),
        "labour-force": _labour_force_doc(),
    }
    paths = {}
    for ds_id, doc in stale_docs.items():
        p = prod_dir / f"{ds_id}.json"
        p.write_text(json.dumps(doc), encoding="utf-8")
        paths[ds_id] = p
    monkeypatch.setattr(statss_mod, "_QLFS_DATASET_JSON", paths)

    adapter = _make_adapter(tmp_path)
    result = adapter.fetch_and_apply(dry_run=False, run_id="test-run")

    assert result["status"] == "ok"
    assert not result["errors"]
    assert len(result["version_ids"]) == 3

    # No dataset JSON was written directly — production files unchanged.
    for ds_id, p in paths.items():
        assert json.loads(p.read_text(encoding="utf-8")) == stale_docs[ds_id]

    # But each dataset has exactly one staged, pending version.
    for ds_id in paths:
        pending = pending_versions(adapter.config.report_dir, ds_id)
        assert len(pending) == 1
        staged_doc = read_staged_dataset(adapter.config.report_dir, ds_id, pending[0].version_id)
        assert staged_doc["statistics"]


def test_fetch_and_apply_no_change_produces_no_change_status(tmp_path, monkeypatch):
    file_bytes = _build_fixture_workbook()
    file_url = "https://www.statssa.gov.za/publications/P0211/fixture.xlsx"
    _patch_network(monkeypatch, file_bytes, file_url)

    prod_dir = tmp_path / "production"
    prod_dir.mkdir()
    current_docs = {
        "unemployment": {
            "_meta": {},
            "statistics": [
                {"id": "unemployment-national", "rawValue": 32.7,
                 "series": [{"data": [{"label": "Q1 2026", "value": 32.7}]}]}
            ],
        },
        "youth-unemployment": {
            "_meta": {},
            "statistics": [
                {"id": "youth-unemployment-narrow", "rawValue": 46.3, "series": []},
                {"id": "youth-unemployment-1524", "rawValue": 60.9, "series": []},
                {"id": "youth-unemployment-expanded", "rawValue": 58.1, "series": []},
                {"id": "youth-neet-rate", "rawValue": 37.6, "series": []},
            ],
        },
        "labour-force": {
            "_meta": {},
            "statistics": [
                {"id": "lfpr-overall", "rawValue": 60.5, "series": []},
                {"id": "female-labour-participation", "rawValue": 55.0, "series": []},
            ],
        },
    }
    paths = {}
    for ds_id, doc in current_docs.items():
        p = prod_dir / f"{ds_id}.json"
        p.write_text(json.dumps(doc), encoding="utf-8")
        paths[ds_id] = p
    monkeypatch.setattr(statss_mod, "_QLFS_DATASET_JSON", paths)

    adapter = _make_adapter(tmp_path)
    result = adapter.fetch_and_apply(dry_run=False, run_id="test-run")

    assert result["status"] == "no_change"
    assert result["version_ids"] == []
    for ds_id in paths:
        assert pending_versions(adapter.config.report_dir, ds_id) == []


def test_fetch_and_apply_unparseable_excel_produces_error_status(tmp_path, monkeypatch):
    file_url = "https://www.statssa.gov.za/publications/P0211/fixture.xlsx"
    _patch_network(monkeypatch, b"not a real xlsx payload", file_url)

    adapter = _make_adapter(tmp_path)
    result = adapter.fetch_and_apply(dry_run=False, run_id="test-run")

    assert result["status"] == "error"
    assert result["version_ids"] == []
    assert any("parse" in e.lower() for e in result["errors"])


def test_fetch_and_apply_pdf_fallback_produces_error_status(tmp_path, monkeypatch):
    file_url = "https://www.statssa.gov.za/publications/P0211/Presentation_QLFS_Q1_2026.pdf"
    _patch_network(monkeypatch, b"%PDF-1.4 fake pdf bytes", file_url)

    adapter = _make_adapter(tmp_path)
    result = adapter.fetch_and_apply(dry_run=False, run_id="test-run")

    assert result["status"] == "error"
    assert any("PDF" in e for e in result["errors"])


def test_fetch_and_apply_protected_field_violation_aborts_only_that_dataset(tmp_path, monkeypatch):
    file_bytes = _build_fixture_workbook()
    file_url = "https://www.statssa.gov.za/publications/P0211/fixture.xlsx"
    _patch_network(monkeypatch, file_bytes, file_url)

    prod_dir = tmp_path / "production"
    prod_dir.mkdir()
    stale_unemployment = {
        "_meta": {},
        "statistics": [
            {"id": "unemployment-national", "rawValue": 31.4,
             "series": [{"data": [{"label": "Q4 2025", "value": 31.4}]}]}
        ],
    }
    paths = {
        "unemployment": prod_dir / "unemployment.json",
        "youth-unemployment": prod_dir / "youth-unemployment.json",
        "labour-force": prod_dir / "labour-force.json",
    }
    paths["unemployment"].write_text(json.dumps(stale_unemployment), encoding="utf-8")
    paths["youth-unemployment"].write_text(json.dumps(_youth_doc()), encoding="utf-8")
    paths["labour-force"].write_text(json.dumps(_labour_force_doc()), encoding="utf-8")
    monkeypatch.setattr(statss_mod, "_QLFS_DATASET_JSON", paths)

    original_transform = statss_mod._transform_unemployment

    def _sabotaged_transform(current_doc, extract, source_url):
        doc = original_transform(current_doc, extract, source_url)
        doc["statistics"][0]["id"] = "unemployment-national-TAMPERED"
        return doc

    monkeypatch.setitem(statss_mod._QLFS_TRANSFORMS, "unemployment", _sabotaged_transform)

    adapter = _make_adapter(tmp_path)
    result = adapter.fetch_and_apply(dry_run=False, run_id="test-run")

    assert any("Protected field violation" in e for e in result["errors"])
    assert pending_versions(adapter.config.report_dir, "unemployment") == []
    assert len(pending_versions(adapter.config.report_dir, "youth-unemployment")) == 1
    assert len(pending_versions(adapter.config.report_dir, "labour-force")) == 1
    # The other two datasets still staged successfully despite one failure.
    assert result["status"] == "ok"


def test_qlfs_staged_candidate_requires_approve_then_promote(tmp_path, monkeypatch):
    """QLFS-specific version of test_full_stage_approve_promote_cycle
    (automation/core/tests/test_pipeline_integration.py), using a real
    version produced by StatsSAAdapter.fetch_and_apply() rather than a
    hand-built fixture. Closes acceptance criterion 4 of
    IMPLEMENTATION-SPEC-STATSSA-PHASE2.md, per the closeout spec §4.
    """
    file_bytes = _build_fixture_workbook()
    file_url = "https://www.statssa.gov.za/publications/P0211/fixture.xlsx"
    _patch_network(monkeypatch, file_bytes, file_url)

    prod_dir = tmp_path / "production"
    prod_dir.mkdir()
    stale_docs = {
        "unemployment": {
            "_meta": {},
            "statistics": [
                {"id": "unemployment-national", "rawValue": 31.4,
                 "series": [{"data": [{"label": "Q4 2025", "value": 31.4}]}]}
            ],
        },
        "youth-unemployment": _youth_doc(),
        "labour-force": _labour_force_doc(),
    }
    paths = {}
    for ds_id, doc in stale_docs.items():
        p = prod_dir / f"{ds_id}.json"
        p.write_text(json.dumps(doc), encoding="utf-8")
        paths[ds_id] = p
    monkeypatch.setattr(statss_mod, "_QLFS_DATASET_JSON", paths)

    # Redirect promote_version()'s production target the same way
    # core/tests/test_pipeline_integration.py does for SARB, so this test
    # never touches the real src/data/datasets/ tree.
    monkeypatch.setattr(
        "automation.core.promote.get_production_dataset_path",
        lambda dataset_id: paths[dataset_id],
    )

    adapter = _make_adapter(tmp_path)
    report_dir = adapter.config.report_dir

    # (a) Stage a genuine change for the unemployment dataset.
    result = adapter.fetch_and_apply(dry_run=False, run_id="test-run")
    assert result["status"] == "ok"
    assert not result["errors"]

    pending = pending_versions(report_dir, "unemployment")
    assert len(pending) == 1
    version_id = pending[0].version_id

    # unemployment.json's on-disk content is unchanged immediately after staging.
    assert json.loads(paths["unemployment"].read_text(encoding="utf-8")) == stale_docs["unemployment"]

    # (b) Promotion must be refused before approval — same guarantee proven
    # for SARB in test_full_stage_approve_promote_cycle, now for a real
    # QLFS-produced version.
    with pytest.raises(ValueError, match="requires 'approved'"):
        promote_version(report_dir, "unemployment", version_id)
    assert json.loads(paths["unemployment"].read_text(encoding="utf-8")) == stale_docs["unemployment"]

    # (c) Approve.
    approve_version(report_dir, "unemployment", version_id, approver="test-reviewer")
    assert pending_versions(report_dir, "unemployment") == []

    # (d) Promote — now allowed. The written file matches the staged document.
    staged_doc = read_staged_dataset(report_dir, "unemployment", version_id)
    result_path = promote_version(report_dir, "unemployment", version_id)
    assert result_path == paths["unemployment"]
    written = json.loads(paths["unemployment"].read_text(encoding="utf-8"))
    assert written == staged_doc
    # (e) Only now — after promotion — has the on-disk content changed.
    assert written != stale_docs["unemployment"]


# ---------------------------------------------------------------------------
# GDP (Phase 3a) — IMPLEMENTATION-SPEC-GDP.md §10
# ---------------------------------------------------------------------------
#
# 15 new tests, appended after the existing QLFS tests above. No existing
# test in this file is modified.

_GDP_GROWTH_LABEL = "GDP growth rate (QoQ, SAAR)"
_GDP_ANNUAL_LABEL = "Annual GDP growth rate"


def _build_gdp_fixture_workbook(
    headers: tuple[str, ...] = ("Q2 2025", "Q3 2025", "Q4 2025", "Q1 2026"),
    values: tuple[float, ...] = (0.8, 0.3, 0.4, 0.5),
    label: str = _GDP_GROWTH_LABEL,
    include_annual_row: bool = True,
) -> bytes:
    """
    Build a minimal, representative GDP-style workbook: a quarter-label
    header row plus a single "GDP growth" indicator row with a value
    under each quarter column. Unlike the QLFS fixture builder (one value
    per metric, only the latest quarter populated), this places values
    under MULTIPLE quarter columns, since parse_gdp_workbook() must read
    every available column, not just the latest.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GDP Time Series"
    ws.cell(row=1, column=1, value="Indicator")
    for i, h in enumerate(headers, start=2):
        ws.cell(row=1, column=i, value=h)

    ws.cell(row=2, column=1, value=label)
    for c, v in enumerate(values, start=2):
        ws.cell(row=2, column=c, value=v)

    if include_annual_row:
        ws.cell(row=3, column=1, value=_GDP_ANNUAL_LABEL)
        for c, v in enumerate(values, start=2):
            ws.cell(row=3, column=c, value=v + 10.0)  # distinct dummy values

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# 1. Parser
# ---------------------------------------------------------------------------


def test_parse_gdp_workbook_extracts_all_quarter_points():
    data = _build_gdp_fixture_workbook()
    extract = parse_gdp_workbook(data)

    assert extract.release_period == "Q1 2026"
    assert extract.growth_points == [
        ("Q2 2025", 0.8),
        ("Q3 2025", 0.3),
        ("Q4 2025", 0.4),
        ("Q1 2026", 0.5),
    ]


def test_parse_gdp_workbook_missing_row_fails_loudly():
    data = _build_gdp_fixture_workbook(include_annual_row=False)
    wb = openpyxl.load_workbook(BytesIO(data))
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == _GDP_GROWTH_LABEL:
                cell.value = "Some unrelated row"
    buf = BytesIO()
    wb.save(buf)

    with pytest.raises(ValueError, match="growth"):
        parse_gdp_workbook(buf.getvalue())


def test_parse_gdp_workbook_no_quarter_headers_fails_loudly():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Indicator")
    ws.cell(row=1, column=2, value="Not a quarter")
    ws.cell(row=2, column=1, value=_GDP_GROWTH_LABEL)
    ws.cell(row=2, column=2, value=0.5)
    buf = BytesIO()
    wb.save(buf)

    with pytest.raises(ValueError, match="quarter-header"):
        parse_gdp_workbook(buf.getvalue())


def test_parse_gdp_workbook_skips_blank_columns():
    data = _build_gdp_fixture_workbook()
    wb = openpyxl.load_workbook(BytesIO(data))
    ws = wb.active
    ws.cell(row=2, column=3).value = None  # blank out Q3 2025's value cell
    buf = BytesIO()
    wb.save(buf)

    extract = parse_gdp_workbook(buf.getvalue())
    labels = [label for label, _ in extract.growth_points]
    assert "Q3 2025" not in labels
    assert labels == ["Q2 2025", "Q4 2025", "Q1 2026"]


# ---------------------------------------------------------------------------
# 2. Validation
# ---------------------------------------------------------------------------


def test_validate_gdp_growth_rate_in_range_and_out_of_range():
    assert _validate_gdp_growth_rate(0.5, "Q1 2026") == []
    # A real historical value from gdp.json itself — proves the range
    # isn't naively [0, 100] like QLFS's _validate_percentage().
    assert _validate_gdp_growth_rate(-6.2, "2020") == []

    errors = _validate_gdp_growth_rate(45.0, "Q1 2026")
    assert len(errors) == 1
    assert "Q1 2026" in errors[0]


# ---------------------------------------------------------------------------
# 3. _apply_gdp_growth_points() / _transform_gdp()
# ---------------------------------------------------------------------------


def _gdp_growth_stat(series_data=None, **overrides) -> dict:
    stat = {
        "id": "gdp-growth",
        "categoryId": "gdp",
        "title": "GDP Growth Rate (Quarter-on-Quarter)",
        "value": "0.4%",
        "rawValue": 0.4,
        "unit": "%",
        "change": 0.1,
        "changeLabel": "from Q3 2025",
        "trend": "up",
        "source": {"name": "Statistics South Africa", "publicationDate": "2026-03-10"},
        "lastUpdated": "2026-03-10",
        "series": [
            {
                "name": "GDP Growth (%, QoQ SAAR)",
                "unit": "%",
                "data": series_data if series_data is not None else [
                    {"label": "Q3 2025", "value": 0.3},
                    {"label": "Q4 2025", "value": 0.4},
                ],
            }
        ],
    }
    stat.update(overrides)
    return stat


def _gdp_doc(growth_stat: dict | None = None) -> dict:
    return {
        "_meta": {"source": "Stats SA", "last_verified": "2026-05-31"},
        "statistics": [
            growth_stat if growth_stat is not None else _gdp_growth_stat(),
            {
                "id": "gdp-annual-growth",
                "rawValue": 1.1,
                "series": [{"data": [
                    {"label": "2024", "value": 0.5},
                    {"label": "2025", "value": 1.1},
                ]}],
            },
            {
                "id": "gdp-nominal",
                "rawValue": 7670,
                "series": [{"data": [
                    {"label": "2024", "value": 7352},
                    {"label": "2025", "value": 7670},
                ]}],
            },
            {
                "id": "gdp-per-capita",
                "rawValue": 120800,
                "series": [{"data": [
                    {"label": "2024", "value": 114869},
                    {"label": "2025", "value": 120800},
                ]}],
            },
        ],
    }


def test_apply_gdp_growth_points_appends_new_point():
    doc = _gdp_doc()
    notes = _apply_gdp_growth_points(doc, [("Q1 2026", 0.5)], publication_date="2026-06-09")

    stat = {s["id"]: s for s in doc["statistics"]}["gdp-growth"]
    data = stat["series"][0]["data"]
    assert data[-1] == {"label": "Q1 2026", "value": 0.5}
    assert len(data) == 3
    assert stat["rawValue"] == 0.5
    assert stat["value"] == "0.5%"
    assert stat["change"] == 0.1
    assert stat["trend"] == "up"
    assert stat["lastUpdated"] == "2026-06-09"
    assert notes == []  # nothing existing changed — pure append


def test_apply_gdp_growth_points_revises_historical_point():
    """The single most important test in this milestone (IMPLEMENTATION-
    SPEC-GDP.md §10 item 7) — the direct proof of the
    ``overwrites_historical_points: true`` requirement from gdp.yaml.
    """
    doc = _gdp_doc()
    notes = _apply_gdp_growth_points(
        doc,
        [("Q3 2025", 0.6), ("Q1 2026", 0.5)],
        publication_date="2026-06-09",
    )

    stat = {s["id"]: s for s in doc["statistics"]}["gdp-growth"]
    data = stat["series"][0]["data"]

    # Q3 2025 overwritten in place — not duplicated.
    q3_points = [pt for pt in data if pt["label"] == "Q3 2025"]
    assert len(q3_points) == 1
    assert q3_points[0]["value"] == 0.6

    # Q1 2026 appended as the new point.
    assert data[-1] == {"label": "Q1 2026", "value": 0.5}
    assert len(data) == 3

    # A revision note was produced, naming the revised period.
    assert any("Q3 2025" in n and "Revised" in n for n in notes)

    # Headline fields are driven by Q1 2026 (the newest point) — NOT by
    # the revised older Q3 2025 point.
    assert stat["rawValue"] == 0.5
    assert stat["value"] == "0.5%"
    assert stat["trend"] == "up"


def test_apply_gdp_growth_points_seeds_empty_series():
    doc = _gdp_doc()
    stat = {s["id"]: s for s in doc["statistics"]}["gdp-growth"]
    stat["series"] = []
    del stat["rawValue"]  # true "first-ever update": nothing to diff against

    notes = _apply_gdp_growth_points(doc, [("Q1 2026", 0.5)], publication_date="2026-06-09")

    stat = {s["id"]: s for s in doc["statistics"]}["gdp-growth"]
    assert stat["series"][0]["data"] == [{"label": "Q1 2026", "value": 0.5}]
    assert stat["change"] == 0.0
    assert stat["trend"] == "stable"
    assert notes == []


def test_transform_gdp_only_touches_gdp_growth():
    doc = _gdp_doc()
    extract = GDPExtract(
        release_period="Q1 2026",
        publication_date="2026-06-09",
        growth_points=[("Q4 2025", 0.4), ("Q1 2026", 0.5)],
    )
    new_doc, warnings = _transform_gdp(doc, extract, "https://statssa.gov.za/gdp.xlsx")

    original_stats = {s["id"]: s for s in doc["statistics"]}
    new_stats = {s["id"]: s for s in new_doc["statistics"]}

    for stat_id in ("gdp-annual-growth", "gdp-nominal", "gdp-per-capita"):
        assert new_stats[stat_id] == original_stats[stat_id]

    assert new_stats["gdp-growth"]["rawValue"] == 0.5
    # Deep copy — original input document is not mutated.
    assert doc["statistics"][0]["rawValue"] == 0.4


# ---------------------------------------------------------------------------
# 4. Quarter-over-quarter anomaly threshold (GDP-specific)
# ---------------------------------------------------------------------------


def test_qoq_jump_flags_large_gdp_swing():
    warning = _check_qoq_jump(0.4, 6.0, "gdp-growth", threshold=_GDP_GROWTH_JUMP_WARNING_THRESHOLD)
    assert warning is not None
    assert "ANOMALY" in warning

    # A swing within the wider GDP threshold (but that would exceed
    # QLFS's tighter 3.0pp default) is NOT flagged — proving the
    # GDP-specific threshold is actually wider, not just re-used.
    assert _check_qoq_jump(0.4, 4.0, "gdp-growth", threshold=_GDP_GROWTH_JUMP_WARNING_THRESHOLD) is None


# ---------------------------------------------------------------------------
# 5. _check_gdp() — hub-change detection
# ---------------------------------------------------------------------------


def _make_adapter_for_gdp(tmp_path: Path) -> StatsSAAdapter:
    config = AutomationConfig(
        report_dir=tmp_path / "reports",
        raw_archive_dir=tmp_path / "raw_archive",
    )
    return StatsSAAdapter(
        config, SourceConfig(source_id="statssa", display_name="Statistics South Africa")
    )


def test_check_gdp_detects_hub_change(tmp_path, monkeypatch):
    adapter = _make_adapter_for_gdp(tmp_path)

    changed_response = HTTPResponse(
        url=statss_mod._GDP_HUB_URL,
        status=200,
        headers={},
        body=b"<html>Q1 2026 GDP release</html>",
        content_sha256="new-hash",
    )
    monkeypatch.setattr(
        "automation.core.http_client.HTTPClient.etag_check",
        lambda self, url, **kwargs: (True, changed_response),
    )
    result = adapter.check_for_updates("gdp", None)
    assert result.status == "update_available"

    # Fresh adapter (new cache), hub now reports unchanged.
    adapter2 = _make_adapter_for_gdp(tmp_path)
    unchanged_response = HTTPResponse(
        url=statss_mod._GDP_HUB_URL,
        status=200,
        headers={},
        body=b"<html>Q1 2026 GDP release</html>",
        content_sha256="new-hash",
    )
    monkeypatch.setattr(
        "automation.core.http_client.HTTPClient.etag_check",
        lambda self, url, **kwargs: (False, unchanged_response),
    )
    result2 = adapter2.check_for_updates("gdp", None)
    assert result2.status == "up_to_date"


# ---------------------------------------------------------------------------
# 6. fetch_and_apply() integration — network layer mocked, GDP + QLFS
# ---------------------------------------------------------------------------


def _patch_qlfs_and_gdp_network(
    monkeypatch,
    qlfs_bytes: bytes,
    qlfs_url: str,
    gdp_bytes: bytes,
    gdp_url: str,
) -> None:
    """Combined network patch: deterministic mocks for both the QLFS and
    GDP discovery/download paths within a single fetch_and_apply() call.
    """
    monkeypatch.setattr(
        statss_mod, "_fetch_release_hub_html",
        lambda client, url: b"<html>Q1 2026 QLFS release</html>",
    )
    monkeypatch.setattr(statss_mod, "_extract_release_period", lambda html: "Q1 2026")
    monkeypatch.setattr(statss_mod, "_determine_current_qlfs_quarter", lambda: (1, 2026))
    monkeypatch.setattr(statss_mod, "_probe_qlfs_publication_url", lambda client, q, y: qlfs_url)

    monkeypatch.setattr(statss_mod, "_determine_current_gdp_quarter", lambda: (1, 2026))
    monkeypatch.setattr(
        statss_mod, "_discover_gdp_excel",
        lambda client, **kwargs: (gdp_url, "Q1 2026", b"<html>Q1 2026 GDP release</html>"),
    )

    def _download(client, url):
        if url == qlfs_url:
            return qlfs_bytes
        if url == gdp_url:
            return gdp_bytes
        raise AssertionError(f"Unexpected download URL in test: {url}")

    monkeypatch.setattr(statss_mod, "_download_publication", _download)


def _stale_qlfs_docs_and_paths(prod_dir: Path) -> tuple[dict, dict]:
    stale_docs = {
        "unemployment": {
            "_meta": {},
            "statistics": [
                {"id": "unemployment-national", "rawValue": 31.4,
                 "series": [{"data": [{"label": "Q4 2025", "value": 31.4}]}]}
            ],
        },
        "youth-unemployment": _youth_doc(),
        "labour-force": _labour_force_doc(),
    }
    paths = {}
    for ds_id, doc in stale_docs.items():
        p = prod_dir / f"{ds_id}.json"
        p.write_text(json.dumps(doc), encoding="utf-8")
        paths[ds_id] = p
    return stale_docs, paths


def test_fetch_and_apply_stages_gdp_without_direct_write(tmp_path, monkeypatch):
    qlfs_bytes = _build_fixture_workbook()
    qlfs_url = "https://www.statssa.gov.za/publications/P0211/fixture.xlsx"
    gdp_bytes = _build_gdp_fixture_workbook(
        headers=("Q4 2025", "Q1 2026"), values=(0.4, 0.5), include_annual_row=False,
    )
    gdp_url = "https://www.statssa.gov.za/publications/P0441/gdp_fixture.xlsx"
    _patch_qlfs_and_gdp_network(monkeypatch, qlfs_bytes, qlfs_url, gdp_bytes, gdp_url)

    prod_dir = tmp_path / "production"
    prod_dir.mkdir()
    stale_qlfs_docs, qlfs_paths = _stale_qlfs_docs_and_paths(prod_dir)
    monkeypatch.setattr(statss_mod, "_QLFS_DATASET_JSON", qlfs_paths)

    stale_gdp_doc = _gdp_doc()
    gdp_path = prod_dir / "gdp.json"
    gdp_path.write_text(json.dumps(stale_gdp_doc), encoding="utf-8")
    monkeypatch.setattr(statss_mod, "_GDP_DATASET_JSON", gdp_path)

    adapter = _make_adapter(tmp_path)
    result = adapter.fetch_and_apply(dry_run=False, run_id="test-run")

    # QLFS portion of the result is unaffected by GDP being added.
    assert result["status"] == "ok"
    assert not result["errors"]
    # result["version_ids"] is QLFS-only — GDP's version_id must never be
    # appended into this shared list (audit fix: version_id
    # cross-contamination).
    assert len(result["version_ids"]) == 3  # 3 QLFS datasets only
    for ds_id, p in qlfs_paths.items():
        assert json.loads(p.read_text(encoding="utf-8")) == stale_qlfs_docs[ds_id]

    # GDP staged, no direct write to gdp.json.
    assert result["gdp"]["status"] == "ok"
    assert result["gdp"]["version_id"] is not None
    # GDP's version_id lives only in result["gdp"]["version_id"] — it must
    # NOT be mixed into the QLFS-only result["version_ids"] list.
    assert result["gdp"]["version_id"] not in result["version_ids"]
    assert json.loads(gdp_path.read_text(encoding="utf-8")) == stale_gdp_doc

    pending = pending_versions(adapter.config.report_dir, "gdp")
    assert len(pending) == 1
    staged_doc = read_staged_dataset(adapter.config.report_dir, "gdp", pending[0].version_id)
    assert staged_doc["statistics"]
    staged_growth = {s["id"]: s for s in staged_doc["statistics"]}["gdp-growth"]
    assert staged_growth["rawValue"] == 0.5


def test_fetch_and_apply_qlfs_hub_exception_still_runs_gdp(tmp_path, monkeypatch):
    """
    Audit finding #8 (regression guard): if the QLFS hub fetch itself
    raises (e.g. network down, non-WAF HTTP error), GDP must still
    execute and stage successfully. Before the fix, this early
    `return result` inside fetch_and_apply() prevented GDP/CPI/Population
    from ever running this cycle.
    """
    gdp_bytes = _build_gdp_fixture_workbook(
        headers=("Q4 2025", "Q1 2026"), values=(0.4, 0.5), include_annual_row=False,
    )
    gdp_url = "https://www.statssa.gov.za/publications/P0441/gdp_fixture.xlsx"

    # QLFS discovery raises before a file_url can ever be resolved.
    monkeypatch.setattr(statss_mod, "_determine_current_qlfs_quarter", lambda: (1, 2026))

    def _boom(client, q, y):
        raise RuntimeError("QLFS hub unreachable")

    monkeypatch.setattr(statss_mod, "_probe_qlfs_publication_url", _boom)

    monkeypatch.setattr(statss_mod, "_determine_current_gdp_quarter", lambda: (1, 2026))
    monkeypatch.setattr(
        statss_mod, "_discover_gdp_excel",
        lambda client, **kwargs: (gdp_url, "Q1 2026", b"<html>Q1 2026 GDP release</html>"),
    )
    monkeypatch.setattr(statss_mod, "_download_publication", lambda client, url: gdp_bytes)

    prod_dir = tmp_path / "production"
    prod_dir.mkdir()
    stale_gdp_doc = _gdp_doc()
    gdp_path = prod_dir / "gdp.json"
    gdp_path.write_text(json.dumps(stale_gdp_doc), encoding="utf-8")
    monkeypatch.setattr(statss_mod, "_GDP_DATASET_JSON", gdp_path)

    adapter = _make_adapter(tmp_path)
    result = adapter.fetch_and_apply(dry_run=False, run_id="test-run")

    # QLFS failed loudly ...
    assert result["status"] == "error"
    assert result["errors"]
    assert not result["version_ids"]

    # ... but GDP still ran independently and staged successfully.
    assert result["gdp"]["status"] == "ok"
    assert result["gdp"]["version_id"] is not None
    assert result["gdp"]["version_id"] not in result["version_ids"]


def test_fetch_and_apply_qlfs_no_publication_found_still_runs_gdp(tmp_path, monkeypatch):
    """
    Audit finding #8 (regression guard): if the QLFS hub is fetched
    successfully but no publication link can be located at all (the
    exact "no_publication_found" scenario the audit called out), GDP
    must still execute and stage successfully.
    """
    gdp_bytes = _build_gdp_fixture_workbook(
        headers=("Q4 2025", "Q1 2026"), values=(0.4, 0.5), include_annual_row=False,
    )
    gdp_url = "https://www.statssa.gov.za/publications/P0441/gdp_fixture.xlsx"

    monkeypatch.setattr(
        statss_mod, "_fetch_release_hub_html",
        lambda client, url: b"<html>no excel link here</html>",
    )
    monkeypatch.setattr(statss_mod, "_extract_release_period", lambda html: "Q1 2026")
    monkeypatch.setattr(statss_mod, "_determine_current_qlfs_quarter", lambda: (1, 2026))
    # Neither the direct probe nor the hub scrape find a publication link.
    monkeypatch.setattr(statss_mod, "_probe_qlfs_publication_url", lambda client, q, y: None)
    monkeypatch.setattr(statss_mod, "_extract_excel_url", lambda html, base_url=None: None)

    monkeypatch.setattr(statss_mod, "_determine_current_gdp_quarter", lambda: (1, 2026))
    monkeypatch.setattr(
        statss_mod, "_discover_gdp_excel",
        lambda client, **kwargs: (gdp_url, "Q1 2026", b"<html>Q1 2026 GDP release</html>"),
    )
    monkeypatch.setattr(statss_mod, "_download_publication", lambda client, url: gdp_bytes)

    prod_dir = tmp_path / "production"
    prod_dir.mkdir()
    stale_gdp_doc = _gdp_doc()
    gdp_path = prod_dir / "gdp.json"
    gdp_path.write_text(json.dumps(stale_gdp_doc), encoding="utf-8")
    monkeypatch.setattr(statss_mod, "_GDP_DATASET_JSON", gdp_path)

    adapter = _make_adapter(tmp_path)
    result = adapter.fetch_and_apply(dry_run=False, run_id="test-run")

    # QLFS correctly reports no_publication_found, not a generic error ...
    assert result["status"] == "no_publication_found"
    assert not result["version_ids"]

    # ... but GDP still ran independently and staged successfully.
    assert result["gdp"]["status"] == "ok"
    assert result["gdp"]["version_id"] is not None
    assert result["gdp"]["version_id"] not in result["version_ids"]


def test_fetch_and_apply_qlfs_download_error_still_runs_gdp(tmp_path, monkeypatch):
    """
    Audit finding #8 (regression guard): if the QLFS publication link is
    found but the file download itself fails, GDP must still execute and
    stage successfully.
    """
    from automation.core.http_client import AutomationHTTPError

    qlfs_url = "https://www.statssa.gov.za/publications/P0211/fixture.xlsx"
    gdp_bytes = _build_gdp_fixture_workbook(
        headers=("Q4 2025", "Q1 2026"), values=(0.4, 0.5), include_annual_row=False,
    )
    gdp_url = "https://www.statssa.gov.za/publications/P0441/gdp_fixture.xlsx"

    monkeypatch.setattr(
        statss_mod, "_fetch_release_hub_html",
        lambda client, url: b"<html>Q1 2026 QLFS release</html>",
    )
    monkeypatch.setattr(statss_mod, "_extract_release_period", lambda html: "Q1 2026")
    monkeypatch.setattr(statss_mod, "_determine_current_qlfs_quarter", lambda: (1, 2026))
    monkeypatch.setattr(statss_mod, "_probe_qlfs_publication_url", lambda client, q, y: qlfs_url)

    monkeypatch.setattr(statss_mod, "_determine_current_gdp_quarter", lambda: (1, 2026))
    monkeypatch.setattr(
        statss_mod, "_discover_gdp_excel",
        lambda client, **kwargs: (gdp_url, "Q1 2026", b"<html>Q1 2026 GDP release</html>"),
    )

    def _download(client, url):
        if url == qlfs_url:
            raise AutomationHTTPError(url, 500, "Internal Server Error")
        if url == gdp_url:
            return gdp_bytes
        raise AssertionError(f"Unexpected download URL in test: {url}")

    monkeypatch.setattr(statss_mod, "_download_publication", _download)

    prod_dir = tmp_path / "production"
    prod_dir.mkdir()
    stale_gdp_doc = _gdp_doc()
    gdp_path = prod_dir / "gdp.json"
    gdp_path.write_text(json.dumps(stale_gdp_doc), encoding="utf-8")
    monkeypatch.setattr(statss_mod, "_GDP_DATASET_JSON", gdp_path)

    adapter = _make_adapter(tmp_path)
    result = adapter.fetch_and_apply(dry_run=False, run_id="test-run")

    # QLFS failed loudly on download ...
    assert result["status"] == "error"
    assert result["errors"]
    assert not result["version_ids"]

    # ... but GDP still ran independently and staged successfully.
    assert result["gdp"]["status"] == "ok"
    assert result["gdp"]["version_id"] is not None
    assert result["gdp"]["version_id"] not in result["version_ids"]


def test_fetch_and_apply_stages_gdp_waf_fallback_real(tmp_path, monkeypatch):
    """
    Ensures that if the GDP hub is WAF-blocked, _discover_gdp_excel successfully
    falls back to the direct-URL probe via _resolve_publication_url.
    """
    from automation.core.http_client import AutomationHTTPError
    qlfs_bytes = _build_fixture_workbook()
    qlfs_url = "https://www.statssa.gov.za/publications/P0211/fixture.xlsx"
    gdp_bytes = _build_gdp_fixture_workbook(
        headers=("Q4 2025", "Q1 2026"), values=(0.4, 0.5), include_annual_row=False,
    )
    gdp_url = "https://www.statssa.gov.za/publications/P0441/gdp_fixture.xlsx"

    monkeypatch.setattr(statss_mod, "_extract_release_period", lambda html: "Q1 2026")
    monkeypatch.setattr(statss_mod, "_determine_current_qlfs_quarter", lambda: (1, 2026))
    monkeypatch.setattr(statss_mod, "_probe_qlfs_publication_url", lambda client, q, y: qlfs_url)

    def mock_fetch_hub(client, url):
        if url == statss_mod._GDP_HUB_URL:
            raise AutomationHTTPError(url, 403, "WAF_BLOCKED: Incapsula WAF challenge detected")
        return b"<html>Q1 2026 QLFS release</html>"
        
    monkeypatch.setattr(statss_mod, "_fetch_release_hub_html", mock_fetch_hub)

    monkeypatch.setattr(statss_mod, "_determine_current_gdp_quarter", lambda: (1, 2026))
    monkeypatch.setattr(statss_mod, "_probe_gdp_publication_url", lambda client, q, y: gdp_url)
    
    def _download(client, url):
        if url == qlfs_url:
            return qlfs_bytes
        if url == gdp_url:
            return gdp_bytes
        raise AssertionError(f"Unexpected download URL in test: {url}")

    monkeypatch.setattr(statss_mod, "_download_publication", _download)

    prod_dir = tmp_path / "production"
    prod_dir.mkdir()
    stale_qlfs_docs, qlfs_paths = _stale_qlfs_docs_and_paths(prod_dir)
    monkeypatch.setattr(statss_mod, "_QLFS_DATASET_JSON", qlfs_paths)

    stale_gdp_doc = _gdp_doc()
    gdp_path = prod_dir / "gdp.json"
    gdp_path.write_text(json.dumps(stale_gdp_doc), encoding="utf-8")
    monkeypatch.setattr(statss_mod, "_GDP_DATASET_JSON", gdp_path)

    adapter = _make_adapter(tmp_path)
    result = adapter.fetch_and_apply(dry_run=False, run_id="test-run")

    assert result["gdp"]["status"] == "ok"
    assert result["gdp"]["version_id"] is not None


def test_fetch_and_apply_gdp_no_change_produces_no_change_status(tmp_path, monkeypatch):
    qlfs_bytes = _build_fixture_workbook()
    qlfs_url = "https://www.statssa.gov.za/publications/P0211/fixture.xlsx"
    gdp_bytes = _build_gdp_fixture_workbook(
        headers=("Q3 2025", "Q4 2025"), values=(0.3, 0.4), include_annual_row=False,
    )
    gdp_url = "https://www.statssa.gov.za/publications/P0441/gdp_fixture.xlsx"
    _patch_qlfs_and_gdp_network(monkeypatch, qlfs_bytes, qlfs_url, gdp_bytes, gdp_url)

    prod_dir = tmp_path / "production"
    prod_dir.mkdir()
    _, qlfs_paths = _stale_qlfs_docs_and_paths(prod_dir)
    # Make the QLFS values already match the fixture so QLFS is no_change,
    # keeping this test focused on the GDP no_change path.
    for ds_id, doc in {
        "unemployment": {
            "_meta": {},
            "statistics": [
                {"id": "unemployment-national", "rawValue": 32.7,
                 "series": [{"data": [{"label": "Q1 2026", "value": 32.7}]}]}
            ],
        },
        "youth-unemployment": {
            "_meta": {},
            "statistics": [
                {"id": "youth-unemployment-narrow", "rawValue": 46.3, "series": []},
                {"id": "youth-unemployment-1524", "rawValue": 60.9, "series": []},
                {"id": "youth-unemployment-expanded", "rawValue": 58.1, "series": []},
                {"id": "youth-neet-rate", "rawValue": 37.6, "series": []},
            ],
        },
        "labour-force": {
            "_meta": {},
            "statistics": [
                {"id": "lfpr-overall", "rawValue": 60.5, "series": []},
                {"id": "female-labour-participation", "rawValue": 55.0, "series": []},
            ],
        },
    }.items():
        qlfs_paths[ds_id].write_text(json.dumps(doc), encoding="utf-8")
    monkeypatch.setattr(statss_mod, "_QLFS_DATASET_JSON", qlfs_paths)

    stale_gdp_doc = _gdp_doc()  # series already ends at Q3 2025: 0.3, Q4 2025: 0.4
    gdp_path = prod_dir / "gdp.json"
    gdp_path.write_text(json.dumps(stale_gdp_doc), encoding="utf-8")
    monkeypatch.setattr(statss_mod, "_GDP_DATASET_JSON", gdp_path)

    adapter = _make_adapter(tmp_path)
    result = adapter.fetch_and_apply(dry_run=False, run_id="test-run")

    assert result["gdp"]["status"] == "no_change"
    assert result["gdp"]["version_id"] is None
    assert pending_versions(adapter.config.report_dir, "gdp") == []
    assert json.loads(gdp_path.read_text(encoding="utf-8")) == stale_gdp_doc


def test_fetch_and_apply_gdp_protected_field_violation_does_not_affect_qlfs(tmp_path, monkeypatch):
    qlfs_bytes = _build_fixture_workbook()
    qlfs_url = "https://www.statssa.gov.za/publications/P0211/fixture.xlsx"
    gdp_bytes = _build_gdp_fixture_workbook(
        headers=("Q4 2025", "Q1 2026"), values=(0.4, 0.5), include_annual_row=False,
    )
    gdp_url = "https://www.statssa.gov.za/publications/P0441/gdp_fixture.xlsx"
    _patch_qlfs_and_gdp_network(monkeypatch, qlfs_bytes, qlfs_url, gdp_bytes, gdp_url)

    prod_dir = tmp_path / "production"
    prod_dir.mkdir()
    stale_qlfs_docs, qlfs_paths = _stale_qlfs_docs_and_paths(prod_dir)
    monkeypatch.setattr(statss_mod, "_QLFS_DATASET_JSON", qlfs_paths)

    stale_gdp_doc = _gdp_doc()
    gdp_path = prod_dir / "gdp.json"
    gdp_path.write_text(json.dumps(stale_gdp_doc), encoding="utf-8")
    monkeypatch.setattr(statss_mod, "_GDP_DATASET_JSON", gdp_path)

    # Sabotage the GDP transform to violate a protected field.
    original_transform_gdp = statss_mod._transform_gdp

    def _sabotaged_transform_gdp(current_doc, extract, source_url=""):
        doc, warnings = original_transform_gdp(current_doc, extract, source_url)
        doc["statistics"][0]["id"] = "gdp-growth-TAMPERED"
        return doc, warnings

    monkeypatch.setattr(statss_mod, "_transform_gdp", _sabotaged_transform_gdp)

    adapter = _make_adapter(tmp_path)
    result = adapter.fetch_and_apply(dry_run=False, run_id="test-run")

    # GDP failed due to the protected-field violation …
    assert result["gdp"]["status"] == "error"
    assert any("Protected field violation" in e for e in result["gdp"]["errors"])
    assert pending_versions(adapter.config.report_dir, "gdp") == []
    assert json.loads(gdp_path.read_text(encoding="utf-8")) == stale_gdp_doc

    # … while the QLFS portion of the SAME fetch_and_apply() call still
    # succeeded normally — direct proof the two flows are isolated.
    assert result["status"] == "ok"
    assert not result["errors"]
    for ds_id, p in qlfs_paths.items():
        assert json.loads(p.read_text(encoding="utf-8")) == stale_qlfs_docs[ds_id]
    for ds_id in qlfs_paths:
        assert len(pending_versions(adapter.config.report_dir, ds_id)) == 1


def test_gdp_staged_candidate_requires_approve_then_promote(tmp_path, monkeypatch):
    """The GDP-specific equivalent of
    test_qlfs_staged_candidate_requires_approve_then_promote — closes this
    acceptance-criterion class for GDP from the start of this milestone
    rather than as a later closeout (IMPLEMENTATION-SPEC-GDP.md §10 item 15).
    """
    qlfs_bytes = _build_fixture_workbook()
    qlfs_url = "https://www.statssa.gov.za/publications/P0211/fixture.xlsx"
    gdp_bytes = _build_gdp_fixture_workbook(
        headers=("Q4 2025", "Q1 2026"), values=(0.4, 0.5), include_annual_row=False,
    )
    gdp_url = "https://www.statssa.gov.za/publications/P0441/gdp_fixture.xlsx"
    _patch_qlfs_and_gdp_network(monkeypatch, qlfs_bytes, qlfs_url, gdp_bytes, gdp_url)

    prod_dir = tmp_path / "production"
    prod_dir.mkdir()
    stale_qlfs_docs, qlfs_paths = _stale_qlfs_docs_and_paths(prod_dir)
    monkeypatch.setattr(statss_mod, "_QLFS_DATASET_JSON", qlfs_paths)

    stale_gdp_doc = _gdp_doc()
    gdp_path = prod_dir / "gdp.json"
    gdp_path.write_text(json.dumps(stale_gdp_doc), encoding="utf-8")
    monkeypatch.setattr(statss_mod, "_GDP_DATASET_JSON", gdp_path)

    # Redirect promote_version()'s production target, same technique used
    # for QLFS/SARB in this file, so this test never touches the real
    # src/data/datasets/ tree.
    monkeypatch.setattr(
        "automation.core.promote.get_production_dataset_path",
        lambda dataset_id: {**qlfs_paths, "gdp": gdp_path}[dataset_id],
    )

    adapter = _make_adapter(tmp_path)
    report_dir = adapter.config.report_dir

    # (a) Stage a genuine GDP change.
    result = adapter.fetch_and_apply(dry_run=False, run_id="test-run")
    assert result["gdp"]["status"] == "ok"

    pending = pending_versions(report_dir, "gdp")
    assert len(pending) == 1
    version_id = pending[0].version_id

    # gdp.json's on-disk content is unchanged immediately after staging.
    assert json.loads(gdp_path.read_text(encoding="utf-8")) == stale_gdp_doc

    # (b) Promotion must be refused before approval.
    with pytest.raises(ValueError, match="requires 'approved'"):
        promote_version(report_dir, "gdp", version_id)
    assert json.loads(gdp_path.read_text(encoding="utf-8")) == stale_gdp_doc

    # (c) Approve.
    approve_version(report_dir, "gdp", version_id, approver="test-reviewer")
    assert pending_versions(report_dir, "gdp") == []

    # (d) Promote — now allowed. The written file matches the staged document.
    staged_doc = read_staged_dataset(report_dir, "gdp", version_id)
    result_path = promote_version(report_dir, "gdp", version_id)
    assert result_path == gdp_path
    written = json.loads(gdp_path.read_text(encoding="utf-8"))
    assert written == staged_doc
    # (e) Only now — after promotion — has the on-disk content changed.
    assert written != stale_gdp_doc


def test_fetch_release_hub_html_waf_raises_automation_http_error(monkeypatch):
    """
    _fetch_release_hub_html() must raise AutomationHTTPError (not TypeError)
    when WAF challenge is detected. This test exercises the bug at line 674
    that previously caused TypeError due to missing url positional argument.
    """
    from automation.core.http_client import HTTPClient, HTTPResponse, AutomationHTTPError
    from automation.core.config import SourceConfig
    import automation.adapters.statss as statss_mod

    hub_url = "https://www.statssa.gov.za/?page_id=1854&PPN=P0211"
    waf_body = b"<html><body>_Incapsula_Resource challenge</body></html>"
    waf_resp = HTTPResponse(
        url=hub_url, status=200, headers={}, body=waf_body,
        content_sha256="waf-hash",
    )
    monkeypatch.setattr(HTTPClient, "get", lambda self, url, **kw: waf_resp)

    client = statss_mod._build_http_client(
        SourceConfig(source_id="statssa", display_name="Stats SA")
    )
    with pytest.raises(AutomationHTTPError) as exc_info:
        statss_mod._fetch_release_hub_html(client, hub_url)

    assert exc_info.value.url == hub_url
    assert exc_info.value.status == 403
    assert "WAF_BLOCKED" in exc_info.value.reason


def test_resolve_publication_url_waf_fallback(monkeypatch):
    """
    If the release hub is WAF-blocked (raising AutomationHTTPError), the fallback
    scrape is bypassed but the primary probed URL (Strategy 1) is still returned.
    This asserts the fix for the fetch-path WAF asymmetry bug.
    """
    from automation.core.http_client import HTTPClient, AutomationHTTPError
    import automation.adapters.statss as statss_mod
    
    hub_url = "https://www.statssa.gov.za/fake_hub"
    probed_url = "https://www.statssa.gov.za/fake_probe.xlsx"
    
    # 1. Strategy 1 (probe) succeeds
    def mock_probe():
        return probed_url
        
    # 2. Strategy 2 (hub scrape) fails due to WAF
    def mock_fetch_hub(client, url):
        raise AutomationHTTPError(url, 403, "WAF_BLOCKED: Incapsula WAF challenge detected")
        
    monkeypatch.setattr(statss_mod, "_fetch_release_hub_html", mock_fetch_hub)
    
    client = HTTPClient()
    
    # Act
    file_url, release_period, hub_html = statss_mod._resolve_publication_url(
        client=client,
        hub_url=hub_url,
        probe_func=mock_probe,
        fallback_period_label="Q1 2026",
    )
    
    # Assert - The probe URL is returned, WAF error is caught and logged
    assert file_url == probed_url
    assert release_period == "Q1 2026"
    assert hub_html is None

def test_download_publication_headers(monkeypatch):
    from automation.core.http_client import HTTPClient
    import automation.adapters.statss as statss_mod

    captured_kwargs = {}
    original_init = HTTPClient.__init__

    def mock_init(self, **kwargs):
        captured_kwargs.update(kwargs)
        # Call original init to avoid breaking things, but we only need the kwargs
        original_init(self, **kwargs)

    monkeypatch.setattr(HTTPClient, "__init__", mock_init)
    
    # Mock get to avoid network call
    monkeypatch.setattr(HTTPClient, "get", lambda self, url: type('Resp', (), {'body': b'A'*2048})())

    statss_mod._download_publication(HTTPClient(), "http://fake")

    headers = captured_kwargs.get("extra_headers", {})
    
    # Assert _STATSSA_BROWSER_HEADERS are present
    for k, v in statss_mod._STATSSA_BROWSER_HEADERS.items():
        if k != "Accept":
            assert headers.get(k) == v
        
    # Assert Accept is also present and correct
    assert headers.get("Accept") == "application/vnd.ms-excel,application/pdf,*/*"


# ---------------------------------------------------------------------------
# 7. IMPLEMENTATION-SPEC-STATSSA-WAF.md §8 — Tier 1 WAF-fallback tests
#
# Covers, for both _check_qlfs() and _check_gdp():
#   1. Hub WAF-blocked, direct-URL probe succeeds -> status="unknown".
#   2. Hub WAF-blocked, direct-URL probe also fails -> status="error"
#      (today's WAF_BLOCKED behaviour preserved).
#   3. Hub succeeds normally (no WAF marker) -> unchanged, and the new
#      fallback probe is provably NOT invoked (call-count assertion).
# Plus, separately:
#   4. _build_http_client() sends the new Tier 1 header set.
# ---------------------------------------------------------------------------


def _waf_response(hub_url: str) -> HTTPResponse:
    """An HTTPResponse whose body contains the Incapsula WAF marker."""
    return HTTPResponse(
        url=hub_url,
        status=200,
        headers={},
        body=b"<html><body>_Incapsula_Resource challenge page</body></html>",
        content_sha256="waf-challenge-hash",
    )


def _clean_response(hub_url: str, body: bytes) -> HTTPResponse:
    """An ordinary, non-WAF HTTPResponse."""
    return HTTPResponse(
        url=hub_url, status=200, headers={}, body=body, content_sha256="clean-hash",
    )


def test_check_qlfs_waf_blocked_fallback_probe_succeeds_returns_unknown(tmp_path, monkeypatch):
    adapter = _make_adapter(tmp_path)
    monkeypatch.setattr(
        "automation.core.http_client.HTTPClient.etag_check",
        lambda self, url, **kwargs: (True, _waf_response(statss_mod._QLFS_HUB_URL)),
    )
    found_url = "https://www.statssa.gov.za/publications/P0211/found.xlsx"
    monkeypatch.setattr(
        statss_mod, "_probe_qlfs_publication_url", lambda client, q, y: found_url
    )

    result = adapter.check_for_updates("unemployment", None)

    assert result.status == "unknown"
    assert "WAF_BLOCKED" in result.message
    assert "probe-based signal" in result.message
    assert found_url in result.notes


def test_check_qlfs_waf_blocked_fallback_probe_also_fails_returns_error(tmp_path, monkeypatch):
    adapter = _make_adapter(tmp_path)
    monkeypatch.setattr(
        "automation.core.http_client.HTTPClient.etag_check",
        lambda self, url, **kwargs: (True, _waf_response(statss_mod._QLFS_HUB_URL)),
    )
    monkeypatch.setattr(
        statss_mod, "_probe_qlfs_publication_url", lambda client, q, y: None
    )

    result = adapter.check_for_updates("unemployment", None)

    # Today's pre-Tier-1 behaviour is preserved exactly when the fallback
    # provides no signal either.
    assert result.status == "error"
    assert result.message == (
        "WAF_BLOCKED: Incapsula WAF challenge detected. Cannot check for updates."
    )


def test_check_qlfs_no_waf_fallback_probe_not_invoked(tmp_path, monkeypatch):
    adapter = _make_adapter(tmp_path)
    monkeypatch.setattr(
        "automation.core.http_client.HTTPClient.etag_check",
        lambda self, url, **kwargs: (
            True, _clean_response(statss_mod._QLFS_HUB_URL, b"<html>Q1 2026 QLFS release</html>"),
        ),
    )
    probe_calls: list[tuple] = []
    monkeypatch.setattr(
        statss_mod,
        "_probe_qlfs_publication_url",
        lambda client, q, y: probe_calls.append((q, y)) or "unused",
    )

    result = adapter.check_for_updates("unemployment", None)

    # Non-WAF path is completely unaffected by Tier 1: status still
    # computed from the hub diff, and the new fallback probe is never
    # reached.
    assert result.status == "update_available"
    assert probe_calls == []


def test_check_gdp_waf_blocked_fallback_probe_succeeds_returns_unknown(tmp_path, monkeypatch):
    adapter = _make_adapter_for_gdp(tmp_path)
    monkeypatch.setattr(
        "automation.core.http_client.HTTPClient.etag_check",
        lambda self, url, **kwargs: (True, _waf_response(statss_mod._GDP_HUB_URL)),
    )
    found_url = "https://www.statssa.gov.za/publications/P0441/found.xlsx"
    monkeypatch.setattr(
        statss_mod, "_probe_gdp_publication_url", lambda client, q, y: found_url
    )

    result = adapter.check_for_updates("gdp", None)

    assert result.status == "unknown"
    assert "WAF_BLOCKED" in result.message
    assert "probe-based signal" in result.message
    assert found_url in result.notes


def test_check_gdp_waf_blocked_fallback_probe_also_fails_returns_error(tmp_path, monkeypatch):
    adapter = _make_adapter_for_gdp(tmp_path)
    monkeypatch.setattr(
        "automation.core.http_client.HTTPClient.etag_check",
        lambda self, url, **kwargs: (True, _waf_response(statss_mod._GDP_HUB_URL)),
    )
    monkeypatch.setattr(
        statss_mod, "_probe_gdp_publication_url", lambda client, q, y: None
    )

    result = adapter.check_for_updates("gdp", None)

    assert result.status == "error"
    assert result.message == (
        "WAF_BLOCKED: Incapsula WAF challenge detected. Cannot check for updates."
    )


def test_check_gdp_no_waf_fallback_probe_not_invoked(tmp_path, monkeypatch):
    adapter = _make_adapter_for_gdp(tmp_path)
    monkeypatch.setattr(
        "automation.core.http_client.HTTPClient.etag_check",
        lambda self, url, **kwargs: (
            True, _clean_response(statss_mod._GDP_HUB_URL, b"<html>Q1 2026 GDP release</html>"),
        ),
    )
    probe_calls: list[tuple] = []
    monkeypatch.setattr(
        statss_mod,
        "_probe_gdp_publication_url",
        lambda client, q, y: probe_calls.append((q, y)) or "unused",
    )

    result = adapter.check_for_updates("gdp", None)

    assert result.status == "update_available"
    assert probe_calls == []


def test_build_http_client_sends_tier1_browser_headers():
    """IMPLEMENTATION-SPEC-STATSSA-WAF.md §8 item 4 — a direct assertion on
    the constructed client's headers, not a live-network test."""
    source_config = SourceConfig(source_id="statssa", display_name="Statistics South Africa")
    client = statss_mod._build_http_client(source_config)

    headers = client.extra_headers
    assert headers["User-Agent"] == statss_mod._STATSSA_BROWSER_HEADERS["User-Agent"]
    assert "data-automation-bot" not in headers["User-Agent"]
    assert headers["Accept-Language"] == "en-ZA,en;q=0.9"
    assert headers["Sec-Fetch-Mode"] == "navigate"
    # Accept-Encoding is deliberately "identity", not "gzip, deflate, br" —
    # core/http_client.py never decompresses a response body (out of scope
    # to change here), so advertising compression support would corrupt
    # the raw body text this adapter WAF-scans and parses.
    assert headers["Accept-Encoding"] == "identity"


# ---------------------------------------------------------------------------
# CPI (Phase 3b) — IMPLEMENTATION-SPEC-CPI.md §15
#
# 23 new tests, appended after the existing QLFS/GDP/WAF tests above (§15's
# 20 numbered items expand to 23 functions once item 6's "/" and item 15's
# "three tests" are counted individually — see the deviation note in the
# completion report). No existing test in this file is modified.
#
# Scope discipline mirrors the spec exactly: only cpi-headline and
# food-inflation are exercised here. repo-rate and annual-cpi-avg appear
# ONLY as the untouched "control" stats these tests prove were not
# modified — never as something this milestone's code is meant to update.
# ---------------------------------------------------------------------------

_CPI_HEADLINE_LABEL = "All items"
_CPI_FOOD_LABEL = "Food and non-alcoholic beverages"


def _build_cpi_fixture_workbook(
    headers: tuple[str, ...] = ("Mar 2026", "Apr 2026", "May 2026"),
    headline_values: tuple[float, ...] = (3.1, 4.0, 4.2),
    food_values: tuple[float, ...] = (3.6, 2.9, 2.7),
    include_food_row: bool = True,
) -> bytes:
    """
    Build a minimal, representative CPI-style workbook: a month-label
    header row plus "All items" and "Food and non-alcoholic beverages"
    indicator rows, with a value under each month column. Only the
    LATEST month's value is expected to be read by parse_cpi_workbook()
    — unlike the GDP fixture builder, this is a deliberate parallel to
    the QLFS fixture builder's single-latest-value philosophy, not GDP's
    multi-quarter one.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CPI Time Series"
    ws.cell(row=1, column=1, value="Indicator")
    for i, h in enumerate(headers, start=2):
        ws.cell(row=1, column=i, value=h)

    ws.cell(row=2, column=1, value=_CPI_HEADLINE_LABEL)
    for c, v in enumerate(headline_values, start=2):
        ws.cell(row=2, column=c, value=v)

    if include_food_row:
        ws.cell(row=3, column=1, value=_CPI_FOOD_LABEL)
        for c, v in enumerate(food_values, start=2):
            ws.cell(row=3, column=c, value=v)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# 1. Parser
# ---------------------------------------------------------------------------


def test_parse_cpi_workbook_extracts_both_metrics():
    data = _build_cpi_fixture_workbook()
    extract = parse_cpi_workbook(data)

    assert extract.release_period == "May 2026"
    assert extract.cpi_headline == 4.2
    assert extract.food_inflation == 2.7


def test_parse_cpi_workbook_missing_metric_fails_loudly():
    data = _build_cpi_fixture_workbook()
    wb = openpyxl.load_workbook(BytesIO(data))
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == _CPI_FOOD_LABEL:
                cell.value = "Some unrelated row"
    buf = BytesIO()
    wb.save(buf)

    with pytest.raises(ValueError, match="food_inflation"):
        parse_cpi_workbook(buf.getvalue())


def test_parse_cpi_workbook_no_month_headers_fails_loudly():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Indicator")
    ws.cell(row=1, column=2, value="Not a month")
    ws.cell(row=2, column=1, value=_CPI_HEADLINE_LABEL)
    ws.cell(row=2, column=2, value=4.2)
    buf = BytesIO()
    wb.save(buf)

    with pytest.raises(ValueError, match="month-header"):
        parse_cpi_workbook(buf.getvalue())


def test_parse_cpi_workbook_not_an_excel_file_fails_loudly():
    with pytest.raises(ValueError, match="Excel workbook"):
        parse_cpi_workbook(b"this is not a valid xlsx file")


# ---------------------------------------------------------------------------
# 2. Validation
# ---------------------------------------------------------------------------


def test_validate_cpi_rate_in_range_and_out_of_range():
    assert _validate_cpi_rate(4.2, _CPI_HEADLINE_STAT_ID) == []
    # A negative (deflationary) value is plausible for CPI — unlike QLFS's
    # [0, 100] percentage range — proving this validator genuinely differs
    # from _validate_percentage(), not merely re-using it.
    assert _validate_cpi_rate(-1.0, _CPI_HEADLINE_STAT_ID) == []

    errors = _validate_cpi_rate(55.0, _CPI_HEADLINE_STAT_ID)
    assert len(errors) == 1
    assert _CPI_HEADLINE_STAT_ID in errors[0]


def test_validate_monthly_label_ok():
    assert _validate_monthly_label("May 2026") == []


def test_validate_monthly_label_bad_format():
    errors = _validate_monthly_label("Q1 2026")
    assert len(errors) == 1
    assert "Q1 2026" in errors[0]


# ---------------------------------------------------------------------------
# 3. Month-over-month anomaly threshold (CPI-specific)
# ---------------------------------------------------------------------------


def test_check_cpi_jump_beyond_threshold_flags_anomaly_not_error():
    warning = _check_qoq_jump(3.1, 5.0, _CPI_HEADLINE_STAT_ID, threshold=_CPI_JUMP_WARNING_THRESHOLD)
    assert warning is not None
    assert "ANOMALY" in warning

    # A swing within CPI's own (narrower than GDP's, wider than nothing)
    # threshold is NOT flagged — proving _CPI_JUMP_WARNING_THRESHOLD is a
    # genuinely distinct value, not a re-use of GDP's or QLFS's default.
    assert _check_qoq_jump(3.1, 4.0, _CPI_HEADLINE_STAT_ID, threshold=_CPI_JUMP_WARNING_THRESHOLD) is None


# ---------------------------------------------------------------------------
# 4. _transform_inflation() / ownership boundary
# ---------------------------------------------------------------------------


def _cpi_headline_stat(series_data=None, **overrides) -> dict:
    stat = {
        "id": _CPI_HEADLINE_STAT_ID,
        "categoryId": "inflation",
        "title": "Headline Inflation (CPI)",
        "value": "3.1%",
        "rawValue": 3.1,
        "unit": "%",
        "change": 0.1,
        "changeLabel": "from Feb 2026",
        "trend": "up",
        "description": "Year-on-year change in the Consumer Price Index.",
        "source": {
            "name": "Statistics South Africa",
            "shortName": "Stats SA",
            "publicationName": "P0141",
            "publicationDate": "2026-04-21",
        },
        "lastUpdated": "2026-04-21",
        "series": [
            {
                "name": "Headline CPI (%, YoY)",
                "unit": "%",
                "data": series_data if series_data is not None else [
                    {"label": "Feb 2026", "value": 3.0},
                    {"label": "Mar 2026", "value": 3.1},
                ],
            }
        ],
    }
    stat.update(overrides)
    return stat


def _food_inflation_stat(series_data=None, **overrides) -> dict:
    stat = {
        "id": _CPI_FOOD_STAT_ID,
        "categoryId": "inflation",
        "title": "Food Inflation",
        "value": "3.6%",
        "rawValue": 3.6,
        "unit": "%",
        "change": -0.2,
        "changeLabel": "from Feb 2026",
        "trend": "down",
        "description": "Year-on-year change in food and non-alcoholic beverage prices.",
        "source": {
            "name": "Statistics South Africa",
            "shortName": "Stats SA",
            "publicationName": "P0141",
            "publicationDate": "2026-04-21",
        },
        "lastUpdated": "2026-04-21",
        "series": [
            {
                "name": "Food CPI (%, YoY)",
                "unit": "%",
                "data": series_data if series_data is not None else [
                    {"label": "Feb 2026", "value": 3.8},
                    {"label": "Mar 2026", "value": 3.6},
                ],
            }
        ],
    }
    stat.update(overrides)
    return stat


def _repo_rate_stat() -> dict:
    return {
        "id": "repo-rate",
        "categoryId": "inflation",
        "title": "Repo Rate",
        "value": "6.75%",
        "rawValue": 6.75,
        "unit": "%",
        "change": 0.0,
        "changeLabel": "unchanged since Jan 2026 MPC",
        "trend": "stable",
        "source": {"name": "South African Reserve Bank", "shortName": "SARB"},
        "lastUpdated": "2026-01-30",
        "series": [{"data": [{"label": "Jan 2026", "value": 6.75}]}],
    }


def _annual_cpi_avg_stat() -> dict:
    return {
        "id": "annual-cpi-avg",
        "categoryId": "inflation",
        "title": "Annual Average CPI",
        "value": "4.4%",
        "rawValue": 4.4,
        "unit": "%",
        "source": {"name": "Statistics South Africa", "shortName": "Stats SA"},
        "lastUpdated": "2026-01-21",
        "series": [{"data": [{"label": "2024", "value": 4.4}]}],
    }


def _inflation_doc(**overrides) -> dict:
    doc = {
        "_meta": {
            "source": "Statistics South Africa (CPI); South African Reserve Bank (repo rate)",
            "source_url": "https://www.statssa.gov.za/?page_id=1854&PPN=P0141",
            "update_frequency": "Monthly (CPI, ~22nd); repo rate updated after each MPC meeting (~6x per year)",
            "last_verified": "2026-04-30",
            "notes": "CPI figures are year-on-year percentage change. Repo rate reflects the SARB Monetary Policy Committee's most recent decision.",
        },
        "statistics": [
            _cpi_headline_stat(),
            _food_inflation_stat(),
            _repo_rate_stat(),
            _annual_cpi_avg_stat(),
        ],
    }
    doc.update(overrides)
    return doc


def test_transform_inflation_updates_only_cpi_headline_and_food_inflation():
    doc = _inflation_doc()
    extract = CPIExtract(
        release_period="Apr 2026",
        publication_date="2026-05-21",
        cpi_headline=4.0,
        food_inflation=2.9,
    )
    new_doc = _transform_inflation(doc, extract, "https://statssa.gov.za/cpi.xlsx")

    original_stats = {s["id"]: s for s in doc["statistics"]}
    new_stats = {s["id"]: s for s in new_doc["statistics"]}

    for stat_id in ("repo-rate", "annual-cpi-avg"):
        assert new_stats[stat_id] == original_stats[stat_id]

    assert new_stats[_CPI_HEADLINE_STAT_ID]["rawValue"] == 4.0
    assert new_stats[_CPI_FOOD_STAT_ID]["rawValue"] == 2.9
    # Deep copy — original input document is not mutated.
    assert doc["statistics"][0]["rawValue"] == 3.1


def test_transform_inflation_never_touches_repo_rate_value():
    """The single most important test in this milestone
    (IMPLEMENTATION-SPEC-CPI.md §7/§11 item 5) — the direct proof that
    repo-rate is never touched, even when a new CPI value happens to be
    numerically close to it.
    """
    doc = _inflation_doc()
    doc["statistics"][2]["rawValue"] = 4.0  # repo-rate deliberately set to
    # the same number extract.cpi_headline below will carry, so a
    # hypothetical bug that let rate_map leak into repo-rate would be
    # invisible under a naive "did it change" check but still fails a
    # direct value/id comparison against the original document.
    original_repo_rate = copy.deepcopy(doc["statistics"][2])

    extract = CPIExtract(
        release_period="Apr 2026",
        publication_date="2026-05-21",
        cpi_headline=4.0,
        food_inflation=2.9,
    )
    new_doc = _transform_inflation(doc, extract, "https://statssa.gov.za/cpi.xlsx")

    new_repo_rate = {s["id"]: s for s in new_doc["statistics"]}["repo-rate"]
    assert new_repo_rate == original_repo_rate


def test_assert_cpi_ownership_boundary_detects_repo_rate_tamper():
    previous = _inflation_doc()
    proposed = copy.deepcopy(previous)
    proposed["statistics"][2]["rawValue"] = 7.00  # tamper repo-rate

    violations = _assert_cpi_ownership_boundary(previous, proposed)
    assert len(violations) == 1
    assert "repo-rate" in violations[0]


def test_assert_cpi_ownership_boundary_detects_stat_removed_or_added():
    previous = _inflation_doc()

    proposed_removed = copy.deepcopy(previous)
    proposed_removed["statistics"] = [
        s for s in proposed_removed["statistics"] if s["id"] != "annual-cpi-avg"
    ]
    removed_violations = _assert_cpi_ownership_boundary(previous, proposed_removed)
    assert any("stat IDs" in v for v in removed_violations)

    proposed_added = copy.deepcopy(previous)
    proposed_added["statistics"].append({"id": "some-new-stat", "rawValue": 1.0})
    added_violations = _assert_cpi_ownership_boundary(previous, proposed_added)
    assert any("stat IDs" in v for v in added_violations)


def test_assert_cpi_ownership_boundary_passes_on_legitimate_cpi_only_change():
    previous = _inflation_doc()
    extract = CPIExtract(
        release_period="Apr 2026",
        publication_date="2026-05-21",
        cpi_headline=4.0,
        food_inflation=2.9,
    )
    proposed = _transform_inflation(previous, extract, "https://statssa.gov.za/cpi.xlsx")

    assert _assert_cpi_ownership_boundary(previous, proposed) == []


def test_update_cpi_meta_does_not_touch_source_or_notes():
    doc = _inflation_doc()
    original_meta = copy.deepcopy(doc["_meta"])

    _update_cpi_meta(doc, release_period="Apr 2026", publication_date="2026-05-21")

    assert doc["_meta"]["source"] == original_meta["source"]
    assert doc["_meta"]["source_url"] == original_meta["source_url"]
    assert doc["_meta"]["update_frequency"] == original_meta["update_frequency"]
    assert doc["_meta"]["notes"] == original_meta["notes"]
    assert "automation" in doc["_meta"]
    assert doc["_meta"]["automation"]["releasePeriod"] == "Apr 2026"


# ---------------------------------------------------------------------------
# 5. _check_cpi() — hub-change detection
# ---------------------------------------------------------------------------


def test_check_cpi_detects_hub_change(tmp_path, monkeypatch):
    adapter = _make_adapter(tmp_path)

    changed_response = HTTPResponse(
        url=statss_mod._CPI_HUB_URL,
        status=200,
        headers={},
        body=b"<html>Apr 2026 CPI release</html>",
        content_sha256="new-hash",
    )
    monkeypatch.setattr(
        "automation.core.http_client.HTTPClient.etag_check",
        lambda self, url, **kwargs: (True, changed_response),
    )
    result = adapter.check_for_updates("inflation", None)
    assert result.status == "update_available"

    # Fresh adapter (new cache), hub now reports unchanged.
    adapter2 = _make_adapter(tmp_path)
    unchanged_response = HTTPResponse(
        url=statss_mod._CPI_HUB_URL,
        status=200,
        headers={},
        body=b"<html>Apr 2026 CPI release</html>",
        content_sha256="new-hash",
    )
    monkeypatch.setattr(
        "automation.core.http_client.HTTPClient.etag_check",
        lambda self, url, **kwargs: (False, unchanged_response),
    )
    result2 = adapter2.check_for_updates("inflation", None)
    assert result2.status == "up_to_date"


# ---------------------------------------------------------------------------
# 6. fetch_and_apply() integration — network layer mocked, QLFS + GDP + CPI
# ---------------------------------------------------------------------------


def _patch_qlfs_gdp_cpi_network(
    monkeypatch,
    qlfs_bytes: bytes,
    qlfs_url: str,
    gdp_bytes: bytes,
    gdp_url: str,
    cpi_bytes: bytes,
    cpi_url: str,
) -> None:
    """Combined network patch, extending _patch_qlfs_and_gdp_network with a
    third, fully independent CPI discovery/download path within a single
    fetch_and_apply() call.
    """
    monkeypatch.setattr(
        statss_mod, "_fetch_release_hub_html",
        lambda client, url: b"<html>Q1 2026 QLFS release</html>",
    )
    monkeypatch.setattr(statss_mod, "_extract_release_period", lambda html: "Q1 2026")
    monkeypatch.setattr(statss_mod, "_determine_current_qlfs_quarter", lambda: (1, 2026))
    monkeypatch.setattr(statss_mod, "_probe_qlfs_publication_url", lambda client, q, y: qlfs_url)

    monkeypatch.setattr(statss_mod, "_determine_current_gdp_quarter", lambda: (1, 2026))
    monkeypatch.setattr(
        statss_mod, "_discover_gdp_excel",
        lambda client, **kwargs: (gdp_url, "Q1 2026", b"<html>Q1 2026 GDP release</html>"),
    )

    monkeypatch.setattr(statss_mod, "_determine_current_cpi_month", lambda: (4, 2026))
    monkeypatch.setattr(
        statss_mod, "_discover_cpi_excel",
        lambda client, **kwargs: (cpi_url, "Apr 2026", b"<html>Apr 2026 CPI release</html>"),
    )

    def _download(client, url):
        if url == qlfs_url:
            return qlfs_bytes
        if url == gdp_url:
            return gdp_bytes
        if url == cpi_url:
            return cpi_bytes
        raise AssertionError(f"Unexpected download URL in test: {url}")

    monkeypatch.setattr(statss_mod, "_download_publication", _download)


def test_fetch_and_apply_stages_cpi_without_direct_write(tmp_path, monkeypatch):
    qlfs_bytes = _build_fixture_workbook()
    qlfs_url = "https://www.statssa.gov.za/publications/P0211/fixture.xlsx"
    gdp_bytes = _build_gdp_fixture_workbook(
        headers=("Q4 2025", "Q1 2026"), values=(0.4, 0.5), include_annual_row=False,
    )
    gdp_url = "https://www.statssa.gov.za/publications/P0441/gdp_fixture.xlsx"
    cpi_bytes = _build_cpi_fixture_workbook(
        headers=("Mar 2026", "Apr 2026"), headline_values=(3.1, 4.0), food_values=(3.6, 2.9),
    )
    cpi_url = "https://www.statssa.gov.za/publications/P0141/cpi_fixture.xlsx"
    _patch_qlfs_gdp_cpi_network(
        monkeypatch, qlfs_bytes, qlfs_url, gdp_bytes, gdp_url, cpi_bytes, cpi_url,
    )

    prod_dir = tmp_path / "production"
    prod_dir.mkdir()
    stale_qlfs_docs, qlfs_paths = _stale_qlfs_docs_and_paths(prod_dir)
    monkeypatch.setattr(statss_mod, "_QLFS_DATASET_JSON", qlfs_paths)

    stale_gdp_doc = _gdp_doc()
    gdp_path = prod_dir / "gdp.json"
    gdp_path.write_text(json.dumps(stale_gdp_doc), encoding="utf-8")
    monkeypatch.setattr(statss_mod, "_GDP_DATASET_JSON", gdp_path)

    stale_inflation_doc = _inflation_doc()  # cpi-headline=3.1, food-inflation=3.6
    inflation_path = prod_dir / "inflation.json"
    inflation_path.write_text(json.dumps(stale_inflation_doc), encoding="utf-8")
    monkeypatch.setattr(statss_mod, "_CPI_DATASET_JSON", inflation_path)

    adapter = _make_adapter(tmp_path)
    result = adapter.fetch_and_apply(dry_run=False, run_id="test-run")

    # QLFS/GDP portions of the result are unaffected by CPI being added.
    assert result["status"] == "ok"
    assert not result["errors"]
    # result["version_ids"] is QLFS-only — GDP/CPI version_ids must never
    # be appended into this shared list (audit fix: version_id
    # cross-contamination).
    assert len(result["version_ids"]) == 3  # 3 QLFS datasets only
    for ds_id, p in qlfs_paths.items():
        assert json.loads(p.read_text(encoding="utf-8")) == stale_qlfs_docs[ds_id]
    assert json.loads(gdp_path.read_text(encoding="utf-8")) == stale_gdp_doc

    # CPI staged, no direct write to inflation.json.
    assert result["cpi"]["status"] == "ok"
    assert result["cpi"]["version_id"] is not None
    assert result["cpi"]["version_id"] not in result["version_ids"]
    assert result["gdp"]["version_id"] not in result["version_ids"]
    assert json.loads(inflation_path.read_text(encoding="utf-8")) == stale_inflation_doc

    pending = pending_versions(adapter.config.report_dir, "inflation")
    assert len(pending) == 1
    staged_doc = read_staged_dataset(adapter.config.report_dir, "inflation", pending[0].version_id)
    staged_by_id = {s["id"]: s for s in staged_doc["statistics"]}
    assert staged_by_id[_CPI_HEADLINE_STAT_ID]["rawValue"] == 4.0
    assert staged_by_id[_CPI_FOOD_STAT_ID]["rawValue"] == 2.9
    # repo-rate / annual-cpi-avg carried through unchanged in the staged doc.
    assert staged_by_id["repo-rate"] == stale_inflation_doc["statistics"][2]
    assert staged_by_id["annual-cpi-avg"] == stale_inflation_doc["statistics"][3]


def test_fetch_and_apply_cpi_no_change_produces_no_change_status(tmp_path, monkeypatch):
    qlfs_bytes = _build_fixture_workbook()
    qlfs_url = "https://www.statssa.gov.za/publications/P0211/fixture.xlsx"
    gdp_bytes = _build_gdp_fixture_workbook(
        headers=("Q3 2025", "Q4 2025"), values=(0.3, 0.4), include_annual_row=False,
    )
    gdp_url = "https://www.statssa.gov.za/publications/P0441/gdp_fixture.xlsx"
    cpi_bytes = _build_cpi_fixture_workbook(
        headers=("Feb 2026", "Mar 2026"), headline_values=(3.0, 3.1), food_values=(3.8, 3.6),
    )
    cpi_url = "https://www.statssa.gov.za/publications/P0141/cpi_fixture.xlsx"
    _patch_qlfs_gdp_cpi_network(
        monkeypatch, qlfs_bytes, qlfs_url, gdp_bytes, gdp_url, cpi_bytes, cpi_url,
    )

    prod_dir = tmp_path / "production"
    prod_dir.mkdir()
    _, qlfs_paths = _stale_qlfs_docs_and_paths(prod_dir)
    # Make the QLFS values already match the fixture so QLFS is no_change,
    # keeping this test focused on the CPI no_change path.
    for ds_id, doc in {
        "unemployment": {
            "_meta": {},
            "statistics": [
                {"id": "unemployment-national", "rawValue": 32.7,
                 "series": [{"data": [{"label": "Q1 2026", "value": 32.7}]}]}
            ],
        },
        "youth-unemployment": {
            "_meta": {},
            "statistics": [
                {"id": "youth-unemployment-narrow", "rawValue": 46.3, "series": []},
                {"id": "youth-unemployment-1524", "rawValue": 60.9, "series": []},
                {"id": "youth-unemployment-expanded", "rawValue": 58.1, "series": []},
                {"id": "youth-neet-rate", "rawValue": 37.6, "series": []},
            ],
        },
        "labour-force": {
            "_meta": {},
            "statistics": [
                {"id": "lfpr-overall", "rawValue": 60.5, "series": []},
                {"id": "female-labour-participation", "rawValue": 55.0, "series": []},
            ],
        },
    }.items():
        qlfs_paths[ds_id].write_text(json.dumps(doc), encoding="utf-8")
    monkeypatch.setattr(statss_mod, "_QLFS_DATASET_JSON", qlfs_paths)

    stale_gdp_doc = _gdp_doc()  # series already ends at Q3 2025: 0.3, Q4 2025: 0.4
    gdp_path = prod_dir / "gdp.json"
    gdp_path.write_text(json.dumps(stale_gdp_doc), encoding="utf-8")
    monkeypatch.setattr(statss_mod, "_GDP_DATASET_JSON", gdp_path)

    # cpi-headline=3.1, food-inflation=3.6 already match the fixture's
    # latest (Mar 2026) column.
    stale_inflation_doc = _inflation_doc()
    inflation_path = prod_dir / "inflation.json"
    inflation_path.write_text(json.dumps(stale_inflation_doc), encoding="utf-8")
    monkeypatch.setattr(statss_mod, "_CPI_DATASET_JSON", inflation_path)

    adapter = _make_adapter(tmp_path)
    result = adapter.fetch_and_apply(dry_run=False, run_id="test-run")

    assert result["cpi"]["status"] == "no_change"
    assert result["cpi"]["version_id"] is None
    assert pending_versions(adapter.config.report_dir, "inflation") == []
    assert json.loads(inflation_path.read_text(encoding="utf-8")) == stale_inflation_doc


def test_fetch_and_apply_cpi_protected_field_violation_aborts_only_that_dataset(tmp_path, monkeypatch):
    qlfs_bytes = _build_fixture_workbook()
    qlfs_url = "https://www.statssa.gov.za/publications/P0211/fixture.xlsx"
    gdp_bytes = _build_gdp_fixture_workbook(
        headers=("Q4 2025", "Q1 2026"), values=(0.4, 0.5), include_annual_row=False,
    )
    gdp_url = "https://www.statssa.gov.za/publications/P0441/gdp_fixture.xlsx"
    cpi_bytes = _build_cpi_fixture_workbook(
        headers=("Mar 2026", "Apr 2026"), headline_values=(3.1, 4.0), food_values=(3.6, 2.9),
    )
    cpi_url = "https://www.statssa.gov.za/publications/P0141/cpi_fixture.xlsx"
    _patch_qlfs_gdp_cpi_network(
        monkeypatch, qlfs_bytes, qlfs_url, gdp_bytes, gdp_url, cpi_bytes, cpi_url,
    )

    prod_dir = tmp_path / "production"
    prod_dir.mkdir()
    stale_qlfs_docs, qlfs_paths = _stale_qlfs_docs_and_paths(prod_dir)
    monkeypatch.setattr(statss_mod, "_QLFS_DATASET_JSON", qlfs_paths)

    stale_gdp_doc = _gdp_doc()
    gdp_path = prod_dir / "gdp.json"
    gdp_path.write_text(json.dumps(stale_gdp_doc), encoding="utf-8")
    monkeypatch.setattr(statss_mod, "_GDP_DATASET_JSON", gdp_path)

    stale_inflation_doc = _inflation_doc()
    inflation_path = prod_dir / "inflation.json"
    inflation_path.write_text(json.dumps(stale_inflation_doc), encoding="utf-8")
    monkeypatch.setattr(statss_mod, "_CPI_DATASET_JSON", inflation_path)

    # Sabotage the CPI transform to violate a protected field (id).
    original_transform_inflation = statss_mod._transform_inflation

    def _sabotaged_transform_inflation(current_doc, extract, source_url=""):
        doc = original_transform_inflation(current_doc, extract, source_url)
        doc["statistics"][0]["id"] = "cpi-headline-TAMPERED"
        return doc

    monkeypatch.setattr(statss_mod, "_transform_inflation", _sabotaged_transform_inflation)

    adapter = _make_adapter(tmp_path)
    result = adapter.fetch_and_apply(dry_run=False, run_id="test-run")

    # CPI failed due to the protected-field violation …
    assert result["cpi"]["status"] == "error"
    assert any("Protected field violation" in e for e in result["cpi"]["errors"])
    assert pending_versions(adapter.config.report_dir, "inflation") == []
    assert json.loads(inflation_path.read_text(encoding="utf-8")) == stale_inflation_doc

    # … while the QLFS and GDP portions of the SAME fetch_and_apply() call
    # still succeeded normally — direct proof the three flows are isolated.
    assert result["status"] == "ok"
    assert not result["errors"]
    assert result["gdp"]["status"] == "ok"
    for ds_id, p in qlfs_paths.items():
        assert json.loads(p.read_text(encoding="utf-8")) == stale_qlfs_docs[ds_id]


def test_fetch_and_apply_cpi_ownership_violation_aborts_staging(tmp_path, monkeypatch):
    """Proves _assert_cpi_ownership_boundary() catches what
    check_protected_fields() cannot: a non-id field (rawValue) on a
    non-owned stat (repo-rate) silently changing.
    """
    qlfs_bytes = _build_fixture_workbook()
    qlfs_url = "https://www.statssa.gov.za/publications/P0211/fixture.xlsx"
    gdp_bytes = _build_gdp_fixture_workbook(
        headers=("Q4 2025", "Q1 2026"), values=(0.4, 0.5), include_annual_row=False,
    )
    gdp_url = "https://www.statssa.gov.za/publications/P0441/gdp_fixture.xlsx"
    cpi_bytes = _build_cpi_fixture_workbook(
        headers=("Mar 2026", "Apr 2026"), headline_values=(3.1, 4.0), food_values=(3.6, 2.9),
    )
    cpi_url = "https://www.statssa.gov.za/publications/P0141/cpi_fixture.xlsx"
    _patch_qlfs_gdp_cpi_network(
        monkeypatch, qlfs_bytes, qlfs_url, gdp_bytes, gdp_url, cpi_bytes, cpi_url,
    )

    prod_dir = tmp_path / "production"
    prod_dir.mkdir()
    _, qlfs_paths = _stale_qlfs_docs_and_paths(prod_dir)
    monkeypatch.setattr(statss_mod, "_QLFS_DATASET_JSON", qlfs_paths)

    gdp_path = prod_dir / "gdp.json"
    gdp_path.write_text(json.dumps(_gdp_doc()), encoding="utf-8")
    monkeypatch.setattr(statss_mod, "_GDP_DATASET_JSON", gdp_path)

    stale_inflation_doc = _inflation_doc()
    inflation_path = prod_dir / "inflation.json"
    inflation_path.write_text(json.dumps(stale_inflation_doc), encoding="utf-8")
    monkeypatch.setattr(statss_mod, "_CPI_DATASET_JSON", inflation_path)

    # Sabotage the CPI transform to silently change repo-rate's rawValue
    # — not the id field, so check_protected_fields() alone would NOT
    # catch this.
    original_transform_inflation = statss_mod._transform_inflation

    def _sabotaged_transform_inflation(current_doc, extract, source_url=""):
        doc = original_transform_inflation(current_doc, extract, source_url)
        for stat in doc["statistics"]:
            if stat["id"] == "repo-rate":
                stat["rawValue"] = 7.00
        return doc

    monkeypatch.setattr(statss_mod, "_transform_inflation", _sabotaged_transform_inflation)

    adapter = _make_adapter(tmp_path)
    result = adapter.fetch_and_apply(dry_run=False, run_id="test-run")

    assert result["cpi"]["status"] == "error"
    assert any("ownership boundary" in e for e in result["cpi"]["errors"])
    assert pending_versions(adapter.config.report_dir, "inflation") == []
    assert json.loads(inflation_path.read_text(encoding="utf-8")) == stale_inflation_doc


def test_cpi_staged_candidate_requires_approve_then_promote(tmp_path, monkeypatch):
    """The CPI-specific equivalent of
    test_gdp_staged_candidate_requires_approve_then_promote — closes this
    acceptance-criterion class for CPI (IMPLEMENTATION-SPEC-CPI.md §15
    item 20).
    """
    qlfs_bytes = _build_fixture_workbook()
    qlfs_url = "https://www.statssa.gov.za/publications/P0211/fixture.xlsx"
    gdp_bytes = _build_gdp_fixture_workbook(
        headers=("Q4 2025", "Q1 2026"), values=(0.4, 0.5), include_annual_row=False,
    )
    gdp_url = "https://www.statssa.gov.za/publications/P0441/gdp_fixture.xlsx"
    cpi_bytes = _build_cpi_fixture_workbook(
        headers=("Mar 2026", "Apr 2026"), headline_values=(3.1, 4.0), food_values=(3.6, 2.9),
    )
    cpi_url = "https://www.statssa.gov.za/publications/P0141/cpi_fixture.xlsx"
    _patch_qlfs_gdp_cpi_network(
        monkeypatch, qlfs_bytes, qlfs_url, gdp_bytes, gdp_url, cpi_bytes, cpi_url,
    )

    prod_dir = tmp_path / "production"
    prod_dir.mkdir()
    _, qlfs_paths = _stale_qlfs_docs_and_paths(prod_dir)
    monkeypatch.setattr(statss_mod, "_QLFS_DATASET_JSON", qlfs_paths)

    gdp_path = prod_dir / "gdp.json"
    gdp_path.write_text(json.dumps(_gdp_doc()), encoding="utf-8")
    monkeypatch.setattr(statss_mod, "_GDP_DATASET_JSON", gdp_path)

    stale_inflation_doc = _inflation_doc()
    inflation_path = prod_dir / "inflation.json"
    inflation_path.write_text(json.dumps(stale_inflation_doc), encoding="utf-8")
    monkeypatch.setattr(statss_mod, "_CPI_DATASET_JSON", inflation_path)

    monkeypatch.setattr(
        "automation.core.promote.get_production_dataset_path",
        lambda dataset_id: {**qlfs_paths, "gdp": gdp_path, "inflation": inflation_path}[dataset_id],
    )

    adapter = _make_adapter(tmp_path)
    report_dir = adapter.config.report_dir

    # (a) Stage a genuine CPI change.
    result = adapter.fetch_and_apply(dry_run=False, run_id="test-run")
    assert result["cpi"]["status"] == "ok"

    pending = pending_versions(report_dir, "inflation")
    assert len(pending) == 1
    version_id = pending[0].version_id

    # inflation.json's on-disk content is unchanged immediately after staging.
    assert json.loads(inflation_path.read_text(encoding="utf-8")) == stale_inflation_doc

    # (b) Promotion must be refused before approval.
    with pytest.raises(ValueError, match="requires 'approved'"):
        promote_version(report_dir, "inflation", version_id)
    assert json.loads(inflation_path.read_text(encoding="utf-8")) == stale_inflation_doc

    # (c) Approve.
    approve_version(report_dir, "inflation", version_id, approver="test-reviewer")
    assert pending_versions(report_dir, "inflation") == []

    # (d) Promote — now allowed. The written file matches the staged document.
    staged_doc = read_staged_dataset(report_dir, "inflation", version_id)
    result_path = promote_version(report_dir, "inflation", version_id)
    assert result_path == inflation_path
    written = json.loads(inflation_path.read_text(encoding="utf-8"))
    assert written == staged_doc
    # (e) Only now — after promotion — has the on-disk content changed.
    assert written != stale_inflation_doc
    # (f) repo-rate / annual-cpi-avg are identical before and after
    # promotion — the ownership boundary held all the way through staging,
    # approval, and promotion.
    written_by_id = {s["id"]: s for s in written["statistics"]}
    stale_by_id = {s["id"]: s for s in stale_inflation_doc["statistics"]}
    assert written_by_id["repo-rate"] == stale_by_id["repo-rate"]
    assert written_by_id["annual-cpi-avg"] == stale_by_id["annual-cpi-avg"]


# ---------------------------------------------------------------------------
# 7. IMPLEMENTATION-SPEC-STATSSA-WAF.md-style Tier 1 WAF-fallback tests for
#    CPI — mirrors the QLFS/GDP WAF tests above exactly, extended to
#    _check_cpi() / _probe_cpi_publication_url().
# ---------------------------------------------------------------------------


def test_check_cpi_waf_blocked_fallback_probe_succeeds_returns_unknown(tmp_path, monkeypatch):
    adapter = _make_adapter(tmp_path)
    monkeypatch.setattr(
        "automation.core.http_client.HTTPClient.etag_check",
        lambda self, url, **kwargs: (True, _waf_response(statss_mod._CPI_HUB_URL)),
    )
    found_url = "https://www.statssa.gov.za/publications/P0141/found.xlsx"
    monkeypatch.setattr(
        statss_mod, "_probe_cpi_publication_url", lambda client, m, y: found_url
    )

    result = adapter.check_for_updates("inflation", None)

    assert result.status == "unknown"
    assert "WAF_BLOCKED" in result.message
    assert "probe-based signal" in result.message
    assert found_url in result.notes


def test_check_cpi_waf_blocked_fallback_probe_also_fails_returns_error(tmp_path, monkeypatch):
    adapter = _make_adapter(tmp_path)
    monkeypatch.setattr(
        "automation.core.http_client.HTTPClient.etag_check",
        lambda self, url, **kwargs: (True, _waf_response(statss_mod._CPI_HUB_URL)),
    )
    monkeypatch.setattr(
        statss_mod, "_probe_cpi_publication_url", lambda client, m, y: None
    )

    result = adapter.check_for_updates("inflation", None)

    assert result.status == "error"
    assert result.message == (
        "WAF_BLOCKED: Incapsula WAF challenge detected. Cannot check for updates."
    )


def test_check_cpi_no_waf_fallback_probe_not_invoked(tmp_path, monkeypatch):
    adapter = _make_adapter(tmp_path)
    monkeypatch.setattr(
        "automation.core.http_client.HTTPClient.etag_check",
        lambda self, url, **kwargs: (
            True, _clean_response(statss_mod._CPI_HUB_URL, b"<html>Apr 2026 CPI release</html>"),
        ),
    )
    probe_calls: list[tuple] = []
    monkeypatch.setattr(
        statss_mod,
        "_probe_cpi_publication_url",
        lambda client, m, y: probe_calls.append((m, y)) or "unused",
    )

    result = adapter.check_for_updates("inflation", None)

    assert result.status == "update_available"
    assert probe_calls == []


# ---------------------------------------------------------------------------
# Population (Phase 4) — IMPLEMENTATION-SPEC-POPULATION.md §13
#
# New tests appended after the existing QLFS/GDP/CPI/WAF tests above. No
# existing test in this file is modified.
#
# Scope discipline mirrors the spec exactly: only population-total is
# exercised here. population-urban and population-median-age appear ONLY
# as the untouched "control" stats these tests prove were not modified —
# never as something this milestone's code is meant to update. The
# source-guard tests are, per the spec, the highest-value tests in this
# section — they are the direct regression test against the actual
# historical World-Bank-sourced data-integrity bug.
# ---------------------------------------------------------------------------

_POPULATION_TOTAL_LABEL = "Total Population (RSA) South Africa"


def _build_population_fixture_workbook(
    headers: tuple[str, ...] = ("2024", "2025", "2026"),
    values: tuple[float, ...] = (62.4, 63.1, 63.8),
    include_total_row: bool = True,
) -> bytes:
    """
    Build a minimal, representative Population (MYPE)-style workbook: a
    bare-year header row plus a "Total Population (RSA) South Africa"
    indicator row, with a value under each year column. Only the LATEST
    year's value is expected to be read by parse_population_workbook() —
    a deliberate parallel to the QLFS/CPI fixture builders' single-
    latest-value philosophy, not GDP's multi-quarter one. Values default
    to the millions representation (§7.2); see
    _build_population_fixture_workbook_raw_headcount() for the raw-
    headcount representation used to test the parser's heuristic.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MYPE Time Series"
    ws.cell(row=1, column=1, value="Indicator")
    for i, h in enumerate(headers, start=2):
        ws.cell(row=1, column=i, value=int(h))

    if include_total_row:
        ws.cell(row=2, column=1, value=_POPULATION_TOTAL_LABEL)
        for c, v in enumerate(values, start=2):
            ws.cell(row=2, column=c, value=v)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# 1. Parser
# ---------------------------------------------------------------------------


def test_parse_population_workbook_extracts_total_population():
    data = _build_population_fixture_workbook()
    extract = parse_population_workbook(data)

    assert extract.release_period == "2026"
    assert extract.total_population_millions == 63.8


def test_parse_population_workbook_missing_metric_fails_loudly():
    data = _build_population_fixture_workbook()
    wb = openpyxl.load_workbook(BytesIO(data))
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == _POPULATION_TOTAL_LABEL:
                cell.value = "Some unrelated row"
    buf = BytesIO()
    wb.save(buf)

    with pytest.raises(ValueError, match="total population"):
        parse_population_workbook(buf.getvalue())


def test_parse_population_workbook_no_year_headers_fails_loudly():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Indicator")
    ws.cell(row=1, column=2, value="Not a year")
    ws.cell(row=2, column=1, value=_POPULATION_TOTAL_LABEL)
    ws.cell(row=2, column=2, value=63.1)
    buf = BytesIO()
    wb.save(buf)

    with pytest.raises(ValueError, match="year-header"):
        parse_population_workbook(buf.getvalue())


def test_parse_population_workbook_not_an_excel_file_fails_loudly():
    with pytest.raises(ValueError, match="Excel workbook"):
        parse_population_workbook(b"this is not a valid xlsx file")


def test_parse_population_workbook_millions_representation():
    """The workbook expresses the total already in millions (e.g. 63.1) —
    parse_population_workbook() must NOT divide this value further."""
    data = _build_population_fixture_workbook(
        headers=("2025", "2026"), values=(63.1, 63.8),
    )
    extract = parse_population_workbook(data)
    assert extract.total_population_millions == 63.8


def test_parse_population_workbook_raw_headcount_representation():
    """The workbook expresses the total as a raw headcount (e.g. 63800000)
    — parse_population_workbook() must divide by 1,000,000 per the
    documented heuristic (IMPLEMENTATION-SPEC-POPULATION.md §7.2 / §14
    item 2)."""
    data = _build_population_fixture_workbook(
        headers=("2025", "2026"), values=(63_100_000, 63_800_000),
    )
    extract = parse_population_workbook(data)
    assert extract.total_population_millions == pytest.approx(63.8)


# ---------------------------------------------------------------------------
# 2. Validation
# ---------------------------------------------------------------------------


def test_validate_population_total_in_range_and_out_of_range():
    assert _validate_population_total(63.8, _POPULATION_TOTAL_STAT_ID) == []

    errors = _validate_population_total(120.0, _POPULATION_TOTAL_STAT_ID)
    assert len(errors) == 1
    assert _POPULATION_TOTAL_STAT_ID in errors[0]

    errors_low = _validate_population_total(5.0, _POPULATION_TOTAL_STAT_ID)
    assert len(errors_low) == 1


def test_validate_annual_label_ok():
    assert _validate_annual_label("2026") == []


def test_validate_annual_label_bad_format():
    errors = _validate_annual_label("Q1 2026")
    assert len(errors) == 1
    assert "Q1 2026" in errors[0]


# ---------------------------------------------------------------------------
# 3. Year-over-year anomaly threshold (Population-specific)
# ---------------------------------------------------------------------------


def test_check_yoy_jump_within_threshold_no_warning():
    # 62.4 -> 63.1 is about +1.1% — within the 2.0% default threshold.
    assert _check_yoy_jump(62.4, 63.1, _POPULATION_TOTAL_STAT_ID) is None


def test_check_yoy_jump_beyond_threshold_flags_anomaly_not_error():
    # 64.0 -> 63.1 is about -1.4%, still within threshold; use a much
    # larger swing to exceed _POPULATION_JUMP_WARNING_THRESHOLD (2.0%) —
    # this is the swing IMPLEMENTATION-SPEC-POPULATION.md §14 item 7
    # expects on the one-time production correction run.
    warning = _check_yoy_jump(
        64.0, 60.0, _POPULATION_TOTAL_STAT_ID,
        threshold=_POPULATION_JUMP_WARNING_THRESHOLD,
    )
    assert warning is not None
    assert "ANOMALY" in warning


def test_check_yoy_jump_no_current_value_no_warning():
    assert _check_yoy_jump(None, 63.8, _POPULATION_TOTAL_STAT_ID) is None


# ---------------------------------------------------------------------------
# 4. Source guard — the highest-value tests in this milestone
#    (IMPLEMENTATION-SPEC-POPULATION.md §13)
# ---------------------------------------------------------------------------


def test_assert_population_source_guard_passes_for_statssa_domain():
    assert _assert_population_source_guard(
        "https://www.statssa.gov.za/publications/P0302/mype2026.xlsx"
    ) == []


def test_assert_population_source_guard_hard_fails_for_worldbank_domain():
    """The direct regression test against the actual historical bug — the
    single most important test in this spec (IMPLEMENTATION-SPEC-
    POPULATION.md §13)."""
    violations = _assert_population_source_guard(
        "https://data.worldbank.org/indicator/SP.POP.TOTL?locations=ZA"
    )
    assert len(violations) == 1
    assert "source guard violation" in violations[0]
    assert "statssa.gov.za" in violations[0]


def test_assert_population_source_guard_hard_fails_for_unrelated_domain():
    violations = _assert_population_source_guard("https://example.com/mype.xlsx")
    assert len(violations) == 1
    assert "source guard violation" in violations[0]


def test_assert_population_source_guard_passes_for_statssa_subdomain():
    assert _assert_population_source_guard(
        "https://census.statssa.gov.za/mype2026.xlsx"
    ) == []


# ---------------------------------------------------------------------------
# 5. _transform_population() / ownership boundary
# ---------------------------------------------------------------------------


def _population_total_stat(series_data=None, **overrides) -> dict:
    stat = {
        "id": _POPULATION_TOTAL_STAT_ID,
        "categoryId": "population",
        "title": "Total Population",
        "value": "64.0M",
        "rawValue": 64_000_000,
        "unit": "people",
        "change": 1.3,
        "changeLabel": "from 2023",
        "trend": "up",
        "description": "South Africa's estimated total population.",
        "source": {
            "name": "Statistics South Africa",
            "shortName": "Stats SA",
            "url": "https://www.statssa.gov.za/?page_id=1854&PPN=P0302",
            "publicationName": "Mid-Year Population Estimates 2024",
            "publicationDate": "2024-07-30",
        },
        "lastUpdated": "2024-07-30",
        "series": [
            {
                "name": "Population (millions)",
                "unit": "million",
                "data": series_data if series_data is not None else [
                    {"label": "2023", "value": 63.2},
                    {"label": "2024", "value": 64.0},
                ],
            }
        ],
    }
    stat.update(overrides)
    return stat


def _population_urban_stat() -> dict:
    return {
        "id": "population-urban",
        "categoryId": "population",
        "title": "Urban Population Share",
        "value": "63.7%",
        "rawValue": 63.7,
        "unit": "%",
        "change": 0.1,
        "changeLabel": "from 2023",
        "trend": "up",
        "source": {"name": "Statistics South Africa", "shortName": "Stats SA"},
        "lastUpdated": "2023-10-10",
    }


def _population_median_age_stat() -> dict:
    return {
        "id": "population-median-age",
        "categoryId": "population",
        "title": "Median Age",
        "value": "28.0",
        "rawValue": 28.0,
        "unit": "years",
        "change": 1.5,
        "changeLabel": "from Census 2011",
        "trend": "up",
        "source": {"name": "Statistics South Africa", "shortName": "Stats SA"},
        "lastUpdated": "2023-10-10",
    }


def _population_doc(**overrides) -> dict:
    doc = {
        "_meta": {
            "source": "Statistics South Africa — Census 2022; Mid-Year Population Estimates",
            "source_url": "https://www.statssa.gov.za/?page_id=1854&PPN=P0302",
            "update_frequency": "Census: decennial; Mid-year estimates: annual (July)",
            "last_verified": "2026-05-30",
            "notes": "Census 2022 is the definitive source.",
        },
        "statistics": [
            _population_total_stat(),
            _population_urban_stat(),
            _population_median_age_stat(),
        ],
    }
    doc.update(overrides)
    return doc


def test_apply_population_total_point_seeds_empty_series():
    doc = _population_doc()
    for stat in doc["statistics"]:
        if stat["id"] == _POPULATION_TOTAL_STAT_ID:
            stat["series"] = []

    _apply_population_total_point(
        doc, 63.8, release_period="2026", publication_date="2026-07-29",
    )

    stat = next(s for s in doc["statistics"] if s["id"] == _POPULATION_TOTAL_STAT_ID)
    assert stat["series"][0]["data"] == [{"label": "2026", "value": 63.8}]
    assert stat["rawValue"] == 63_800_000
    assert stat["value"] == "63.8M"


def test_apply_population_total_point_revises_existing_year():
    doc = _population_doc()
    _apply_population_total_point(
        doc, 63.3, release_period="2024", publication_date="2024-08-01",
    )
    stat = next(s for s in doc["statistics"] if s["id"] == _POPULATION_TOTAL_STAT_ID)
    labels = [pt["label"] for pt in stat["series"][0]["data"]]
    assert labels.count("2024") == 1
    revised = next(pt for pt in stat["series"][0]["data"] if pt["label"] == "2024")
    assert revised["value"] == 63.3


def test_apply_population_total_point_appends_new_year():
    doc = _population_doc()
    _apply_population_total_point(
        doc, 63.8, release_period="2026", publication_date="2026-07-29",
    )
    stat = next(s for s in doc["statistics"] if s["id"] == _POPULATION_TOTAL_STAT_ID)
    labels = [pt["label"] for pt in stat["series"][0]["data"]]
    assert labels == ["2023", "2024", "2026"]
    assert stat["rawValue"] == 63_800_000
    assert stat["value"] == "63.8M"
    assert stat["changeLabel"] == "from 2025"


def test_transform_population_only_touches_population_total():
    doc = _population_doc()
    extract = PopulationExtract(
        release_period="2026",
        publication_date="2026-07-29",
        total_population_millions=63.8,
        source_domain="www.statssa.gov.za",
    )
    new_doc = _transform_population(doc, extract, "https://www.statssa.gov.za/mype2026.xlsx")

    original_stats = {s["id"]: s for s in doc["statistics"]}
    new_stats = {s["id"]: s for s in new_doc["statistics"]}

    for stat_id in ("population-urban", "population-median-age"):
        assert new_stats[stat_id] == original_stats[stat_id]

    assert new_stats[_POPULATION_TOTAL_STAT_ID]["rawValue"] == 63_800_000
    # Deep copy — original input document is not mutated.
    assert doc["statistics"][0]["rawValue"] == 64_000_000


def test_assert_population_ownership_boundary_detects_tamper():
    previous = _population_doc()
    proposed = copy.deepcopy(previous)
    proposed["statistics"][1]["rawValue"] = 70.0  # tamper population-urban

    violations = _assert_population_ownership_boundary(previous, proposed)
    assert len(violations) == 1
    assert "population-urban" in violations[0]


def test_assert_population_ownership_boundary_detects_stat_removed_or_added():
    previous = _population_doc()

    proposed_removed = copy.deepcopy(previous)
    proposed_removed["statistics"] = [
        s for s in proposed_removed["statistics"] if s["id"] != "population-median-age"
    ]
    removed_violations = _assert_population_ownership_boundary(previous, proposed_removed)
    assert any("stat IDs" in v for v in removed_violations)

    proposed_added = copy.deepcopy(previous)
    proposed_added["statistics"].append({"id": "some-new-stat", "rawValue": 1.0})
    added_violations = _assert_population_ownership_boundary(previous, proposed_added)
    assert any("stat IDs" in v for v in added_violations)


def test_assert_population_ownership_boundary_passes_on_legitimate_change():
    previous = _population_doc()
    extract = PopulationExtract(
        release_period="2026",
        publication_date="2026-07-29",
        total_population_millions=63.8,
        source_domain="www.statssa.gov.za",
    )
    proposed = _transform_population(previous, extract, "https://www.statssa.gov.za/mype2026.xlsx")

    assert _assert_population_ownership_boundary(previous, proposed) == []


# ---------------------------------------------------------------------------
# 6. _check_population() — hub-change detection
# ---------------------------------------------------------------------------


def test_check_population_detects_hub_change(tmp_path, monkeypatch):
    adapter = _make_adapter(tmp_path)

    changed_response = HTTPResponse(
        url=statss_mod._POPULATION_HUB_URL,
        status=200,
        headers={},
        body=b"<html>2026 MYPE release</html>",
        content_sha256="new-hash",
    )
    monkeypatch.setattr(
        "automation.core.http_client.HTTPClient.etag_check",
        lambda self, url, **kwargs: (True, changed_response),
    )
    result = adapter.check_for_updates("population", None)
    assert result.status == "update_available"

    # Fresh adapter (new cache), hub now reports unchanged.
    adapter2 = _make_adapter(tmp_path)
    unchanged_response = HTTPResponse(
        url=statss_mod._POPULATION_HUB_URL,
        status=200,
        headers={},
        body=b"<html>2026 MYPE release</html>",
        content_sha256="new-hash",
    )
    monkeypatch.setattr(
        "automation.core.http_client.HTTPClient.etag_check",
        lambda self, url, **kwargs: (False, unchanged_response),
    )
    result2 = adapter2.check_for_updates("population", None)
    assert result2.status == "up_to_date"


# ---------------------------------------------------------------------------
# 7. fetch_and_apply() integration — network layer mocked,
#    QLFS + GDP + CPI + Population combined
# ---------------------------------------------------------------------------


def _patch_all_network(
    monkeypatch,
    qlfs_bytes: bytes,
    qlfs_url: str,
    gdp_bytes: bytes,
    gdp_url: str,
    cpi_bytes: bytes,
    cpi_url: str,
    population_bytes: bytes,
    population_url: str,
) -> None:
    """Combined network patch, extending _patch_qlfs_gdp_cpi_network with a
    fourth, fully independent Population discovery/download path within a
    single fetch_and_apply() call.
    """
    _patch_qlfs_gdp_cpi_network(
        monkeypatch, qlfs_bytes, qlfs_url, gdp_bytes, gdp_url, cpi_bytes, cpi_url,
    )

    monkeypatch.setattr(statss_mod, "_determine_current_population_year", lambda: 2026)
    monkeypatch.setattr(
        statss_mod, "_discover_population_excel",
        lambda client, **kwargs: (population_url, "2026", b"<html>2026 MYPE release</html>"),
    )

    def _download(client, url):
        if url == qlfs_url:
            return qlfs_bytes
        if url == gdp_url:
            return gdp_bytes
        if url == cpi_url:
            return cpi_bytes
        if url == population_url:
            return population_bytes
        raise AssertionError(f"Unexpected download URL in test: {url}")

    monkeypatch.setattr(statss_mod, "_download_publication", _download)


def _setup_prod_dir_for_all_flows(tmp_path, monkeypatch):
    """Write stale QLFS/GDP/CPI production docs to disk and monkeypatch
    their dataset-JSON paths, returning (prod_dir, stale docs/paths) so
    each Population test can focus on population.json without repeating
    this boilerplate."""
    prod_dir = tmp_path / "production"
    prod_dir.mkdir()
    stale_qlfs_docs, qlfs_paths = _stale_qlfs_docs_and_paths(prod_dir)
    monkeypatch.setattr(statss_mod, "_QLFS_DATASET_JSON", qlfs_paths)

    stale_gdp_doc = _gdp_doc()
    gdp_path = prod_dir / "gdp.json"
    gdp_path.write_text(json.dumps(stale_gdp_doc), encoding="utf-8")
    monkeypatch.setattr(statss_mod, "_GDP_DATASET_JSON", gdp_path)

    stale_inflation_doc = _inflation_doc()
    inflation_path = prod_dir / "inflation.json"
    inflation_path.write_text(json.dumps(stale_inflation_doc), encoding="utf-8")
    monkeypatch.setattr(statss_mod, "_CPI_DATASET_JSON", inflation_path)

    return prod_dir, qlfs_paths, gdp_path, inflation_path


def test_fetch_and_apply_stages_population_without_direct_write(tmp_path, monkeypatch):
    qlfs_bytes = _build_fixture_workbook()
    qlfs_url = "https://www.statssa.gov.za/publications/P0211/fixture.xlsx"
    gdp_bytes = _build_gdp_fixture_workbook(
        headers=("Q4 2025", "Q1 2026"), values=(0.4, 0.5), include_annual_row=False,
    )
    gdp_url = "https://www.statssa.gov.za/publications/P0441/gdp_fixture.xlsx"
    cpi_bytes = _build_cpi_fixture_workbook(
        headers=("Mar 2026", "Apr 2026"), headline_values=(3.1, 4.0), food_values=(3.6, 2.9),
    )
    cpi_url = "https://www.statssa.gov.za/publications/P0141/cpi_fixture.xlsx"
    population_bytes = _build_population_fixture_workbook(
        headers=("2025", "2026"), values=(63.1, 63.8),
    )
    population_url = "https://www.statssa.gov.za/publications/P0302/population_fixture.xlsx"
    _patch_all_network(
        monkeypatch, qlfs_bytes, qlfs_url, gdp_bytes, gdp_url, cpi_bytes, cpi_url,
        population_bytes, population_url,
    )

    prod_dir, qlfs_paths, gdp_path, inflation_path = _setup_prod_dir_for_all_flows(
        tmp_path, monkeypatch
    )

    stale_population_doc = _population_doc()  # population-total rawValue=64,000,000
    population_path = prod_dir / "population.json"
    population_path.write_text(json.dumps(stale_population_doc), encoding="utf-8")
    monkeypatch.setattr(statss_mod, "_POPULATION_DATASET_JSON", population_path)

    adapter = _make_adapter(tmp_path)
    result = adapter.fetch_and_apply(dry_run=False, run_id="test-run")

    # QLFS/GDP/CPI portions of the result are unaffected by Population
    # being added.
    assert result["status"] == "ok"
    assert not result["errors"]
    # result["version_ids"] is QLFS-only — GDP/CPI/Population version_ids
    # must never be appended into this shared list (audit fix: version_id
    # cross-contamination).
    assert len(result["version_ids"]) == 3  # 3 QLFS datasets only

    # Population staged, no direct write to population.json.
    assert result["population"]["status"] == "ok"
    assert result["population"]["version_id"] is not None
    assert result["population"]["version_id"] not in result["version_ids"]
    assert result["gdp"]["version_id"] not in result["version_ids"]
    assert result["cpi"]["version_id"] not in result["version_ids"]
    assert json.loads(population_path.read_text(encoding="utf-8")) == stale_population_doc

    pending = pending_versions(adapter.config.report_dir, "population")
    assert len(pending) == 1
    staged_doc = read_staged_dataset(adapter.config.report_dir, "population", pending[0].version_id)
    staged_by_id = {s["id"]: s for s in staged_doc["statistics"]}
    assert staged_by_id[_POPULATION_TOTAL_STAT_ID]["rawValue"] == 63_800_000
    # population-urban / population-median-age carried through unchanged
    # in the staged doc.
    assert staged_by_id["population-urban"] == stale_population_doc["statistics"][1]
    assert staged_by_id["population-median-age"] == stale_population_doc["statistics"][2]


def test_fetch_and_apply_population_no_change_produces_no_change_status(tmp_path, monkeypatch):
    qlfs_bytes = _build_fixture_workbook()
    qlfs_url = "https://www.statssa.gov.za/publications/P0211/fixture.xlsx"
    gdp_bytes = _build_gdp_fixture_workbook(
        headers=("Q3 2025", "Q4 2025"), values=(0.3, 0.4), include_annual_row=False,
    )
    gdp_url = "https://www.statssa.gov.za/publications/P0441/gdp_fixture.xlsx"
    cpi_bytes = _build_cpi_fixture_workbook(
        headers=("Feb 2026", "Mar 2026"), headline_values=(3.0, 3.1), food_values=(3.8, 3.6),
    )
    cpi_url = "https://www.statssa.gov.za/publications/P0141/cpi_fixture.xlsx"
    population_bytes = _build_population_fixture_workbook(
        headers=("2025", "2026"), values=(63.1, 64.0),
    )
    population_url = "https://www.statssa.gov.za/publications/P0302/population_fixture.xlsx"
    _patch_all_network(
        monkeypatch, qlfs_bytes, qlfs_url, gdp_bytes, gdp_url, cpi_bytes, cpi_url,
        population_bytes, population_url,
    )

    prod_dir = tmp_path / "production"
    prod_dir.mkdir()
    # Make QLFS values already match the fixture so QLFS/GDP/CPI are all
    # no_change, keeping this test focused on the Population no_change path.
    qlfs_paths = {}
    for ds_id, doc in {
        "unemployment": {
            "_meta": {},
            "statistics": [
                {"id": "unemployment-national", "rawValue": 32.7,
                 "series": [{"data": [{"label": "Q1 2026", "value": 32.7}]}]}
            ],
        },
        "youth-unemployment": {
            "_meta": {},
            "statistics": [
                {"id": "youth-unemployment-narrow", "rawValue": 46.3, "series": []},
                {"id": "youth-unemployment-1524", "rawValue": 60.9, "series": []},
                {"id": "youth-unemployment-expanded", "rawValue": 58.1, "series": []},
                {"id": "youth-neet-rate", "rawValue": 37.6, "series": []},
            ],
        },
        "labour-force": {
            "_meta": {},
            "statistics": [
                {"id": "lfpr-overall", "rawValue": 60.5, "series": []},
                {"id": "female-labour-participation", "rawValue": 55.0, "series": []},
            ],
        },
    }.items():
        p = prod_dir / f"{ds_id}.json"
        p.write_text(json.dumps(doc), encoding="utf-8")
        qlfs_paths[ds_id] = p
    monkeypatch.setattr(statss_mod, "_QLFS_DATASET_JSON", qlfs_paths)

    stale_gdp_doc = _gdp_doc()  # series already ends at Q3 2025: 0.3, Q4 2025: 0.4
    gdp_path = prod_dir / "gdp.json"
    gdp_path.write_text(json.dumps(stale_gdp_doc), encoding="utf-8")
    monkeypatch.setattr(statss_mod, "_GDP_DATASET_JSON", gdp_path)

    stale_inflation_doc = _inflation_doc()  # cpi-headline=3.1, food-inflation=3.6
    inflation_path = prod_dir / "inflation.json"
    inflation_path.write_text(json.dumps(stale_inflation_doc), encoding="utf-8")
    monkeypatch.setattr(statss_mod, "_CPI_DATASET_JSON", inflation_path)

    # population-total rawValue already matches the fixture's latest
    # (2026) column: 64.0M == 64,000,000.
    stale_population_doc = _population_doc()
    population_path = prod_dir / "population.json"
    population_path.write_text(json.dumps(stale_population_doc), encoding="utf-8")
    monkeypatch.setattr(statss_mod, "_POPULATION_DATASET_JSON", population_path)

    adapter = _make_adapter(tmp_path)
    result = adapter.fetch_and_apply(dry_run=False, run_id="test-run")

    assert result["population"]["status"] == "no_change"
    assert result["population"]["version_id"] is None
    assert pending_versions(adapter.config.report_dir, "population") == []
    assert json.loads(population_path.read_text(encoding="utf-8")) == stale_population_doc


def test_fetch_and_apply_population_protected_field_violation_aborts_only_that_dataset(tmp_path, monkeypatch):
    qlfs_bytes = _build_fixture_workbook()
    qlfs_url = "https://www.statssa.gov.za/publications/P0211/fixture.xlsx"
    gdp_bytes = _build_gdp_fixture_workbook(
        headers=("Q4 2025", "Q1 2026"), values=(0.4, 0.5), include_annual_row=False,
    )
    gdp_url = "https://www.statssa.gov.za/publications/P0441/gdp_fixture.xlsx"
    cpi_bytes = _build_cpi_fixture_workbook(
        headers=("Mar 2026", "Apr 2026"), headline_values=(3.1, 4.0), food_values=(3.6, 2.9),
    )
    cpi_url = "https://www.statssa.gov.za/publications/P0141/cpi_fixture.xlsx"
    population_bytes = _build_population_fixture_workbook(
        headers=("2025", "2026"), values=(63.1, 63.8),
    )
    population_url = "https://www.statssa.gov.za/publications/P0302/population_fixture.xlsx"
    _patch_all_network(
        monkeypatch, qlfs_bytes, qlfs_url, gdp_bytes, gdp_url, cpi_bytes, cpi_url,
        population_bytes, population_url,
    )

    prod_dir, qlfs_paths, gdp_path, inflation_path = _setup_prod_dir_for_all_flows(
        tmp_path, monkeypatch
    )

    stale_population_doc = _population_doc()
    population_path = prod_dir / "population.json"
    population_path.write_text(json.dumps(stale_population_doc), encoding="utf-8")
    monkeypatch.setattr(statss_mod, "_POPULATION_DATASET_JSON", population_path)

    # Sabotage the Population transform to violate a protected field (id).
    original_transform_population = statss_mod._transform_population

    def _sabotaged_transform_population(current_doc, extract, source_url=""):
        doc = original_transform_population(current_doc, extract, source_url)
        doc["statistics"][0]["id"] = "population-total-TAMPERED"
        return doc

    monkeypatch.setattr(statss_mod, "_transform_population", _sabotaged_transform_population)

    adapter = _make_adapter(tmp_path)
    result = adapter.fetch_and_apply(dry_run=False, run_id="test-run")

    # Population failed due to the protected-field violation …
    assert result["population"]["status"] == "error"
    assert any("Protected field violation" in e for e in result["population"]["errors"])
    assert pending_versions(adapter.config.report_dir, "population") == []
    assert json.loads(population_path.read_text(encoding="utf-8")) == stale_population_doc

    # … while the QLFS/GDP/CPI portions of the SAME fetch_and_apply() call
    # still succeeded normally — direct proof the four flows are isolated.
    assert result["status"] == "ok"
    assert not result["errors"]
    assert result["gdp"]["status"] == "ok"
    assert result["cpi"]["status"] == "ok"


def test_fetch_and_apply_population_ownership_violation_aborts_staging(tmp_path, monkeypatch):
    """Proves _assert_population_ownership_boundary() catches what
    check_protected_fields() cannot: a non-id field (rawValue) on a
    non-owned stat (population-urban) silently changing.
    """
    qlfs_bytes = _build_fixture_workbook()
    qlfs_url = "https://www.statssa.gov.za/publications/P0211/fixture.xlsx"
    gdp_bytes = _build_gdp_fixture_workbook(
        headers=("Q4 2025", "Q1 2026"), values=(0.4, 0.5), include_annual_row=False,
    )
    gdp_url = "https://www.statssa.gov.za/publications/P0441/gdp_fixture.xlsx"
    cpi_bytes = _build_cpi_fixture_workbook(
        headers=("Mar 2026", "Apr 2026"), headline_values=(3.1, 4.0), food_values=(3.6, 2.9),
    )
    cpi_url = "https://www.statssa.gov.za/publications/P0141/cpi_fixture.xlsx"
    population_bytes = _build_population_fixture_workbook(
        headers=("2025", "2026"), values=(63.1, 63.8),
    )
    population_url = "https://www.statssa.gov.za/publications/P0302/population_fixture.xlsx"
    _patch_all_network(
        monkeypatch, qlfs_bytes, qlfs_url, gdp_bytes, gdp_url, cpi_bytes, cpi_url,
        population_bytes, population_url,
    )

    prod_dir, qlfs_paths, gdp_path, inflation_path = _setup_prod_dir_for_all_flows(
        tmp_path, monkeypatch
    )

    stale_population_doc = _population_doc()
    population_path = prod_dir / "population.json"
    population_path.write_text(json.dumps(stale_population_doc), encoding="utf-8")
    monkeypatch.setattr(statss_mod, "_POPULATION_DATASET_JSON", population_path)

    # Sabotage the Population transform to silently change
    # population-urban's rawValue — not the id field, so
    # check_protected_fields() alone would NOT catch this.
    original_transform_population = statss_mod._transform_population

    def _sabotaged_transform_population(current_doc, extract, source_url=""):
        doc = original_transform_population(current_doc, extract, source_url)
        for stat in doc["statistics"]:
            if stat["id"] == "population-urban":
                stat["rawValue"] = 70.0
        return doc

    monkeypatch.setattr(statss_mod, "_transform_population", _sabotaged_transform_population)

    adapter = _make_adapter(tmp_path)
    result = adapter.fetch_and_apply(dry_run=False, run_id="test-run")

    assert result["population"]["status"] == "error"
    assert any("ownership boundary" in e for e in result["population"]["errors"])
    assert pending_versions(adapter.config.report_dir, "population") == []
    assert json.loads(population_path.read_text(encoding="utf-8")) == stale_population_doc


def test_fetch_and_apply_population_non_statssa_source_never_stages(tmp_path, monkeypatch):
    """The direct end-to-end proof of the source guard
    (IMPLEMENTATION-SPEC-POPULATION.md §13): a mocked non-statssa.gov.za
    discovery result must never reach _transform_population() and must
    never stage anything.
    """
    qlfs_bytes = _build_fixture_workbook()
    qlfs_url = "https://www.statssa.gov.za/publications/P0211/fixture.xlsx"
    gdp_bytes = _build_gdp_fixture_workbook(
        headers=("Q4 2025", "Q1 2026"), values=(0.4, 0.5), include_annual_row=False,
    )
    gdp_url = "https://www.statssa.gov.za/publications/P0441/gdp_fixture.xlsx"
    cpi_bytes = _build_cpi_fixture_workbook(
        headers=("Mar 2026", "Apr 2026"), headline_values=(3.1, 4.0), food_values=(3.6, 2.9),
    )
    cpi_url = "https://www.statssa.gov.za/publications/P0141/cpi_fixture.xlsx"
    population_bytes = _build_population_fixture_workbook(
        headers=("2025", "2026"), values=(63.1, 63.8),
    )
    # Deliberately a World Bank URL — the exact wrong source the sourcing
    # plan identified as the historical bug.
    population_url = "https://data.worldbank.org/indicator/mype2026.xlsx"
    _patch_all_network(
        monkeypatch, qlfs_bytes, qlfs_url, gdp_bytes, gdp_url, cpi_bytes, cpi_url,
        population_bytes, population_url,
    )

    prod_dir, qlfs_paths, gdp_path, inflation_path = _setup_prod_dir_for_all_flows(
        tmp_path, monkeypatch
    )

    stale_population_doc = _population_doc()
    population_path = prod_dir / "population.json"
    population_path.write_text(json.dumps(stale_population_doc), encoding="utf-8")
    monkeypatch.setattr(statss_mod, "_POPULATION_DATASET_JSON", population_path)

    transform_calls: list = []
    original_transform_population = statss_mod._transform_population

    def _tracking_transform_population(current_doc, extract, source_url=""):
        transform_calls.append(source_url)
        return original_transform_population(current_doc, extract, source_url)

    monkeypatch.setattr(statss_mod, "_transform_population", _tracking_transform_population)

    adapter = _make_adapter(tmp_path)
    result = adapter.fetch_and_apply(dry_run=False, run_id="test-run")

    assert result["population"]["status"] == "error"
    assert any("source guard" in e.lower() for e in result["population"]["errors"])
    assert transform_calls == []
    assert pending_versions(adapter.config.report_dir, "population") == []
    assert json.loads(population_path.read_text(encoding="utf-8")) == stale_population_doc

    # QLFS/GDP/CPI are unaffected by the Population source-guard failure.
    assert result["status"] == "ok"
    assert not result["errors"]
    assert result["gdp"]["status"] == "ok"
    assert result["cpi"]["status"] == "ok"


def test_population_staged_candidate_requires_approve_then_promote(tmp_path, monkeypatch):
    """The Population-specific equivalent of
    test_cpi_staged_candidate_requires_approve_then_promote — closes this
    acceptance-criterion class for population (IMPLEMENTATION-SPEC-
    POPULATION.md §13/§15 item 7). This is also the vehicle proving the
    one-time production correction (§9) works end-to-end.
    """
    qlfs_bytes = _build_fixture_workbook()
    qlfs_url = "https://www.statssa.gov.za/publications/P0211/fixture.xlsx"
    gdp_bytes = _build_gdp_fixture_workbook(
        headers=("Q4 2025", "Q1 2026"), values=(0.4, 0.5), include_annual_row=False,
    )
    gdp_url = "https://www.statssa.gov.za/publications/P0441/gdp_fixture.xlsx"
    cpi_bytes = _build_cpi_fixture_workbook(
        headers=("Mar 2026", "Apr 2026"), headline_values=(3.1, 4.0), food_values=(3.6, 2.9),
    )
    cpi_url = "https://www.statssa.gov.za/publications/P0141/cpi_fixture.xlsx"
    # The one-time production correction: this run's swing is deliberately
    # large enough to exceed _POPULATION_JUMP_WARNING_THRESHOLD (2.0%) —
    # per IMPLEMENTATION-SPEC-POPULATION.md §14 item 7, this is expected
    # and correct on such a run, not a parser bug.
    population_bytes = _build_population_fixture_workbook(
        headers=("2024", "2025"), values=(64.0, 61.0),
    )
    population_url = "https://www.statssa.gov.za/publications/P0302/population_fixture.xlsx"
    _patch_all_network(
        monkeypatch, qlfs_bytes, qlfs_url, gdp_bytes, gdp_url, cpi_bytes, cpi_url,
        population_bytes, population_url,
    )

    prod_dir, qlfs_paths, gdp_path, inflation_path = _setup_prod_dir_for_all_flows(
        tmp_path, monkeypatch
    )

    stale_population_doc = _population_doc()  # population-total = 64.0M (2024)
    population_path = prod_dir / "population.json"
    population_path.write_text(json.dumps(stale_population_doc), encoding="utf-8")
    monkeypatch.setattr(statss_mod, "_POPULATION_DATASET_JSON", population_path)

    monkeypatch.setattr(
        "automation.core.promote.get_production_dataset_path",
        lambda dataset_id: {
            **qlfs_paths, "gdp": gdp_path, "inflation": inflation_path,
            "population": population_path,
        }[dataset_id],
    )

    adapter = _make_adapter(tmp_path)
    report_dir = adapter.config.report_dir

    # (a) Stage a genuine Population change — the one-time correction.
    result = adapter.fetch_and_apply(dry_run=False, run_id="test-run")
    assert result["population"]["status"] == "ok"
    # The year-over-year anomaly check fires on this specific run, per
    # §14 item 7 — this is expected and correct, not a parser bug.
    assert "ANOMALY" in result["population"]["notes"]

    pending = pending_versions(report_dir, "population")
    assert len(pending) == 1
    version_id = pending[0].version_id

    # population.json's on-disk content is unchanged immediately after staging.
    assert json.loads(population_path.read_text(encoding="utf-8")) == stale_population_doc

    # (b) Promotion must be refused before approval.
    with pytest.raises(ValueError, match="requires 'approved'"):
        promote_version(report_dir, "population", version_id)
    assert json.loads(population_path.read_text(encoding="utf-8")) == stale_population_doc

    # (c) Approve.
    approve_version(report_dir, "population", version_id, approver="test-reviewer")
    assert pending_versions(report_dir, "population") == []

    # (d) Promote — now allowed. The written file matches the staged document.
    staged_doc = read_staged_dataset(report_dir, "population", version_id)
    result_path = promote_version(report_dir, "population", version_id)
    assert result_path == population_path
    written = json.loads(population_path.read_text(encoding="utf-8"))
    assert written == staged_doc
    # (e) Only now — after promotion — has the on-disk content changed,
    # and it reflects the corrected figure, not the stale 64.0M.
    assert written != stale_population_doc
    written_by_id = {s["id"]: s for s in written["statistics"]}
    assert written_by_id[_POPULATION_TOTAL_STAT_ID]["rawValue"] == 61_000_000
    # (f) population-urban / population-median-age are identical before
    # and after promotion — the ownership boundary held all the way
    # through staging, approval, and promotion.
    stale_by_id = {s["id"]: s for s in stale_population_doc["statistics"]}
    assert written_by_id["population-urban"] == stale_by_id["population-urban"]
    assert written_by_id["population-median-age"] == stale_by_id["population-median-age"]


# ---------------------------------------------------------------------------
# 8. IMPLEMENTATION-SPEC-STATSSA-WAF.md-style Tier 1 WAF-fallback tests for
#    Population — mirrors the QLFS/GDP/CPI WAF tests above exactly,
#    extended to _check_population() / _probe_population_publication_url().
# ---------------------------------------------------------------------------


def test_check_population_waf_blocked_fallback_probe_succeeds_returns_unknown(tmp_path, monkeypatch):
    adapter = _make_adapter(tmp_path)
    monkeypatch.setattr(
        "automation.core.http_client.HTTPClient.etag_check",
        lambda self, url, **kwargs: (True, _waf_response(statss_mod._POPULATION_HUB_URL)),
    )
    found_url = "https://www.statssa.gov.za/publications/P0302/found.xlsx"
    monkeypatch.setattr(
        statss_mod, "_probe_population_publication_url", lambda client, y: found_url
    )

    result = adapter.check_for_updates("population", None)

    assert result.status == "unknown"
    assert "WAF_BLOCKED" in result.message
    assert "probe-based signal" in result.message
    assert found_url in result.notes


def test_check_population_waf_blocked_fallback_probe_also_fails_returns_error(tmp_path, monkeypatch):
    adapter = _make_adapter(tmp_path)
    monkeypatch.setattr(
        "automation.core.http_client.HTTPClient.etag_check",
        lambda self, url, **kwargs: (True, _waf_response(statss_mod._POPULATION_HUB_URL)),
    )
    monkeypatch.setattr(
        statss_mod, "_probe_population_publication_url", lambda client, y: None
    )

    result = adapter.check_for_updates("population", None)

    assert result.status == "error"
    assert result.message == (
        "WAF_BLOCKED: Incapsula WAF challenge detected. Cannot check for updates."
    )


def test_check_population_no_waf_fallback_probe_not_invoked(tmp_path, monkeypatch):
    adapter = _make_adapter(tmp_path)
    monkeypatch.setattr(
        "automation.core.http_client.HTTPClient.etag_check",
        lambda self, url, **kwargs: (
            True, _clean_response(statss_mod._POPULATION_HUB_URL, b"<html>2026 MYPE release</html>"),
        ),
    )
    probe_calls: list[tuple] = []
    monkeypatch.setattr(
        statss_mod,
        "_probe_population_publication_url",
        lambda client, y: probe_calls.append((y,)) or "unused",
    )

    result = adapter.check_for_updates("population", None)

    assert result.status == "update_available"
    assert probe_calls == []
